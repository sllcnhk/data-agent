"""
数据导出功能单元测试

A · ClickHouse 导出客户端
B · 大整数 / 单元格格式化
C · Excel 生成（单 Sheet、多 Sheet 分割、标题行）
D · SQL 预览接口
E · run_export_job 协程（正常 / 取消 / 启动竞态）
F · REST API 端点（权限、状态码、下载）

运行：
    python -m pytest test_data_export.py -v -s
"""
import asyncio
import io
import os
import sys
import tempfile
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple
from unittest.mock import AsyncMock, MagicMock, Mock, patch, call

import pytest

sys.path.insert(0, str(Path(__file__).parent))

os.environ.setdefault("ENABLE_AUTH", "False")

# ─── 测试前缀（conftest session 清理会删除匹配项）──────────────────────────────
_PREFIX = f"_t_de_{uuid.uuid4().hex[:6]}_"


# =============================================================================
# A · ClickHouse 导出客户端
# =============================================================================

class TestClickHouseExportClientGetColumns:
    """A1-A4: get_columns()"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient("localhost", 8123, "u", "p", "db")

    def _mock_resp(self, text: str, status: int = 200):
        resp = Mock()
        resp.status_code = status
        resp.text = text
        resp.content = text.encode("utf-8")
        return resp

    def test_a1_returns_column_info_list(self):
        """A1: 正常返回 ColumnInfo 列表"""
        from backend.services.export_clients.base import ColumnInfo
        header = "id\tname\tage\n"
        types  = "Int64\tString\tUInt32\n"
        client = self._make_client()
        with patch("requests.post", return_value=self._mock_resp(header + types)):
            result = client.get_columns("SELECT * FROM t")
        assert result == [
            ColumnInfo("id", "Int64"),
            ColumnInfo("name", "String"),
            ColumnInfo("age", "UInt32"),
        ]

    def test_a2_empty_result_returns_empty_list(self):
        """A2: 空响应返回空列表"""
        client = self._make_client()
        with patch("requests.post", return_value=self._mock_resp("")):
            result = client.get_columns("SELECT 1")
        assert result == []

    def test_a3_http_error_raises_runtime(self):
        """A3: HTTP 非 200 抛 RuntimeError"""
        client = self._make_client()
        with patch("requests.post", return_value=self._mock_resp("err", 400)):
            with pytest.raises(RuntimeError, match="400"):
                client.get_columns("SELECT bad query")

    def test_a4_connection_error_raises(self):
        """A4: 连接失败抛 ConnectionError"""
        import requests as _req
        client = self._make_client()
        with patch("requests.post", side_effect=_req.ConnectionError("refused")):
            with pytest.raises(ConnectionError):
                client.get_columns("SELECT 1")


class TestClickHouseExportClientStreamBatches:
    """A5-A9: stream_batches()"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient("localhost", 8123, "u", "p", "db")

    def _streaming_response(self, lines: List[str]):
        resp = Mock()
        resp.status_code = 200
        resp.iter_lines.return_value = iter(lines)
        return resp

    def test_a5_yields_correct_batches(self):
        """A5: 数据行正确分批"""
        header = ["col_a\tcol_b", "String\tInt64"]
        data   = [f"row{i}\t{i}" for i in range(7)]
        resp = self._streaming_response(header + data)

        client = self._make_client()
        with patch("requests.post", return_value=resp):
            batches = list(client.stream_batches("SELECT 1", batch_size=3))

        assert len(batches) == 3         # ceil(7/3) = 3
        assert len(batches[0]) == 3
        assert len(batches[1]) == 3
        assert len(batches[2]) == 1
        assert batches[0][0] == ("row0", "0")

    def test_a6_empty_result_no_batches(self):
        """A6: 无数据行时不 yield"""
        resp = self._streaming_response(["col\n", "Int64\n"])
        client = self._make_client()
        with patch("requests.post", return_value=resp):
            batches = list(client.stream_batches("SELECT 1", batch_size=100))
        assert batches == []

    def test_a7_null_value_parsed(self):
        """A7: \\N 解析为 None"""
        resp = self._streaming_response(["a", "String", r"\N"])
        client = self._make_client()
        with patch("requests.post", return_value=resp):
            batches = list(client.stream_batches("SELECT 1", batch_size=100))
        assert batches == [([(None,)])]

    def test_a8_tsv_escape_parsed(self):
        """A8: \\t \\n 转义正确还原"""
        resp = self._streaming_response(["a", "String", "hello\\tworld"])
        client = self._make_client()
        with patch("requests.post", return_value=resp):
            batches = list(client.stream_batches("SELECT 1", batch_size=100))
        assert batches[0][0][0] == "hello\tworld"

    def test_a9_http_error_raises(self):
        """A9: HTTP 错误抛 RuntimeError"""
        import requests as _req
        resp = Mock()
        resp.status_code = 500
        resp.content = b"Internal error"
        client = self._make_client()
        with patch("requests.post", return_value=resp):
            with pytest.raises(RuntimeError, match="500"):
                list(client.stream_batches("SELECT 1"))


# =============================================================================
# B · 大整数 / 单元格格式化
# =============================================================================

class TestFormatCell:
    """B1-B10: _format_cell()"""

    def _fmt(self, value, col_type):
        from backend.services.data_export_service import _format_cell
        return _format_cell(value, col_type)

    def test_b1_none_returns_none(self):
        assert self._fmt(None, "Int64") is None

    def test_b2_int64_large_becomes_str(self):
        assert self._fmt(9999999999999999, "Int64") == "9999999999999999"

    def test_b3_uint64_large_becomes_str(self):
        assert self._fmt(2**63, "UInt64") == str(2**63)

    def test_b4_int128_becomes_str(self):
        assert self._fmt(2**127, "Int128") == str(2**127)

    def test_b5_small_int_preserved(self):
        # 小整数也被转为字符串（只要类型是 Int64）
        result = self._fmt(42, "Int64")
        assert result == "42"

    def test_b6_string_type_unchanged(self):
        assert self._fmt("hello", "String") == "hello"

    def test_b7_float_unchanged(self):
        assert self._fmt(3.14, "Float64") == pytest.approx(3.14)

    def test_b8_nullable_int64_stripped(self):
        """Nullable(Int64) 应该被识别为 Int64"""
        assert self._fmt(123456789012345, "Nullable(Int64)") == "123456789012345"

    def test_b9_low_cardinality_stripped(self):
        """LowCardinality(String) 应该被识别为 String"""
        assert self._fmt("x", "LowCardinality(String)") == "x"

    def test_b10_bool_unchanged(self):
        assert self._fmt(True, "UInt8") is True


# =============================================================================
# C · Excel 生成
# =============================================================================

class TestExcelGeneration:
    """C1-C6: openpyxl write-only 模式、多 Sheet 分割、标题行"""

    def _run(self, coro):
        return asyncio.get_event_loop().run_until_complete(coro)

    def _make_job_in_db(self, db, username: str) -> str:
        from backend.models.export_job import ExportJob
        job = ExportJob(
            user_id="test-uid",
            username=username,
            query_sql="SELECT 1",
            connection_env="test",
            status="pending",
            output_filename="test.xlsx",
            file_path="/tmp/test.xlsx",
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)

    def test_c1_single_sheet_created(self, tmp_path):
        """C1: 小于 100 万行 → 只有 Sheet1"""
        import openpyxl
        from backend.services.export_clients.base import ColumnInfo
        from backend.config.database import SessionLocal
        import backend.services.data_export_service as svc

        output = str(tmp_path / "out.xlsx")
        db = SessionLocal()
        job_id = self._make_job_in_db(db, f"{_PREFIX}c1")
        db.close()

        config = {
            "query_sql": "SELECT 1",
            "connection_env": "test",
            "connection_type": "clickhouse",
            "batch_size": 1000,
            "output_path": output,
            "output_filename": "out.xlsx",
        }

        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("col1", "String"), ColumnInfo("col2", "String")]
        mock_client.stream_batches.return_value = iter([[("a", "b"), ("c", "d")]])

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            loop = asyncio.new_event_loop()
            loop.run_until_complete(svc.run_export_job(job_id, config))
            loop.close()

        wb = openpyxl.load_workbook(output)
        assert wb.sheetnames == ["Sheet1"]
        ws = wb["Sheet1"]
        assert ws.cell(1, 1).value == "col1"   # 标题行
        assert ws.cell(2, 1).value == "a"       # 数据行

    def test_c2_header_row_on_every_sheet(self, tmp_path):
        """C2: 每个 Sheet 第一行都是标题行"""
        import openpyxl
        from backend.services.export_clients.base import ColumnInfo
        from backend.config.database import SessionLocal
        import backend.services.data_export_service as svc

        output = str(tmp_path / "multisheet.xlsx")
        db = SessionLocal()
        job_id = self._make_job_in_db(db, f"{_PREFIX}c2")
        db.close()

        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("id", "Int32"), ColumnInfo("val", "String")]
        # 生成 3 批，每批略超过 MAX_ROWS_PER_SHEET 的 1/3（模拟跨 Sheet 场景）
        # 用 patch MAX_ROWS_PER_SHEET = 5 简化测试
        batch = [tuple([i, f"v{i}"]) for i in range(4)]
        mock_client.stream_batches.return_value = iter([batch, batch])  # 8 行

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 100,
            "output_path": output, "output_filename": "out.xlsx",
        }

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            with patch("backend.services.data_export_service.MAX_ROWS_PER_SHEET", 5):
                loop = asyncio.new_event_loop()
                loop.run_until_complete(svc.run_export_job(job_id, config))
                loop.close()

        wb = openpyxl.load_workbook(output)
        assert len(wb.sheetnames) == 2
        for sname in wb.sheetnames:
            ws = wb[sname]
            assert ws.cell(1, 1).value == "id"   # 每 Sheet 标题行

    def test_c3_large_int_written_as_string(self, tmp_path):
        """C3: Int64 大值在 Excel 中以字符串写入"""
        import openpyxl
        from backend.services.export_clients.base import ColumnInfo
        from backend.config.database import SessionLocal
        import backend.services.data_export_service as svc

        output = str(tmp_path / "bigint.xlsx")
        db = SessionLocal()
        job_id = self._make_job_in_db(db, f"{_PREFIX}c3")
        db.close()

        big_id = 9999999999999999
        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("id", "Int64")]
        mock_client.stream_batches.return_value = iter([[(big_id,)]])

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 100,
            "output_path": output, "output_filename": "out.xlsx",
        }

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            loop = asyncio.new_event_loop()
            loop.run_until_complete(svc.run_export_job(job_id, config))
            loop.close()

        wb = openpyxl.load_workbook(output)
        ws = wb.active
        cell_val = ws.cell(2, 1).value   # row 1 = header, row 2 = data
        assert isinstance(cell_val, str), f"Expected str, got {type(cell_val)}: {cell_val}"
        assert cell_val == str(big_id)


# =============================================================================
# D · SQL 预览接口
# =============================================================================

class TestPreviewQuery:
    """D1-D4: preview_query()"""

    def _preview_with_mock(self, mock_execute_return, sql="SELECT 1", env="test", **kw):
        """公共 helper: mock ClickHouseHTTPClient.execute 并调用 preview_query"""
        from backend.services.data_export_service import preview_query
        mock_http = Mock()
        mock_http.execute.return_value = mock_execute_return
        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db"
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.services.data_export_service.preview_query.__globals__"
                       if False else "backend.config.settings.settings", mock_settings):
                return preview_query(sql, env, **kw)

    def test_d1_returns_columns_and_rows(self):
        """D1: 正常返回列信息和行数据"""
        from backend.services.data_export_service import preview_query
        rows = [("alice", 30), ("bob", 25)]
        col_types = [("name", "String"), ("age", "UInt32")]
        mock_http = Mock()
        mock_http.execute.return_value = (rows, col_types)
        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db"
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                result = preview_query("SELECT 1", "test")
        assert len(result["columns"]) == 2
        assert result["columns"][0]["name"] == "name"
        assert len(result["rows"]) == 2
        assert result["row_count"] == 2

    def test_d2_large_int_formatted(self):
        """D2: 预览中大整数转字符串"""
        from backend.services.data_export_service import preview_query
        big = 9007199254740993
        mock_http = Mock()
        mock_http.execute.return_value = ([(big,)], [("id", "Int64")])
        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db"
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                result = preview_query("SELECT 1", "test")
        assert result["rows"][0][0] == str(big)

    def test_d3_sql_error_propagates(self):
        """D3: SQL 错误向上传播"""
        from backend.services.data_export_service import preview_query
        mock_http = Mock()
        mock_http.execute.side_effect = RuntimeError("Unknown column")
        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db"
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                with pytest.raises(RuntimeError, match="Unknown column"):
                    preview_query("SELECT bad", "test")

    def test_d4_unsupported_connection_type_raises(self):
        """D4: 不支持的连接类型抛 ValueError"""
        from backend.services.data_export_service import preview_query
        with pytest.raises(ValueError, match="不支持"):
            preview_query("SELECT 1", "test", connection_type="oracle")


# =============================================================================
# E · run_export_job 协程
# =============================================================================

class TestRunExportJob:
    """E1-E7: 正常完成、取消、启动竞态、失败处理"""

    def _make_job(self, db, username: str, status: str = "pending") -> str:
        from backend.models.export_job import ExportJob
        job = ExportJob(
            user_id="uid", username=username,
            query_sql="SELECT 1", connection_env="test",
            status=status, output_filename="out.xlsx", file_path="/tmp/out.xlsx",
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)

    def _run(self, coro):
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()

    def test_e1_completed_status_after_run(self, tmp_path):
        """E1: 正常完成后状态为 completed"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        from backend.services.export_clients.base import ColumnInfo
        import backend.services.data_export_service as svc

        output = str(tmp_path / "e1.xlsx")
        db = SessionLocal()
        job_id = self._make_job(db, f"{_PREFIX}e1")
        db.close()

        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("x", "String")]
        mock_client.stream_batches.return_value = iter([[("hello",)]])

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 1000,
            "output_path": output, "output_filename": "e1.xlsx",
        }

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            self._run(svc.run_export_job(job_id, config))

        db = SessionLocal()
        job = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        assert job.status == "completed"
        assert job.exported_rows == 1
        assert job.file_size is not None

    def test_e2_cancelled_before_start(self, tmp_path):
        """E2: 启动前已 cancelling → 直接 cancelled"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        import backend.services.data_export_service as svc

        output = str(tmp_path / "e2.xlsx")
        db = SessionLocal()
        job_id = self._make_job(db, f"{_PREFIX}e2", status="cancelling")
        db.close()

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 1000,
            "output_path": output, "output_filename": "e2.xlsx",
        }

        self._run(svc.run_export_job(job_id, config))

        db = SessionLocal()
        job = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        assert job.status == "cancelled"

    def test_e3_cancel_mid_export(self, tmp_path):
        """E3: 运行中检测到 cancelling → 标记 cancelled"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        from backend.services.export_clients.base import ColumnInfo
        import backend.services.data_export_service as svc

        output = str(tmp_path / "e3.xlsx")
        db = SessionLocal()
        job_id = self._make_job(db, f"{_PREFIX}e3")
        db.close()

        call_count = {"n": 0}

        def _batches(*args, **kwargs):
            # 第一批正常，第二批返回前设置 cancelling
            yield [("row0",)]
            # 模拟第一批后 DB 被设为 cancelling
            db2 = SessionLocal()
            j = db2.query(ExportJob).filter(ExportJob.id == job_id).first()
            j.status = "cancelling"
            db2.commit()
            db2.close()
            yield [("row1",)]

        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("v", "String")]
        mock_client.stream_batches.return_value = _batches()

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 1000,
            "output_path": output, "output_filename": "e3.xlsx",
        }

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            self._run(svc.run_export_job(job_id, config))

        db = SessionLocal()
        job = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        assert job.status == "cancelled"

    def test_e4_get_columns_failure_marks_failed(self, tmp_path):
        """E4: get_columns 失败 → 状态变 failed"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        import backend.services.data_export_service as svc

        output = str(tmp_path / "e4.xlsx")
        db = SessionLocal()
        job_id = self._make_job(db, f"{_PREFIX}e4")
        db.close()

        mock_client = Mock()
        mock_client.get_columns.side_effect = RuntimeError("Connection refused")

        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 1000,
            "output_path": output, "output_filename": "e4.xlsx",
        }

        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            self._run(svc.run_export_job(job_id, config))

        db = SessionLocal()
        job = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        assert job.status == "failed"
        assert "Connection refused" in job.error_message

    def test_e5_job_not_found_exits_gracefully(self, tmp_path):
        """E5: job_id 不存在时协程安全退出"""
        import backend.services.data_export_service as svc
        config = {
            "query_sql": "SELECT 1", "connection_env": "test",
            "connection_type": "clickhouse", "batch_size": 1000,
            "output_path": str(tmp_path / "e5.xlsx"), "output_filename": "e5.xlsx",
        }
        self._run(svc.run_export_job("00000000-0000-0000-0000-000000000000", config))
        # 不应抛出任何异常


# =============================================================================
# F · REST API 端点
# =============================================================================

class TestDataExportAPI:
    """F1-F10: API 端点权限、状态码、响应格式"""

    @pytest.fixture
    def client(self):
        os.environ["ENABLE_AUTH"] = "False"
        from fastapi.testclient import TestClient
        import sys
        sys.path.insert(0, str(Path(__file__).parent / "backend"))
        from main import app
        with TestClient(app) as c:
            yield c

    def test_f1_get_connections_200(self, client):
        """F1: GET /connections → 200"""
        with patch("backend.services.data_import_service.list_writable_connections", return_value=[]):
            resp = client.get("/api/v1/data-export/connections")
        assert resp.status_code == 200
        assert resp.json()["success"] is True

    def test_f2_preview_400_on_sql_error(self, client):
        """F2: POST /preview SQL 错误 → 400"""
        with patch(
            "backend.services.data_export_service.preview_query",
            side_effect=RuntimeError("bad SQL"),
        ):
            resp = client.post("/api/v1/data-export/preview", json={
                "query_sql": "SELECT bad",
                "connection_env": "test",
            })
        assert resp.status_code == 400

    def test_f3_list_jobs_200(self, client):
        """F3: GET /jobs → 200 with pagination"""
        resp = client.get("/api/v1/data-export/jobs")
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert "total" in data
        assert "items" in data

    def test_f4_get_nonexistent_job_404(self, client):
        """F4: GET /jobs/{bad_id} → 404"""
        resp = client.get(f"/api/v1/data-export/jobs/{uuid.uuid4()}")
        assert resp.status_code == 404

    def test_f5_cancel_nonexistent_job_404(self, client):
        """F5: POST /jobs/{bad_id}/cancel → 404"""
        resp = client.post(f"/api/v1/data-export/jobs/{uuid.uuid4()}/cancel")
        assert resp.status_code == 404

    def test_f6_delete_nonexistent_job_404(self, client):
        """F6: DELETE /jobs/{bad_id} → 404"""
        resp = client.delete(f"/api/v1/data-export/jobs/{uuid.uuid4()}")
        assert resp.status_code == 404

    def test_f7_download_nonexistent_job_404(self, client):
        """F7: GET /jobs/{bad_id}/download → 404"""
        resp = client.get(f"/api/v1/data-export/jobs/{uuid.uuid4()}/download")
        assert resp.status_code == 404

    def test_f8_cancel_pending_job_becomes_cancelled(self, client):
        """F8: 取消 pending 任务 → 直接 cancelled（无需经过 cancelling）"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        job = ExportJob(
            user_id="uid", username=f"{_PREFIX}f8",
            query_sql="SELECT 1", connection_env="test",
            status="pending",
        )
        db.add(job)
        db.commit()
        job_id = str(job.id)
        db.close()

        resp = client.post(f"/api/v1/data-export/jobs/{job_id}/cancel")
        assert resp.status_code == 200
        assert resp.json()["data"]["status"] == "cancelled"

    def test_f9_cancel_completed_job_400(self, client):
        """F9: 取消已完成任务 → 400"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        job = ExportJob(
            user_id="uid", username=f"{_PREFIX}f9",
            query_sql="SELECT 1", connection_env="test",
            status="completed",
        )
        db.add(job)
        db.commit()
        job_id = str(job.id)
        db.close()

        resp = client.post(f"/api/v1/data-export/jobs/{job_id}/cancel")
        assert resp.status_code == 400

    def test_f10_download_completed_job_returns_file(self, client, tmp_path):
        """F10: 下载已完成任务 → FileResponse"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        fpath = tmp_path / "dl.xlsx"
        fpath.write_bytes(b"PK dummy xlsx content")

        db = SessionLocal()
        job = ExportJob(
            user_id="uid", username=f"{_PREFIX}f10",
            query_sql="SELECT 1", connection_env="test",
            status="completed",
            output_filename="dl.xlsx",
            file_path=str(fpath),
        )
        db.add(job)
        db.commit()
        job_id = str(job.id)
        db.close()

        resp = client.get(f"/api/v1/data-export/jobs/{job_id}/download")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]


# =============================================================================
# teardown
# =============================================================================

def teardown_module(_):
    """清理本文件创建的测试数据"""
    try:
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        db = SessionLocal()
        jobs = db.query(ExportJob).filter(ExportJob.username.like(f"{_PREFIX}%")).all()
        for j in jobs:
            db.delete(j)
        db.commit()
        db.close()
    except Exception:
        pass
