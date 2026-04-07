"""
test_export_chunked.py — 分批提取机制单元测试

覆盖维度：

  A (3)  — is_ch_timeout_estimate_error() 错误检测
           A1: Code 160 消息 → True
           A2: ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED 消息 → True
           A3: 其他错误（Code 241、超时、连接失败）→ False

  B (4)  — count_rows() 行数预扫描
           B1: 正常响应 → 返回正确整数
           B2: HTTP 非 200 → 抛出 RuntimeError
           B3: 请求超时 → 抛出 TimeoutError
           B4: 连接失败 → 抛出 ConnectionError

  C (4)  — stream_batches() extra_settings 注入
           C1: extra_settings 合并进 HTTP params
           C2: max_execution_time=300 出现在请求参数中
           C3: extra_settings=None 不破坏现有行为
           C4: Code 160 响应 → 抛出含 "Code: 160" 的 RuntimeError

  D (5)  — stream_batches_chunked() 分批逻辑
           D1: 整除分批 → 正确窗口数和 LIMIT/OFFSET
           D2: 不整除分批 → 最后一批 limit=remaining
           D3: 合并所有分批结果 = 完整数据集
           D4: total_rows=0 → 不发送任何请求
           D5: extra_settings 传递到每个子查询

  E (3)  — run_export_job() Code 160 自动切换
           E1: 正常流返回数据 → stream 模式，不触发 chunked
           E2: 第一次抛 Code 160 → 自动切换 chunked，最终导出成功
           E3: 非 Code 160 RuntimeError → 标记 failed，不重试

共计: 19 个测试用例

运行：
    /d/ProgramData/Anaconda3/envs/dataagent/python.exe -m pytest test_export_chunked.py -v -s
"""
from __future__ import annotations

import asyncio
import io
import os
import sys
import tempfile
import unittest
import uuid
from pathlib import Path
from unittest.mock import MagicMock, Mock, call, patch

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(1, os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_chk_{uuid.uuid4().hex[:6]}_"

# ═════════════════════════════════════════════════════════════════════════════
# Section A — is_ch_timeout_estimate_error() 检测
# ═════════════════════════════════════════════════════════════════════════════

class TestTimeoutErrorDetection(unittest.TestCase):
    """A1-A3: 正确识别 Code 160，不误判其他错误"""

    def setUp(self):
        from backend.services.export_clients.clickhouse import is_ch_timeout_estimate_error
        self.detect = is_ch_timeout_estimate_error

    def test_A1_code_160_in_message_returns_true(self):
        """A1: 消息含 'Code: 160' → True"""
        err = RuntimeError(
            "ClickHouse 错误 500: Code: 160, e.displayText() = DB::Exception: "
            "Estimated query execution time (60.9 seconds) is too long. Maximum: 60."
        )
        self.assertTrue(self.detect(err), "A1 应识别 Code: 160")

    def test_A2_estimated_execution_timeout_keyword_returns_true(self):
        """A2: 消息含 'ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED' → True"""
        err = RuntimeError("ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED: time budget exceeded")
        self.assertTrue(self.detect(err), "A2 应识别 ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED")

    def test_A3_other_errors_return_false(self):
        """A3: 其他错误类型不被误判"""
        cases = [
            RuntimeError("ClickHouse 错误 500: Code: 241, Memory limit exceeded"),
            TimeoutError("HTTP 请求超时"),
            ConnectionError("ClickHouse 连接失败"),
            RuntimeError("Code: 62, Syntax error in SQL"),
            RuntimeError("Code: 60, Table not found"),
        ]
        for err in cases:
            with self.subTest(err=err):
                self.assertFalse(self.detect(err), f"A3 不应误判: {err}")


# ═════════════════════════════════════════════════════════════════════════════
# Section B — count_rows() 预扫描
# ═════════════════════════════════════════════════════════════════════════════

class TestCountRows(unittest.TestCase):
    """B1-B4: count_rows() 正确处理各种 HTTP 响应"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient(
            host="localhost", port=8123,
            user="default", password="", database="test",
        )

    def test_B1_normal_response_returns_int(self):
        """B1: HTTP 200 + 数字文本 → 返回正确整数"""
        client = self._make_client()
        mock_resp = Mock(status_code=200, text="3901410\n")

        with patch("requests.post", return_value=mock_resp) as mock_post:
            result = client.count_rows("SELECT * FROM t", timeout=60)

        self.assertEqual(result, 3901410, "B1 应返回 3901410")
        # 验证 max_execution_time 被注入
        call_kwargs = mock_post.call_args
        params = call_kwargs[1].get("params", {}) or call_kwargs[0][1] if call_kwargs[0] else {}
        params = mock_post.call_args.kwargs.get("params", mock_post.call_args.args[1] if len(mock_post.call_args.args) > 1 else {})
        # 检查 params 包含 max_execution_time
        all_params = {}
        if mock_post.call_args.kwargs.get("params"):
            all_params = mock_post.call_args.kwargs["params"]
        self.assertIn("max_execution_time", all_params, "B1 应注入 max_execution_time")
        self.assertEqual(all_params["max_execution_time"], 60)

    def test_B2_non_200_raises_runtime_error(self):
        """B2: HTTP 非 200 → 抛出 RuntimeError"""
        client = self._make_client()
        mock_resp = Mock(status_code=500, text="Code: 160, ...")

        with patch("requests.post", return_value=mock_resp):
            with self.assertRaises(RuntimeError, msg="B2 应抛出 RuntimeError"):
                client.count_rows("SELECT * FROM t")

    def test_B3_requests_timeout_raises_timeout_error(self):
        """B3: requests.Timeout → 抛出 TimeoutError"""
        import requests as req_lib
        client = self._make_client()

        with patch("requests.post", side_effect=req_lib.Timeout()):
            with self.assertRaises(TimeoutError, msg="B3 应抛出 TimeoutError"):
                client.count_rows("SELECT * FROM t")

    def test_B4_connection_error_raises_connection_error(self):
        """B4: requests.ConnectionError → 抛出 ConnectionError"""
        import requests as req_lib
        client = self._make_client()

        with patch("requests.post", side_effect=req_lib.ConnectionError()):
            with self.assertRaises(ConnectionError, msg="B4 应抛出 ConnectionError"):
                client.count_rows("SELECT * FROM t")


# ═════════════════════════════════════════════════════════════════════════════
# Section C — stream_batches() extra_settings 注入
# ═════════════════════════════════════════════════════════════════════════════

class TestStreamBatchesExtraSettings(unittest.TestCase):
    """C1-C4: extra_settings 正确合并进 HTTP 参数"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient(
            host="localhost", port=8123,
            user="default", password="", database="test",
        )

    def _mock_tsv_response(self, rows: list[tuple]) -> Mock:
        """生成带表头的 TSV mock 响应"""
        lines = ["col_a\tcol_b", "String\tInt32"]
        for row in rows:
            lines.append("\t".join(str(v) for v in row))
        mock_resp = Mock(status_code=200)
        mock_resp.iter_lines.return_value = iter(lines)
        return mock_resp

    def test_C1_extra_settings_merged_into_params(self):
        """C1: extra_settings 字典合并进 HTTP params"""
        client = self._make_client()
        mock_resp = self._mock_tsv_response([("a", 1)])

        captured_params = {}

        def capture(*args, **kwargs):
            captured_params.update(kwargs.get("params", {}))
            return mock_resp

        with patch("requests.post", side_effect=capture):
            list(client.stream_batches(
                "SELECT 1",
                extra_settings={"max_execution_time": 300, "max_memory_usage": 10000000000},
            ))

        self.assertEqual(captured_params.get("max_execution_time"), 300, "C1 max_execution_time 应为 300")
        self.assertEqual(captured_params.get("max_memory_usage"), 10000000000, "C1 max_memory_usage 应注入")

    def test_C2_max_execution_time_300_in_params(self):
        """C2: max_execution_time=300 出现在请求参数中"""
        client = self._make_client()
        mock_resp = self._mock_tsv_response([])

        captured_params = {}
        with patch("requests.post", side_effect=lambda *a, **kw: (captured_params.update(kw.get("params", {})) or mock_resp)):
            list(client.stream_batches("SELECT 1", extra_settings={"max_execution_time": 300}))

        self.assertIn("max_execution_time", captured_params)
        self.assertEqual(captured_params["max_execution_time"], 300)

    def test_C3_no_extra_settings_does_not_break(self):
        """C3: extra_settings=None 时行为与原来一致"""
        client = self._make_client()
        mock_resp = self._mock_tsv_response([("x", 42)])

        with patch("requests.post", return_value=mock_resp):
            batches = list(client.stream_batches("SELECT 1", extra_settings=None))

        self.assertEqual(len(batches), 1)
        self.assertEqual(batches[0], [("x", "42")])

    def test_C4_code_160_response_raises_runtime_error_with_code(self):
        """C4: ClickHouse 返回 Code 160 文本 → RuntimeError 含 'Code: 160'"""
        from backend.services.export_clients.clickhouse import is_ch_timeout_estimate_error
        client = self._make_client()

        code160_body = (
            "Code: 160, e.displayText() = DB::Exception: Estimated query execution time "
            "(60.926 seconds) is too long. Maximum: 60. Estimated rows to process: 3901410"
        )
        mock_resp = Mock(status_code=500, content=code160_body.encode())

        with patch("requests.post", return_value=mock_resp):
            with self.assertRaises(RuntimeError) as ctx:
                list(client.stream_batches("SELECT * FROM huge_table"))

        err = ctx.exception
        self.assertTrue(
            is_ch_timeout_estimate_error(err),
            f"C4 抛出的 RuntimeError 应被识别为 Code 160，实际: {err}",
        )


# ═════════════════════════════════════════════════════════════════════════════
# Section D — stream_batches_chunked() 分批逻辑
# ═════════════════════════════════════════════════════════════════════════════

class TestStreamBatchesChunked(unittest.TestCase):
    """D1-D5: LIMIT/OFFSET 窗口分批正确性"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient(
            host="localhost", port=8123,
            user="default", password="", database="test",
        )

    def test_D1_exact_division_generates_correct_windows(self):
        """D1: total_rows=400 chunk_size=200 → 2 个窗口，LIMIT/OFFSET 正确"""
        client = self._make_client()
        emitted_sqls = []

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            emitted_sqls.append(sql)
            # 返回空批次（只测窗口逻辑）
            return iter([])

        client.stream_batches = fake_stream_batches

        list(client.stream_batches_chunked("SELECT * FROM t", chunk_size=200, total_rows=400))

        self.assertEqual(len(emitted_sqls), 2, "D1 应生成 2 个 SQL 窗口")
        self.assertIn("LIMIT 200 OFFSET 0", emitted_sqls[0])
        self.assertIn("LIMIT 200 OFFSET 200", emitted_sqls[1])

    def test_D2_non_divisible_last_chunk_uses_remaining(self):
        """D2: total_rows=350 chunk_size=200 → 最后一批 LIMIT=150"""
        client = self._make_client()
        emitted_sqls = []

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            emitted_sqls.append(sql)
            return iter([])

        client.stream_batches = fake_stream_batches

        list(client.stream_batches_chunked("SELECT * FROM t", chunk_size=200, total_rows=350))

        self.assertEqual(len(emitted_sqls), 2, "D2 应生成 2 个窗口")
        self.assertIn("LIMIT 200 OFFSET 0", emitted_sqls[0])
        self.assertIn("LIMIT 150 OFFSET 200", emitted_sqls[1], "D2 最后批应为 150 行")

    def test_D3_all_chunks_combined_equal_full_dataset(self):
        """D3: 合并所有分批的行 = 完整数据集（顺序正确）"""
        client = self._make_client()

        # 模拟 5 行数据，chunk_size=2 → 3 个窗口
        all_data = [(i, f"v{i}") for i in range(5)]

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            # 从 SQL 提取 LIMIT 和 OFFSET
            import re
            m = re.search(r"LIMIT (\d+) OFFSET (\d+)", sql)
            if m:
                limit = int(m.group(1))
                offset = int(m.group(2))
                yield all_data[offset:offset + limit]

        client.stream_batches = fake_stream_batches

        result = []
        for batch in client.stream_batches_chunked("SELECT * FROM t", chunk_size=2, total_rows=5):
            result.extend(batch)

        self.assertEqual(result, all_data, f"D3 合并结果应等于原始数据，实际: {result}")

    def test_D4_total_rows_zero_yields_nothing(self):
        """D4: total_rows=0 → 不发送任何请求，不 yield 任何批次"""
        client = self._make_client()
        call_count = [0]

        def fake_stream_batches(sql, **kwargs):
            call_count[0] += 1
            return iter([])

        client.stream_batches = fake_stream_batches

        result = list(client.stream_batches_chunked("SELECT * FROM t", chunk_size=100, total_rows=0))

        self.assertEqual(call_count[0], 0, "D4 total_rows=0 不应发起请求")
        self.assertEqual(result, [], "D4 不应返回数据")

    def test_D5_extra_settings_passed_to_each_chunk_query(self):
        """D5: extra_settings 传递到每个子查询的 stream_batches 调用"""
        client = self._make_client()
        received_settings = []

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            received_settings.append(extra_settings)
            return iter([])

        client.stream_batches = fake_stream_batches

        list(client.stream_batches_chunked(
            "SELECT * FROM t",
            chunk_size=100,
            total_rows=250,
            extra_settings={"max_execution_time": 300},
        ))

        self.assertEqual(len(received_settings), 3, "D5 应有 3 次子查询")
        for i, s in enumerate(received_settings):
            self.assertEqual(
                s, {"max_execution_time": 300},
                f"D5 第 {i} 个子查询的 extra_settings 不正确: {s}",
            )


# ═════════════════════════════════════════════════════════════════════════════
# Section E — run_export_job() Code 160 自动切换端到端
# ═════════════════════════════════════════════════════════════════════════════

class TestRunExportJobChunkedFallback(unittest.TestCase):
    """E1-E3: run_export_job 在 Code 160 时自动切换分批，非 Code 160 正常失败"""

    @classmethod
    def setUpClass(cls):
        from backend.config.database import SessionLocal
        cls.db = SessionLocal()
        cls.tmp_dir = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        from backend.models.export_job import ExportJob
        try:
            cls.db.query(ExportJob).filter(
                ExportJob.username.like(f"{_PREFIX}%")
            ).delete(synchronize_session=False)
            cls.db.commit()
        finally:
            cls.db.close()

    def _make_job(self, suffix=""):
        from backend.models.export_job import ExportJob
        job = ExportJob(
            user_id="test-uid",
            username=f"{_PREFIX}{suffix}",
            query_sql="SELECT 1",
            connection_env="test",
            connection_type="clickhouse",
            status="pending",
            output_filename=f"{_PREFIX}{suffix}.xlsx",
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def _run(self, job_id, output_path, stream_side_effect):
        """
        运行 run_export_job，mock ClickHouseExportClient。
        stream_side_effect: list of return values/exceptions for stream_batches calls.
        """
        import asyncio
        from backend.services.data_export_service import run_export_job
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        # 3 行假数据，2 列
        fake_columns_info = [
            type("C", (), {"name": "id", "type": "Int32"})(),
            type("C", (), {"name": "val", "type": "String"})(),
        ]
        fake_rows = [(1, "a"), (2, "b"), (3, "c")]

        call_idx = [0]

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            idx = call_idx[0]
            call_idx[0] += 1
            effect = stream_side_effect[idx] if idx < len(stream_side_effect) else iter([])

            # 返回延迟生成器，匹配真实 stream_batches 的懒执行特性：
            # 真实代码是生成器函数（yield），调用时不执行，迭代时才执行。
            # 若这里直接 raise，异常会在赋值行（batch_source = ...）而非迭代行触发，
            # 导致 try/except 包裹的 for 循环无法捕获。
            def _gen():
                if isinstance(effect, Exception):
                    raise effect
                yield from effect

            return _gen()

        def fake_count_rows(sql, timeout=300):
            return 3

        def fake_stream_batches_chunked(sql, chunk_size, total_rows, batch_size=50000, extra_settings=None):
            yield fake_rows

        with patch.object(ClickHouseExportClient, "get_columns", return_value=fake_columns_info), \
             patch.object(ClickHouseExportClient, "stream_batches", side_effect=fake_stream_batches), \
             patch.object(ClickHouseExportClient, "count_rows", side_effect=fake_count_rows), \
             patch.object(ClickHouseExportClient, "stream_batches_chunked", side_effect=fake_stream_batches_chunked):

            config = {
                "query_sql": "SELECT * FROM t",
                "connection_env": "test",
                "connection_type": "clickhouse",
                "batch_size": 50000,
                "output_path": output_path,
                "output_filename": "test.xlsx",
            }

            with patch("backend.services.data_export_service._build_export_client") as mock_build:
                mock_build.return_value = ClickHouseExportClient(
                    "localhost", 8123, "default", "", "test"
                )
                asyncio.run(run_export_job(job_id, config))

    def test_E1_normal_stream_no_chunked_mode(self):
        """E1: 正常流数据返回 → stream 模式成功，不触发 chunked"""
        job = self._make_job("e1")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}e1.xlsx")

        fake_rows = [(1, "a"), (2, "b")]
        self._run(str(job.id), output_path, stream_side_effect=[iter([fake_rows])])

        self.db.refresh(job)
        self.assertEqual(job.status, "completed", f"E1 任务应 completed，实际: {job.status} {job.error_message}")
        self.assertTrue(Path(output_path).exists(), "E1 Excel 文件应存在")

        # 验证内容
        import openpyxl
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        self.assertEqual(ws.max_row, 3, f"E1 应有 1 表头 + 2 数据行，实际: {ws.max_row}")

    def test_E2_code160_triggers_chunked_retry_and_succeeds(self):
        """E2: 第一次 stream_batches 抛 Code 160 → 自动切换 chunked → 最终 completed"""
        job = self._make_job("e2")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}e2.xlsx")

        code160_err = RuntimeError(
            "ClickHouse 错误 500: Code: 160, e.displayText() = DB::Exception: "
            "Estimated query execution time (60.9 seconds) is too long. Maximum: 60."
        )

        # 第一次调用抛 Code 160，chunked_mode 下由 stream_batches_chunked 接管
        self._run(str(job.id), output_path, stream_side_effect=[code160_err])

        self.db.refresh(job)
        self.assertEqual(
            job.status, "completed",
            f"E2 Code 160 后应切换分批并完成，实际: {job.status} {job.error_message}",
        )
        self.assertTrue(Path(output_path).exists(), "E2 Excel 文件应存在")

    def test_E3_non_code160_error_marks_failed_no_retry(self):
        """E3: 非 Code 160 RuntimeError → 标记 failed，不触发 chunked 重试"""
        job = self._make_job("e3")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}e3.xlsx")

        other_err = RuntimeError("ClickHouse 错误 500: Code: 241, Memory limit exceeded")
        self._run(str(job.id), output_path, stream_side_effect=[other_err])

        self.db.refresh(job)
        self.assertEqual(job.status, "failed", f"E3 应标记 failed，实际: {job.status}")
        self.assertIn("241", job.error_message or "", "E3 错误信息应含 Code 241")
        self.assertFalse(Path(output_path).exists(), "E3 不完整文件应被清理")


if __name__ == "__main__":
    unittest.main(verbosity=2)
