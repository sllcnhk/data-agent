"""
数据导出服务

实现 SQL → Excel 的核心业务逻辑：
- SQL 预览查询（前 N 行）
- 流式导出执行（HTTP 流 + openpyxl write-only）
- 多 Sheet 自动分割（每 100 万行新建一个 Sheet）
- 大整数安全转字符串（避免 Excel 科学计数法）
- 批次级取消检查（协作式取消）
- 分批提取模式（规避 ClickHouse max_execution_time 估算拒绝 Code 160）
- 按日期分块多文件导出（v2.13）

线程安全说明：
  run_export_job 是 async 包装器，将同步阻塞工作交给线程池（run_in_executor）。
  这样 ClickHouse HTTP 流式读取（requests 同步库）和 openpyxl 写入不会阻塞 asyncio
  事件循环，使其他 API 请求（如任务列表轮询 GET /data-export/jobs）能正常响应。
"""
import asyncio
import logging
import math
import os
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# 导出专用线程池（最多 4 个并发导出，超出则排队）
_EXPORT_EXECUTOR = ThreadPoolExecutor(max_workers=4, thread_name_prefix="export-worker")

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# 常量
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_BATCH_SIZE = 50_000          # 每批行数
PREVIEW_LIMIT = 100                  # 预览行数
MAX_ROWS_PER_SHEET = 1_000_000       # 每 Sheet 最大数据行（不含表头）
PROGRESS_UPDATE_EVERY = 10           # 每 N 批更新一次 DB 进度

# ClickHouse 大整数类型（超过 JS Number.MAX_SAFE_INTEGER = 2^53-1，需转字符串）
_LARGE_INT_TYPES = frozenset([
    "Int64", "UInt64",
    "Int128", "UInt128",
    "Int256", "UInt256",
])

# ─────────────────────────────────────────────────────────────────────────────
# 辅助：构建导出客户端
# ─────────────────────────────────────────────────────────────────────────────

def _build_export_client(env: str, connection_type: str = "clickhouse"):
    """
    工厂函数：根据 env 和 connection_type 构建对应的导出客户端。
    扩展点：新增 connection_type 时在此处添加分支。
    """
    if connection_type == "clickhouse":
        from backend.config.settings import settings
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        cfg = settings.get_clickhouse_config(env, level="admin")
        return ClickHouseExportClient(
            host=cfg["host"],
            port=cfg["http_port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            timeout=3600,
        )
    raise ValueError(f"不支持的连接类型: {connection_type}")


# ─────────────────────────────────────────────────────────────────────────────
# 辅助：单元格值格式化
# ─────────────────────────────────────────────────────────────────────────────

def _format_cell(value: Any, col_type: str) -> Any:
    """
    将数据库原始值转为 Excel 安全值：
    - 大整数类型（Int64/UInt64/...） → str，避免 Excel 科学计数法
    - None → None（openpyxl 写入为空单元格）
    - 其余保持原值
    """
    if value is None:
        return None

    # 去掉 Nullable(...) 包装，如 "Nullable(Int64)" → "Int64"
    bare_type = col_type
    if bare_type.startswith("Nullable(") and bare_type.endswith(")"):
        bare_type = bare_type[9:-1]
    # 去掉 LowCardinality(...) 包装
    if bare_type.startswith("LowCardinality(") and bare_type.endswith(")"):
        bare_type = bare_type[15:-1]

    if bare_type in _LARGE_INT_TYPES and isinstance(value, int):
        return str(value)

    return value


# ─────────────────────────────────────────────────────────────────────────────
# 1. SQL 预览
# ─────────────────────────────────────────────────────────────────────────────

def preview_query(
    sql: str,
    env: str,
    connection_type: str = "clickhouse",
    limit: int = PREVIEW_LIMIT,
    preview_date: Optional[str] = None,
) -> Dict[str, Any]:
    """
    执行 SQL（加 LIMIT），返回列信息和前 N 行数据，用于前端展示预览。

    占位符处理（v2.13）：
      若 SQL 含 {{date_start}} / {{date_end}} 占位符（用于按日期分块导出），
      preview_query 会用 preview_date（或默认昨日）自动替换，让预览能正常执行。
      这样用户在写好分块 SQL 后无需手动修改即可预览验证。

    Args:
        preview_date: ISO YYYY-MM-DD 字符串；占位符替换的样本日期。
                      None 时默认昨日（避免今日数据可能尚未写入）。
                      仅当 SQL 含占位符时生效。

    Returns:
        {
            "columns": [{"name": str, "type": str}, ...],
            "rows": [[cell, ...], ...],
            "row_count": int,
            "preview_date": str | None,    # 实际使用的样本日期（占位符模式时非空）
        }
    """
    from backend.services.data_export_chunker import (
        has_placeholders, has_partial_placeholders,
    )

    # 防御：单占位符（仅 {{date_start}} 或仅 {{date_end}}）会让 ClickHouse 误把
    # 字面量当作日期解析，必须拒绝
    if has_partial_placeholders(sql):
        raise ValueError(
            "SQL 中 {{date_start}} 与 {{date_end}} 必须成对出现；"
            "仅写一个会导致预览失败"
        )

    actual_preview_date: Optional[str] = None
    if has_placeholders(sql):
        if preview_date:
            try:
                d = datetime.strptime(preview_date, "%Y-%m-%d").date()
            except ValueError as exc:
                raise ValueError(
                    f"preview_date 格式必须为 YYYY-MM-DD，收到 {preview_date!r}"
                ) from exc
        else:
            from datetime import timedelta
            d = (datetime.utcnow().date()) - timedelta(days=1)
        ds = d.isoformat()
        sql = sql.replace("{{date_start}}", ds).replace("{{date_end}}", ds)
        actual_preview_date = ds
        logger.info(
            "[preview_query] Substituted placeholders with sample date %s "
            "(provided=%s)", ds, bool(preview_date),
        )

    if connection_type == "clickhouse":
        from backend.config.settings import settings
        cfg = settings.get_clickhouse_config(env, level="admin")
        from backend.mcp.clickhouse.http_client import ClickHouseHTTPClient
        client = ClickHouseHTTPClient(
            host=cfg["host"],
            port=cfg["http_port"],
            user=cfg["user"],
            password=cfg["password"],
            database=cfg["database"],
            timeout=30,
        )
        stripped = sql.rstrip().rstrip(";")
        preview_sql = f"SELECT * FROM ({stripped}) AS _preview LIMIT {limit}"
        rows, col_types = client.execute(preview_sql, with_column_types=True)
        columns = [{"name": name, "type": tp} for name, tp in col_types]
        # 格式化单元格（大整数转字符串）
        formatted_rows = [
            [_format_cell(v, col_types[i][1]) for i, v in enumerate(row)]
            for row in rows
        ]
        return {
            "columns": columns,
            "rows": formatted_rows,
            "row_count": len(formatted_rows),
            "preview_date": actual_preview_date,
        }
    raise ValueError(f"不支持的连接类型: {connection_type}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. 单文件导出 — 内核（被单文件路径与分块路径共用）
# ─────────────────────────────────────────────────────────────────────────────

def _humanize_error(exc: Exception) -> str:
    """
    把技术异常翻译为面向用户的可读提示，附可能原因 + 建议处置。
    用于 ExportJob.error_message — 前端直接展示给最终用户。

    保留原始 exception 字符串作为「技术细节」段，便于排查。
    """
    raw = str(exc)
    # 流式断开 — 最常见的瞬时错误
    if any(fp in raw for fp in (
        "Connection broken", "IncompleteRead", "ProtocolError",
        "ChunkedEncodingError", "Connection aborted", "Connection reset",
        "Response ended prematurely", "Read timed out",
        "Remote end closed connection",
    )):
        return (
            "ClickHouse 数据流中途断开（已自动尝试 LIMIT/OFFSET 回退仍未成功）。"
            "可能原因：① 云上 LB / NAT / 反向代理空闲连接超时（最常见，~5 分钟切断）；"
            "② 跨境网络抖动；③ ClickHouse 服务端 OOM 或主动 abort 查询。"
            "已自动注入服务端心跳设置（send_progress_in_http_headers=1, "
            "http_headers_progress_interval_ms=10000）；若仍失败，建议："
            "① 减小单块天数（如 chunk_days=2~3）让单次查询时长 < 5 分钟；"
            "② 简化 SQL（减少每行 CPU 消耗，如 decrypt/JSONExtract）；"
            "③ 错峰重试；④ 联系 DBA 检查 ClickHouse 实例资源水位 / 调整云 LB 空闲超时。\n"
            f"[技术细节] {raw}"
        )
    # Code 160 — 估算超时
    if "Code: 160" in raw or "ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED" in raw:
        return (
            "ClickHouse 估算查询执行时间超出 max_execution_time 限制（已自动 "
            "LIMIT/OFFSET 回退仍未恢复）。建议：① 减小单块天数；② 在数据库层"
            "增加索引/分区；③ 调高 EXPORT_QUERY_MAX_EXECUTION_TIME 环境变量。\n"
            f"[技术细节] {raw}"
        )
    # 网络连接失败 — 配置或环境问题
    if "Connection refused" in raw or "Failed to establish" in raw:
        return (
            f"无法连接到 ClickHouse 服务器。请检查连接配置（host/port）、网络"
            f"可达性或服务状态。\n[技术细节] {raw}"
        )
    # 权限/认证类
    if "Authentication failed" in raw or "Code: 192" in raw:
        return f"ClickHouse 认证失败 — 请检查用户名密码配置。\n[技术细节] {raw}"
    # 表不存在
    if "Code: 60" in raw or "doesn't exist" in raw:
        return f"SQL 引用的表/视图不存在。\n[技术细节] {raw}"
    # 兜底
    return raw


def _is_cancelling(job_id: str) -> bool:
    from backend.config.database import SessionLocal
    from backend.models.export_job import ExportJob
    db = SessionLocal()
    try:
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        return j is not None and j.status == "cancelling"
    finally:
        db.close()


def _update_job(job_id: str, **fields) -> None:
    """通用 ExportJob 字段更新（自动 set updated_at + commit）"""
    from backend.config.database import SessionLocal
    from backend.models.export_job import ExportJob
    db = SessionLocal()
    try:
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        if j is None:
            return
        for k, v in fields.items():
            setattr(j, k, v)
        j.updated_at = datetime.utcnow()
        db.commit()
    finally:
        db.close()


def _run_single_export(
    job_id: str,
    sql: str,
    env: str,
    conn_type: str,
    batch_size: int,
    output_path: str,
    *,
    sheet_prefix: str = "Sheet",
    progress_offset: int = 0,
    progress_total: Optional[int] = None,
    chunk_label: Optional[str] = None,
    on_cancel: Optional[Any] = None,
) -> Dict[str, Any]:
    """
    执行单个 SQL 的 ClickHouse → Excel 流式导出。
    被单文件模式与分块模式共用。

    返回结果摘要：
        {
            "exported_rows": int,
            "total_sheets": int,
            "file_size": int,
            "cancelled": bool,
        }

    分块模式下使用 progress_offset / progress_total 让进度条以全 Job 维度更新。
    若发生 ClickHouse Code 160 错误，自动切换到 LIMIT/OFFSET 分批模式重试一次。
    """
    import openpyxl
    from backend.config.settings import settings as app_settings
    from backend.services.export_clients.clickhouse import (
        is_ch_timeout_estimate_error,
        is_transient_stream_error,
    )

    export_settings = {"max_execution_time": app_settings.export_query_max_execution_time}

    # 列信息
    export_client = _build_export_client(env, conn_type)
    columns = export_client.get_columns(sql)
    col_names = [c.name for c in columns]
    col_types = [c.type for c in columns]

    use_chunked = False
    total_sql_chunks = 1

    for attempt in range(2):
        if attempt == 1 and not use_chunked:
            break

        # 初始化工作簿
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook(write_only=True)
        sheet_num = 1
        ws = wb.create_sheet(f"{sheet_prefix}{sheet_num}")
        ws.append(col_names)

        exported_rows = 0
        done_batches = 0
        sheet_row_count = 0

        # 分块模式下 current_sheet 形如「块 N/M (...) - Sheet1」让前端能看到块进度；
        # 单文件模式下保持「Sheet1」原样。
        def _format_sheet_label(n: int) -> str:
            base = f"{sheet_prefix}{n}"
            return f"{chunk_label} - {base}" if chunk_label else base

        _update_job(job_id, current_sheet=_format_sheet_label(sheet_num))

        if use_chunked:
            try:
                total_rows = export_client.count_rows(
                    sql, timeout=app_settings.export_query_max_execution_time,
                )
            except Exception as cnt_err:
                raise RuntimeError(f"分批模式预扫描行数失败: {cnt_err}") from cnt_err

            total_sql_chunks = max(1, math.ceil(total_rows / app_settings.export_chunk_size))
            logger.info(
                "[ExportJob %s] Chunked mode: total_rows=%d → %d SQL chunks (chunk_size=%d)",
                job_id, total_rows, total_sql_chunks, app_settings.export_chunk_size,
            )
            # 分批模式下保留 progress_total 优先级（分块导出已设置全 Job 进度分母）
            update_kw: Dict[str, Any] = {}
            if progress_total is None:
                update_kw["total_rows"] = total_rows
                update_kw["total_batches"] = total_sql_chunks
            if update_kw:
                _update_job(job_id, **update_kw)

            batch_source = export_client.stream_batches_chunked(
                sql,
                chunk_size=app_settings.export_chunk_size,
                total_rows=total_rows,
                batch_size=batch_size,
                extra_settings=export_settings,
            )
        else:
            batch_source = export_client.stream_batches(
                sql, batch_size=batch_size, extra_settings=export_settings,
            )

        try:
            for batch in batch_source:
                # 检查取消
                cancelled = (on_cancel and on_cancel()) or _is_cancelling(job_id)
                if cancelled:
                    logger.info(
                        "[ExportJob %s] Cancelled mid-export after %d rows.",
                        job_id, exported_rows,
                    )
                    try:
                        wb.save(output_path)
                    except Exception:
                        pass
                    return {
                        "exported_rows": exported_rows,
                        "total_sheets": sheet_num,
                        "done_batches": done_batches,
                        "total_sql_chunks": total_sql_chunks if use_chunked else None,
                        "file_size": (
                            Path(output_path).stat().st_size
                            if Path(output_path).exists() else 0
                        ),
                        "cancelled": True,
                    }

                for row in batch:
                    if sheet_row_count >= MAX_ROWS_PER_SHEET:
                        sheet_num += 1
                        ws = wb.create_sheet(f"{sheet_prefix}{sheet_num}")
                        ws.append(col_names)
                        sheet_row_count = 0
                        _update_job(job_id, current_sheet=_format_sheet_label(sheet_num))

                    formatted = [_format_cell(v, col_types[i]) for i, v in enumerate(row)]
                    ws.append(formatted)
                    sheet_row_count += 1
                    exported_rows += 1

                done_batches += 1

                if done_batches % PROGRESS_UPDATE_EVERY == 0:
                    if progress_total is None:
                        # 单文件模式：本函数全权管理进度字段
                        if use_chunked:
                            _update_job(
                                job_id,
                                exported_rows=exported_rows,
                                done_batches=exported_rows // app_settings.export_chunk_size,
                                total_batches=total_sql_chunks,
                            )
                        else:
                            _update_job(
                                job_id,
                                exported_rows=exported_rows,
                                done_batches=done_batches,
                            )
                    else:
                        # 分块模式由父循环更新 done_batches；此处仅累加 exported_rows
                        _update_job(
                            job_id,
                            exported_rows=progress_offset + exported_rows,
                        )

            wb.save(output_path)
            file_size = Path(output_path).stat().st_size
            logger.info(
                "[ExportJob %s] Single-export OK: %d rows, %d sheet(s), %.1f MB, "
                "mode=%s chunks=%d",
                job_id, exported_rows, sheet_num, file_size / 1024 / 1024,
                "chunked" if use_chunked else "stream", total_sql_chunks,
            )
            return {
                "exported_rows": exported_rows,
                "total_sheets": sheet_num,
                "done_batches": done_batches,
                "total_sql_chunks": total_sql_chunks if use_chunked else None,
                "file_size": file_size,
                "cancelled": False,
            }

        except Exception as exc:
            # 触发自动 LIMIT/OFFSET 回退的错误类型：
            #   1. Code 160（估算超时）— ClickHouse 拒绝执行
            #   2. 流式响应中途断开 — 服务端/网络/代理切断长连接
            # 回退策略：每个窗口是独立 HTTP 请求，短小独立连接，重试大概率成功
            should_retry = attempt == 0 and (
                is_ch_timeout_estimate_error(exc)
                or is_transient_stream_error(exc)
            )
            if should_retry:
                reason = (
                    "Code 160 (estimated timeout)"
                    if is_ch_timeout_estimate_error(exc)
                    else "transient stream disconnect"
                )
                logger.warning(
                    "[ExportJob %s] %s — switching to LIMIT/OFFSET fallback mode: %s",
                    job_id, reason, exc,
                )
                use_chunked = True
                try:
                    wb.close()
                except Exception:
                    pass
                continue
            raise

    # 不应到达此处
    raise RuntimeError("_run_single_export 未能完成（未知状态）")


# ─────────────────────────────────────────────────────────────────────────────
# 3. 调度器：根据 export_mode 路由到单文件 / 分块
# ─────────────────────────────────────────────────────────────────────────────

async def run_export_job(job_id: str, config: Dict[str, Any]) -> None:
    """
    后台协程包装器：将同步阻塞的导出工作提交到线程池执行。

    config 结构：
    {
        "query_sql": str,
        "connection_env": str,
        "connection_type": str,         # 默认 "clickhouse"
        "batch_size": int,              # 默认 50_000
        "output_path": str,             # 单文件模式：输出文件绝对路径
        "output_filename": str,         # 单文件模式：文件名（展示用）
        # 分块模式（当存在 chunk_config 时启用）
        "export_mode": "single" | "date_chunked",
        "chunk_config": {
            "date_column": str | None,
            "date_start": "YYYY-MM-DD",
            "date_end": "YYYY-MM-DD",
            "chunk_days": int,
            "mode": "placeholder" | "wrapper",
        },
        "output_dir": str,              # 分块模式：输出目录绝对路径
        "job_name": str,                # 分块模式：用于生成子文件名
    }
    """
    loop = asyncio.get_event_loop()
    mode = config.get("export_mode", "single")
    if mode == "date_chunked":
        await loop.run_in_executor(_EXPORT_EXECUTOR, _run_chunked_export_sync, job_id, config)
    else:
        await loop.run_in_executor(_EXPORT_EXECUTOR, _run_single_job_sync, job_id, config)


def _mark_running(job_id: str) -> bool:
    """
    把 Job 从 pending 标记为 running（同时检查启动竞态：cancelling → 直接 cancelled）。
    返回 True 表示可继续执行；False 表示已终止（cancelled 或 not found）。
    """
    from backend.config.database import SessionLocal
    from backend.models.export_job import ExportJob
    db = SessionLocal()
    try:
        job = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        if not job:
            logger.error("[ExportJob %s] Job not found, aborting.", job_id)
            return False
        if job.status == "cancelling":
            job.status = "cancelled"
            job.finished_at = datetime.utcnow()
            job.updated_at = datetime.utcnow()
            db.commit()
            logger.info("[ExportJob %s] Cancelled before start.", job_id)
            return False
        job.status = "running"
        job.started_at = datetime.utcnow()
        job.updated_at = datetime.utcnow()
        db.commit()
        return True
    finally:
        db.close()


def _mark_failed(job_id: str, msg: str, exported: int = 0, done_b: int = 0) -> None:
    _update_job(
        job_id,
        status="failed",
        finished_at=datetime.utcnow(),
        error_message=msg,
        exported_rows=exported,
        done_batches=done_b,
    )


def _mark_cancelled(job_id: str, exported: int = 0, done_b: int = 0, sheets: int = 0) -> None:
    _update_job(
        job_id,
        status="cancelled",
        finished_at=datetime.utcnow(),
        exported_rows=exported,
        done_batches=done_b,
        total_sheets=sheets,
    )


def _mark_completed(
    job_id: str,
    exported: int,
    done_b: Optional[int],
    total_b: Optional[int],
    sheets: int,
    file_size: int,
) -> None:
    """
    标记 Job 完成。

    done_b/total_b 为 None 时不覆盖 — 用于单文件模式保留 inner 函数已写入的
    Python 批次计数（旧契约：done_batches=Python 批次数）。分块模式传具体值。
    """
    fields: Dict[str, Any] = {
        "status": "completed",
        "finished_at": datetime.utcnow(),
        "exported_rows": exported,
        "total_sheets": sheets,
        "file_size": file_size,
    }
    if done_b is not None:
        fields["done_batches"] = done_b
    if total_b is not None:
        fields["total_batches"] = total_b
    _update_job(job_id, **fields)


# ─────────────────────────────────────────────────────────────────────────────
# 4. 单文件模式协程
# ─────────────────────────────────────────────────────────────────────────────

def _run_single_job_sync(job_id: str, config: Dict[str, Any]) -> None:
    """单文件模式同步实现（线程池中运行）"""
    sql = config["query_sql"]
    env = config["connection_env"]
    conn_type = config.get("connection_type", "clickhouse")
    batch_size = config.get("batch_size", DEFAULT_BATCH_SIZE)
    output_path = config["output_path"]

    if not _mark_running(job_id):
        return

    # 列信息预检（失败 → failed，不进入 _run_single_export 内部）
    try:
        client = _build_export_client(env, conn_type)
        client.get_columns(sql)
    except Exception as e:
        msg = f"获取列信息失败: {_humanize_error(e)}"
        logger.error("[ExportJob %s] %s", job_id, msg)
        _mark_failed(job_id, msg)
        return

    try:
        result = _run_single_export(
            job_id=job_id, sql=sql, env=env, conn_type=conn_type,
            batch_size=batch_size, output_path=output_path,
        )
    except Exception as exc:
        msg = f"导出执行失败：{_humanize_error(exc)}"
        logger.error("[ExportJob %s] %s", job_id, msg, exc_info=True)
        _mark_failed(job_id, msg)
        try:
            os.unlink(output_path)
        except Exception:
            pass
        return

    if result["cancelled"]:
        _mark_cancelled(
            job_id,
            exported=result["exported_rows"],
            done_b=result.get("done_batches", 0),
            sheets=result["total_sheets"],
        )
        return

    # 单文件模式：done_batches=Python 批次数（旧契约）
    # 若发生 Code 160 回退，total_batches=SQL 分批数；否则保持 None
    _mark_completed(
        job_id,
        exported=result["exported_rows"],
        done_b=result.get("done_batches", 0),
        total_b=result.get("total_sql_chunks"),
        sheets=result["total_sheets"],
        file_size=result["file_size"],
    )


# ─────────────────────────────────────────────────────────────────────────────
# 5. 分块模式协程
# ─────────────────────────────────────────────────────────────────────────────

def _run_chunked_export_sync(job_id: str, config: Dict[str, Any]) -> None:
    """
    按日期分块导出 — 同步实现（线程池中运行）。

    流程：
      1. _mark_running（启动竞态检查）
      2. validate_chunk_config 校验
      3. split_date_range 切分块
      4. 逐块构建 SQL → 调用 _run_single_export → 累加进度 + output_files 清单
      5. 任一块失败 → 整体 failed（已完成块文件保留，便于排查）
      6. 中途 cancelling → 当前块写到一半保留，后续块跳过 → cancelled
    """
    from datetime import date as _date

    from backend.services.data_export_chunker import (
        build_chunk_filename,
        inject_date_filter,
        split_date_range,
        subdivide_date_range,
        validate_chunk_config,
    )
    from backend.services.export_clients.clickhouse import (
        is_ch_timeout_estimate_error,
        is_transient_stream_error,
    )

    # 子块自动分裂的最大递归深度
    # 5 天 → 2,3 → 1,1,1,2 → 1,1,1,1,1（最多 4 层即可降到 1 天）
    MAX_SUBDIVISION_DEPTH = 4

    sql = config["query_sql"]
    env = config["connection_env"]
    conn_type = config.get("connection_type", "clickhouse")
    batch_size = config.get("batch_size", DEFAULT_BATCH_SIZE)
    output_dir = Path(config["output_dir"])
    job_name = config.get("job_name") or "export"
    raw_chunk_cfg = config["chunk_config"]

    if not _mark_running(job_id):
        return

    # 校验 + 切分
    try:
        ncfg = validate_chunk_config(raw_chunk_cfg, sql)
        chunks = split_date_range(ncfg.date_start, ncfg.date_end, ncfg.chunk_days)
    except Exception as e:
        msg = f"分块配置校验失败: {e}"
        logger.error("[ExportJob %s] %s", job_id, msg)
        _mark_failed(job_id, msg)
        return

    output_dir.mkdir(parents=True, exist_ok=True)

    # 列预检（任何块失败前一次性确认 SQL 可用）
    try:
        first_sql, _ = inject_date_filter(
            sql, ncfg.date_column,
            chunks[0].start, chunks[0].end,
        )
        client = _build_export_client(env, conn_type)
        client.get_columns(first_sql)
    except Exception as e:
        msg = f"获取列信息失败：{_humanize_error(e)}"
        logger.error("[ExportJob %s] %s", job_id, msg)
        _mark_failed(job_id, msg)
        return

    # 初始化清单（以 list 形式，索引可变，支持中途插入子块）
    def _make_entry(start_d, end_d, depth: int = 0) -> Dict[str, Any]:
        fn = build_chunk_filename(job_name, start_d, end_d)
        return {
            "index": -1,  # 动态分配（subdivision 时会重排）
            "date_start": start_d.isoformat(),
            "date_end": end_d.isoformat(),
            "filename": fn,
            "file_path": str(output_dir / fn),
            "file_size": None,
            "rows": 0,
            "sheets": 0,
            "status": "pending",
            "_depth": depth,  # 内部字段，跟踪递归深度（不影响前端展示）
        }

    output_files: List[Dict[str, Any]] = [
        _make_entry(c.start, c.end) for c in chunks
    ]
    # 重排索引
    for i, entry in enumerate(output_files):
        entry["index"] = i
    _update_job(
        job_id,
        total_batches=len(output_files),
        done_batches=0,
        output_files=list(output_files),  # 拷贝触发 JSONB diff
    )

    cumulative_rows = 0
    cumulative_sheets = 0

    # 索引迭代而非 for-each，便于中途插入子块再重试
    completed_count = 0
    cur_idx = 0
    while cur_idx < len(output_files):
        entry = output_files[cur_idx]
        chunk_start = date.fromisoformat(entry["date_start"])
        chunk_end = date.fromisoformat(entry["date_end"])
        days_in_chunk = (chunk_end - chunk_start).days + 1
        cur_depth = entry.get("_depth", 0)

        # 启动每块前检查是否已被取消
        if _is_cancelling(job_id):
            logger.info(
                "[ExportJob %s] Cancelled before chunk %d/%d",
                job_id, cur_idx + 1, len(output_files),
            )
            _mark_cancelled(
                job_id,
                exported=cumulative_rows,
                done_b=completed_count,
                sheets=cumulative_sheets,
            )
            _update_job(job_id, output_files=list(output_files))
            return

        chunk_sql, _ = inject_date_filter(sql, ncfg.date_column, chunk_start, chunk_end)
        chunk_path = entry["file_path"]
        chunk_label = (
            f"块 {cur_idx + 1}/{len(output_files)} "
            f"({chunk_start}~{chunk_end})"
            + (f" [子块 L{cur_depth}]" if cur_depth > 0 else "")
        )

        # 标记本块为 running
        entry["status"] = "running"
        _update_job(
            job_id,
            current_sheet=chunk_label,
            output_files=list(output_files),
        )

        try:
            result = _run_single_export(
                job_id=job_id,
                sql=chunk_sql,
                env=env,
                conn_type=conn_type,
                batch_size=batch_size,
                output_path=chunk_path,
                sheet_prefix="Sheet",
                progress_offset=cumulative_rows,
                progress_total=len(output_files),
                chunk_label=chunk_label,
            )
        except Exception as exc:
            # ── 自动日期再细分（v2.13）─────────────────────────────────────
            # 触发条件：流式断开 / Code 160 / 且仍有可分裂空间 / 未达递归上限
            is_retryable = (
                is_transient_stream_error(exc)
                or is_ch_timeout_estimate_error(exc)
            )
            can_subdivide = (
                is_retryable
                and days_in_chunk > 1
                and cur_depth < MAX_SUBDIVISION_DEPTH
            )

            if can_subdivide:
                logger.warning(
                    "[ExportJob %s] Chunk %s~%s (depth=%d) 失败 (%s) "
                    "— 自动对半分裂为更小子块重试",
                    job_id, chunk_start, chunk_end, cur_depth,
                    type(exc).__name__,
                )
                # 删除可能已写入的部分文件
                try:
                    os.unlink(chunk_path)
                except FileNotFoundError:
                    pass
                except Exception:
                    pass

                # 对半分裂日期范围，新建子条目，原位替换
                sub_ranges = subdivide_date_range(chunk_start, chunk_end)
                sub_entries = [
                    _make_entry(s, e, depth=cur_depth + 1)
                    for s, e in sub_ranges
                ]
                # 原位替换：output_files[cur_idx] → sub_entries（不前进 cur_idx）
                output_files = (
                    output_files[:cur_idx]
                    + sub_entries
                    + output_files[cur_idx + 1:]
                )
                # 重排索引
                for i, e in enumerate(output_files):
                    e["index"] = i
                _update_job(
                    job_id,
                    total_batches=len(output_files),
                    output_files=list(output_files),
                )
                # 不前进 cur_idx，下次循环执行第一个子块
                continue

            # 不可重试 / 不可再分 → 整个 Job 失败
            msg = (
                f"块 {cur_idx + 1}/{len(output_files)} "
                f"({chunk_start}~{chunk_end}) 执行失败：{_humanize_error(exc)}"
            )
            logger.error("[ExportJob %s] %s", job_id, msg, exc_info=True)
            entry["status"] = "failed"
            _mark_failed(
                job_id, msg,
                exported=cumulative_rows,
                done_b=completed_count,
            )
            _update_job(job_id, output_files=list(output_files))
            return

        chunk_size_bytes = (
            Path(chunk_path).stat().st_size
            if Path(chunk_path).exists() else None
        )

        if result["cancelled"]:
            entry.update({
                "status": "cancelled",
                "rows": result["exported_rows"],
                "sheets": result["total_sheets"],
                "file_size": chunk_size_bytes,
            })
            cumulative_rows += result["exported_rows"]
            cumulative_sheets += result["total_sheets"]
            _mark_cancelled(
                job_id,
                exported=cumulative_rows,
                done_b=completed_count,
                sheets=cumulative_sheets,
            )
            _update_job(job_id, output_files=list(output_files))
            return

        # 正常完成本块
        entry.update({
            "status": "completed",
            "rows": result["exported_rows"],
            "sheets": result["total_sheets"],
            "file_size": chunk_size_bytes,
        })
        cumulative_rows += result["exported_rows"]
        cumulative_sheets += result["total_sheets"]
        completed_count += 1

        _update_job(
            job_id,
            exported_rows=cumulative_rows,
            done_batches=completed_count,
            total_sheets=cumulative_sheets,
            output_files=list(output_files),
        )
        cur_idx += 1

    # 全部完成
    total_size = sum(
        (f.get("file_size") or 0) for f in output_files
    )
    _mark_completed(
        job_id,
        exported=cumulative_rows,
        done_b=len(output_files),
        total_b=len(output_files),
        sheets=cumulative_sheets,
        file_size=total_size,
    )
    _update_job(job_id, output_files=list(output_files))
    logger.info(
        "[ExportJob %s] Chunked export completed: %d files, %d rows total, %.1f MB"
        " (含自动分裂的子块: %d)",
        job_id, len(output_files), cumulative_rows, total_size / 1024 / 1024,
        sum(1 for f in output_files if f.get("_depth", 0) > 0),
    )
