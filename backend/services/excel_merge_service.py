"""
合并 Excel 文件服务（小工具）

核心能力：
- 多个 xlsx 按文件名排序后合并为一个文件
- 首文件表头保留一次（若含表头），后续文件只追加数据行
- 单 Sheet 超过 MAX_ROWS_PER_SHEET 自动分割到下一个 Sheet
- 合并前做列结构一致性校验（列数不同 → 阻断；表头文字不同 → 非阻断 warning）
- 协作式取消 + 进度定期落库，供前端轮询

线程安全说明：
  run_merge_job 是 async 包装器，将同步阻塞的 openpyxl/xlsxwriter 读写工作
  交给线程池（run_in_executor），避免阻塞 asyncio 事件循环。
"""
import asyncio
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import xlsxwriter

logger = logging.getLogger(__name__)

_MERGE_EXECUTOR = ThreadPoolExecutor(max_workers=4, thread_name_prefix="merge-excel-worker")

MAX_ROWS_PER_SHEET = 1_000_000       # 每 Sheet 最大数据行（不含表头），与数据导出保持一致
PROGRESS_UPDATE_EVERY_ROWS = 20_000  # 每 N 行更新一次 DB 进度
CANCEL_CHECK_EVERY_ROWS = 5_000      # 每 N 行检查一次取消状态

# Excel 最大列数。部分来源工具生成的 xlsx 会写入错误的 <dimension> 元数据
# （声明的列范围比 sheetData 里实际写入的单元格少），openpyxl 的 read_only
# 流式模式默认信任 <dimension> 来决定每行读多少列，坏元数据会导致静默截断。
# 显式传入这个硬上限作为 max_col 可以绕开该信任、读到真实写入的单元格。
_EXCEL_MAX_COLS = 16384

# 部分来源文件把 UTF-8 BOM 字符错误地写进了字符串单元格内容本身
# （而不是文件级 BOM），导致表头显示为 "﻿enterprise_name"。合并时顺手清理。
_BOM = "﻿"


def _clean_cell(value: Any) -> Any:
    if isinstance(value, str) and value.startswith(_BOM):
        return value.lstrip(_BOM)
    return value


_XLSX_OPTIONS = {
    "constant_memory": True,
    "strings_to_numbers": False,
    "strings_to_formulas": False,
    "strings_to_urls": False,
}


# ─────────────────────────────────────────────────────────────────────────────
# Job 存取辅助
# ─────────────────────────────────────────────────────────────────────────────

def _is_cancelling(job_id: str) -> bool:
    from backend.config.database import SessionLocal
    from backend.models.merge_excel_job import MergeExcelJob
    db = SessionLocal()
    try:
        j = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        return j is not None and j.status == "cancelling"
    finally:
        db.close()


def _update_job(job_id: str, **fields) -> None:
    """通用 MergeExcelJob 字段更新（自动 set updated_at + commit）"""
    from backend.config.database import SessionLocal
    from backend.models.merge_excel_job import MergeExcelJob
    db = SessionLocal()
    try:
        j = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        if j is None:
            return
        for k, v in fields.items():
            setattr(j, k, v)
        j.updated_at = datetime.utcnow()
        db.commit()
    finally:
        db.close()


def _mark_running(job_id: str) -> Optional["Any"]:
    """
    把 Job 从 pending 标记为 running（同时检查启动竞态：cancelling → 直接 cancelled）。
    返回 Job 快照字典；None 表示已终止（cancelled 或 not found），调用方应停止。
    """
    from backend.config.database import SessionLocal
    from backend.models.merge_excel_job import MergeExcelJob
    db = SessionLocal()
    try:
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        if not job:
            logger.error("[MergeExcelJob %s] Job not found, aborting.", job_id)
            return None
        if job.status == "cancelling":
            job.status = "cancelled"
            job.finished_at = datetime.utcnow()
            job.updated_at = datetime.utcnow()
            db.commit()
            logger.info("[MergeExcelJob %s] Cancelled before start.", job_id)
            return None
        job.status = "running"
        job.started_at = datetime.utcnow()
        job.updated_at = datetime.utcnow()
        db.commit()
        return {
            "username": job.username,
            "has_header": job.has_header,
            "source_files": job.source_files or [],
            "job_name": job.job_name,
        }
    finally:
        db.close()


def _mark_failed(job_id: str, msg: str) -> None:
    _update_job(job_id, status="failed", finished_at=datetime.utcnow(), error_message=msg)


def _mark_cancelled(job_id: str, merged_rows: int, done_files: int, sheets: int) -> None:
    _update_job(
        job_id,
        status="cancelled",
        finished_at=datetime.utcnow(),
        merged_rows=merged_rows,
        done_files=done_files,
        total_sheets=sheets,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 列结构一致性校验
# ─────────────────────────────────────────────────────────────────────────────

def read_header_and_colcount(file_path: str) -> Tuple[List[Any], int]:
    """
    读取工作簿首个 Sheet 的第一行（表头或首条数据），返回 (值列表, 列数)。

    显式传入 max_col=_EXCEL_MAX_COLS，不依赖 openpyxl read_only 模式默认使用的
    <dimension> 元数据（部分来源文件该元数据本身就是错的，会导致列被截断）。
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        first_row = next(
            ws.iter_rows(min_row=1, max_row=1, max_col=_EXCEL_MAX_COLS, values_only=True), ()
        )
        # 去掉末尾全 None 的空列（openpyxl 有时会把已用区域外的空列也算进来）
        row = [_clean_cell(v) for v in first_row]
        while row and row[-1] is None:
            row.pop()
        return row, len(row)
    finally:
        wb.close()


def validate_files(
    sorted_files: List[Dict[str, Any]],
    has_header: bool,
) -> Tuple[bool, Optional[str], List[str], List[Any], int]:
    """
    以第一个文件为基准做一致性校验。

    Returns:
        (is_valid, error_message, warnings, baseline_header, baseline_colcount)
        - 列数不同 → is_valid=False，error_message 列出问题文件
        - 列数相同但表头文字不同（仅 has_header=True 时有意义）→ warnings 非阻断提示
        - baseline_header/baseline_colcount 供调用方复用，避免重复读取首文件
    """
    if not sorted_files:
        return False, "未提供任何待合并文件", [], [], 0

    baseline_header, baseline_colcount = read_header_and_colcount(sorted_files[0]["file_path"])
    mismatched_colcount: List[str] = []
    header_warnings: List[str] = []

    for f in sorted_files[1:]:
        header, colcount = read_header_and_colcount(f["file_path"])
        if colcount != baseline_colcount:
            mismatched_colcount.append(
                f"{f['filename']}（{colcount} 列，基准 {baseline_colcount} 列）"
            )
        elif has_header and header != baseline_header:
            header_warnings.append(f"{f['filename']} 的表头与首文件不一致，已按位置合并")

    if mismatched_colcount:
        return (
            False,
            "以下文件列数与首文件（按文件名排序）不一致，无法合并：" + "；".join(mismatched_colcount),
            header_warnings,
            baseline_header,
            baseline_colcount,
        )
    return True, None, header_warnings, baseline_header, baseline_colcount


# ─────────────────────────────────────────────────────────────────────────────
# 数据行读取
# ─────────────────────────────────────────────────────────────────────────────

def _iter_data_rows(file_path: str, has_header: bool, colcount: int):
    """
    流式读取工作簿首个 Sheet 的数据行（跳过表头，若有）。

    colcount 为 validate_files 已校验一致的真实列数，作为 max_col 显式传入，
    不依赖该文件自身的 <dimension> 元数据（可能出错导致截断，见 read_header_and_colcount）。
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(max_col=colcount, values_only=True)
        if has_header:
            next(rows, None)
        for row in rows:
            # 跳过 openpyxl 可能产出的全空行（已用区域末尾的空行）
            if row and any(v is not None for v in row):
                yield tuple(_clean_cell(v) for v in row)
    finally:
        wb.close()


def _write_data_row(
    ws,
    row_idx: int,
    values: Tuple[Any, ...],
    date_only_format,
    datetime_format,
) -> None:
    """
    逐列写入数据行，仅对日期/时间类型的单元格套用日期 num_format。

    之前用 ws.write_row(row_idx, 0, values, date_format) 把同一个日期格式
    当作 cell_format 套到了整行的每一列——xlsxwriter 会把该列的原始数值当成
    Excel 日期序列号来显示，导致普通整数指标（如 234020）被显示成荒谬的未来
    日期（如 2540-09-20）。这里改为按值类型逐列判断，数值/字符串列不受影响。
    """
    for col_idx, v in enumerate(values):
        if isinstance(v, datetime):
            fmt = date_only_format if v.time() == time(0, 0, 0) else datetime_format
            ws.write_datetime(row_idx, col_idx, v, fmt)
        elif isinstance(v, date):
            ws.write_datetime(row_idx, col_idx, v, date_only_format)
        else:
            ws.write(row_idx, col_idx, v)


# ─────────────────────────────────────────────────────────────────────────────
# 合并主流程
# ─────────────────────────────────────────────────────────────────────────────

def _run_merge_sync(job_id: str) -> None:
    snapshot = _mark_running(job_id)
    if snapshot is None:
        return

    username: str = snapshot["username"]
    has_header: bool = snapshot["has_header"]
    source_files: List[Dict[str, Any]] = snapshot["source_files"]
    job_name: Optional[str] = snapshot["job_name"]

    sorted_files = sorted(source_files, key=lambda f: f["filename"])
    _update_job(job_id, source_files=sorted_files, total_files=len(sorted_files))

    try:
        is_valid, error_message, warnings, baseline_header, colcount = validate_files(
            sorted_files, has_header
        )
    except Exception as exc:
        _mark_failed(job_id, f"读取源文件失败：{exc}")
        return

    if not is_valid:
        _mark_failed(job_id, error_message)
        return
    if warnings:
        _update_job(job_id, warnings=warnings)

    from backend.config.settings import settings

    customer_data_root = (
        Path(settings.allowed_directories[0])
        if settings.allowed_directories
        else Path("customer_data")
    )
    output_dir = customer_data_root / username / "tools" / "merge_excel" / "jobs"
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_name = (job_name or "merged").strip() or "merged"
    output_filename = f"{safe_name}_{job_id}.xlsx"
    output_path = output_dir / output_filename

    wb = xlsxwriter.Workbook(str(output_path), options=_XLSX_OPTIONS)
    date_only_format = wb.add_format({"num_format": "yyyy-mm-dd"})
    datetime_format = wb.add_format({"num_format": "yyyy-mm-dd hh:mm:ss"})
    sheet_num = 1
    ws = wb.add_worksheet(f"Sheet{sheet_num}")
    cur_row = 0
    sheet_row_count = 0

    def _write_header() -> int:
        nonlocal cur_row
        if has_header:
            ws.write_row(0, 0, baseline_header)
            return 1
        return 0

    cur_row = _write_header()

    merged_rows = 0
    done_files = 0

    def _new_sheet() -> None:
        nonlocal sheet_num, ws, cur_row, sheet_row_count
        sheet_num += 1
        ws = wb.add_worksheet(f"Sheet{sheet_num}")
        cur_row = _write_header()
        sheet_row_count = 0
        _update_job(job_id, current_sheet=f"Sheet{sheet_num}", total_sheets=sheet_num)

    _update_job(job_id, current_sheet=f"Sheet{sheet_num}", total_sheets=sheet_num)

    try:
        for f in sorted_files:
            for row in _iter_data_rows(f["file_path"], has_header, colcount):
                if merged_rows % CANCEL_CHECK_EVERY_ROWS == 0 and _is_cancelling(job_id):
                    wb.close()
                    _mark_cancelled(job_id, merged_rows, done_files, sheet_num)
                    logger.info("[MergeExcelJob %s] Cancelled after %d rows.", job_id, merged_rows)
                    return

                if sheet_row_count >= MAX_ROWS_PER_SHEET:
                    _new_sheet()

                _write_data_row(ws, cur_row, row, date_only_format, datetime_format)
                cur_row += 1
                sheet_row_count += 1
                merged_rows += 1

                if merged_rows % PROGRESS_UPDATE_EVERY_ROWS == 0:
                    _update_job(job_id, merged_rows=merged_rows)

            done_files += 1
            _update_job(job_id, done_files=done_files, merged_rows=merged_rows)

        wb.close()
        file_size = output_path.stat().st_size
        _update_job(
            job_id,
            status="completed",
            finished_at=datetime.utcnow(),
            merged_rows=merged_rows,
            total_rows=merged_rows,
            done_files=done_files,
            total_sheets=sheet_num,
            output_filename=output_filename,
            file_path=str(output_path),
            file_size=file_size,
        )
        logger.info(
            "[MergeExcelJob %s] Completed: %d rows, %d sheets, %d files.",
            job_id, merged_rows, sheet_num, done_files,
        )
    except Exception as exc:
        try:
            wb.close()
        except Exception:
            pass
        logger.exception("[MergeExcelJob %s] Merge failed.", job_id)
        _mark_failed(job_id, f"合并失败：{exc}")


async def run_merge_job(job_id: str) -> None:
    """后台协程包装器：将同步阻塞的合并工作提交到线程池执行。"""
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(_MERGE_EXECUTOR, _run_merge_sync, job_id)
