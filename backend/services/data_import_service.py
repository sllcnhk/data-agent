"""
数据导入服务

实现 Excel → ClickHouse 的核心业务逻辑：
- 可写连接枚举
- Schema/Table 查询
- Excel 文件预览解析
- 分批导入执行（abort on first batch failure）
"""
import asyncio
import logging
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

# 复用导出侧已在生产验证过的 Code 202 判错逻辑（三种报错写法都覆盖），
# 不另写一套。该模块只依赖 stdlib + requests，无循环导入风险。
from backend.services.export_clients.clickhouse import is_ch_too_many_queries_error

logger = logging.getLogger(__name__)

# 每批默认行数（TabSeparated 格式下 5000 行仍是小请求，HTTP 往返次数少 5 倍）
DEFAULT_BATCH_SIZE = 5000
# 文件大小上限 100 MB
MAX_FILE_SIZE = 100 * 1024 * 1024
# 预览行数
PREVIEW_ROWS = 5

# ClickHouse Code 202 (TOO_MANY_SIMULTANEOUS_QUERIES) 退避重试默认值。
# 该错误由 CH 的准入控制在查询注册阶段抛出，此时服务端尚未开始写入数据块，
# 因此失败批次 0 行入库，重发同一批是安全的（不会产生重复行）。
# 可用 IMPORT_TOO_MANY_RETRY_MAX / IMPORT_TOO_MANY_RETRY_BACKOFF 覆盖。
DEFAULT_TOO_MANY_RETRY_MAX = 5
DEFAULT_TOO_MANY_RETRY_BACKOFF = 15

# 导入批插入的 HTTP 超时（秒）。查询类操作仍用 60s。
# 理由：batch_size 提到 5000 后，含长文本列（如 call_record 的对话详情）的单批
# body 可达 ~20 MB。THAI 实测（2026-09-08，19.62 MB）典型 6.3s、最差 19.8s，
# 跨公网抖动明显，60s 余量偏薄。而超时抛出的是 TimeoutError —— 这是唯一**不能**
# 重试的错误（客户端放弃等待时服务端可能已写入成功，重发会产生无法察觉的重复
# 行），因此一次假超时会直接让任务失败且无法自愈。放宽到 180s 换取余量。
DEFAULT_INSERT_TIMEOUT = 180


class _ImportCancelled(Exception):
    """内部信号：重试等待期间收到取消请求，需按 cancelled 而非 failed 收尾。"""


class _BatchInsertFailed(Exception):
    """一批插入最终失败，携带原始异常与已重试次数，便于写入断点信息。"""

    def __init__(self, cause: BaseException, attempts: int) -> None:
        super().__init__(str(cause))
        self.cause = cause
        self.attempts = attempts

# ─────────────────────────────────────────────────────────────────────────────
# call_record_imported 常量
# ─────────────────────────────────────────────────────────────────────────────

# 这 15 列在目标表中有独立字段，其余列打包进 tag_array
_CR_FIXED_COLS = [
    "Task Name", "Dialogue Name", "Contact ID", "Audio Name", "Call ID",
    "Result", "Call Time", "Call Duration", "Agent Call Duration",
    "Dialogue Round", "Have Read", "Agent", "Tags", "Transfer Status",
    "Call Record Text Detail Masked",
]
_CR_FIXED_SET = set(_CR_FIXED_COLS)

# 目标表显式插入列（不插 id/import_time，有 DEFAULT）
_CR_TARGET_COLS = [
    "import_job_id", "source_file",
    "task_name", "dialogue_name", "contact_id", "audio_name", "call_id",
    "result", "call_start_time", "call_duration", "agent_call_duration",
    "dialogue_round", "have_read", "agent", "tags", "transfer_status",
    "call_record_text_detail_masked", "tag_array",
]


def _parse_call_time(raw) -> Optional[str]:
    """DD/MM/YYYY HH:MM:SS → 'YYYY-MM-DD HH:MM:SS'，解析失败返回 None。"""
    if raw is None:
        return None
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def _safe_int(v) -> Optional[int]:
    """转 int，失败返回 None。"""
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _safe_str(v) -> str:
    """转 str，None/NaN → 空串。"""
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def transform_call_record_batch(
    rows: List[Tuple],
    headers: List[str],
    job_id: str,
    source_file: str,
) -> List[Dict[str, Any]]:
    """
    将 call_record_imported 类型的一批原始 Excel 行转换为目标表 dict 列表。

    - 固定 15 列 → 独立字段（Call Time 解析为 DateTime）
    - 其余非空列 → tag_array（格式 "ColName=value"）
    - 返回 list[dict]，key 与 _CR_TARGET_COLS 一致
    """
    # 按列名定位索引（容错列序）
    col_idx: Dict[str, int] = {h: i for i, h in enumerate(headers)}

    def _get(row, col_name) -> Any:
        idx = col_idx.get(col_name)
        return row[idx] if idx is not None and idx < len(row) else None

    extra_cols = [h for h in headers if h not in _CR_FIXED_SET]

    result = []
    for row in rows:
        tag_array = []
        for col in extra_cols:
            val = _get(row, col)
            s = _safe_str(val)
            if s:
                tag_array.append(f"{col}={s}")

        result.append({
            "import_job_id": job_id,
            "source_file":   source_file,
            "task_name":     _safe_str(_get(row, "Task Name")),
            "dialogue_name": _safe_str(_get(row, "Dialogue Name")),
            "contact_id":    _safe_str(_get(row, "Contact ID")),
            "audio_name":    _safe_str(_get(row, "Audio Name")),
            "call_id":       _safe_str(_get(row, "Call ID")),
            "result":        _safe_str(_get(row, "Result")),
            "call_start_time":               _parse_call_time(_get(row, "Call Time")),
            "call_duration":                 _safe_int(_get(row, "Call Duration")),
            "agent_call_duration":           _safe_int(_get(row, "Agent Call Duration")),
            "dialogue_round":                _safe_int(_get(row, "Dialogue Round")),
            "have_read":                     _safe_str(_get(row, "Have Read")),
            "agent":                         _safe_str(_get(row, "Agent")),
            "tags":                          _safe_str(_get(row, "Tags")),
            "transfer_status":               _safe_str(_get(row, "Transfer Status")),
            "call_record_text_detail_masked": _safe_str(_get(row, "Call Record Text Detail Masked")),
            "tag_array": tag_array,
        })
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 辅助：构建 ClickHouseHTTPClient
# ─────────────────────────────────────────────────────────────────────────────

def _build_ch_client(env: str, timeout: int = 60):
    """
    根据 env 构建 ClickHouseHTTPClient（admin 级别）。

    timeout 默认 60s 适用于 schema 查询；批插入请传更大的值
    （见 DEFAULT_INSERT_TIMEOUT 的说明）。
    """
    from backend.config.settings import settings
    from backend.mcp.clickhouse.http_client import ClickHouseHTTPClient

    cfg = settings.get_clickhouse_config(env, level="admin")
    return ClickHouseHTTPClient(
        host=cfg["host"],
        port=cfg["http_port"],
        user=cfg["user"],
        password=cfg["password"],
        database=cfg["database"],
        timeout=timeout,
    )


# ─────────────────────────────────────────────────────────────────────────────
# 1. 连接列表
# ─────────────────────────────────────────────────────────────────────────────

def list_writable_connections() -> List[Dict[str, Any]]:
    """
    返回所有可写（admin 级别）的 ClickHouse 连接信息。
    过滤依据：MCPServerManager 中名称不含 -ro 的 clickhouse-* 连接。
    """
    from backend.config.settings import settings
    from backend.mcp.manager import get_mcp_manager

    manager = get_mcp_manager()
    result = []

    for name, server in manager.servers.items():
        # 只取 ClickHouse 服务，排除只读副本
        if not name.startswith("clickhouse-"):
            continue
        if name.endswith("-ro"):
            continue

        # 从 name 反推 env（clickhouse-sg-azure → sg-azure → sg_azure）
        env_dash = name[len("clickhouse-"):]   # 去掉前缀
        env = env_dash.replace("-", "_")       # 连字符→下划线

        cfg = settings.get_clickhouse_config(env, level="admin")
        result.append({
            "env": env,
            "server_name": name,
            "host": cfg["host"],
            "http_port": cfg["http_port"],
            "database": cfg["database"],
            "display_name": name,
        })

    return result


# ─────────────────────────────────────────────────────────────────────────────
# 2. Schema / Table 查询
# ─────────────────────────────────────────────────────────────────────────────

def list_databases(env: str) -> List[str]:
    """查询指定环境的数据库列表（排除系统库）"""
    client = _build_ch_client(env)
    rows = client.execute(
        "SELECT name FROM system.databases "
        "WHERE name NOT IN ('system', 'information_schema', 'INFORMATION_SCHEMA') "
        "ORDER BY name"
    )
    return [r[0] for r in rows]


def list_tables(env: str, database: str) -> List[str]:
    """查询指定环境和数据库的表列表"""
    client = _build_ch_client(env)
    rows = client.execute(
        f"SELECT name FROM system.tables "
        f"WHERE database = '{database}' ORDER BY name"
    )
    return [r[0] for r in rows]


def describe_table(env: str, database: str, table: str) -> List[Dict[str, str]]:
    """获取表字段信息（name, type）"""
    client = _build_ch_client(env)
    rows, col_types = client.execute(
        f"DESCRIBE TABLE `{database}`.`{table}`",
        with_column_types=True,
    )
    return [{"name": r[0], "type": r[1]} for r in rows]


# ─────────────────────────────────────────────────────────────────────────────
# 3. Excel 解析预览
# ─────────────────────────────────────────────────────────────────────────────

def parse_excel_preview(file_path: str) -> List[Dict[str, Any]]:
    """
    用 openpyxl 流式模式读取 Excel，返回每个 Sheet 的预览信息。

    性能优化：
    - 行数使用 ws.max_row 元数据（O(1)），不再遍历全部行
    - 迭代器只读取前 PREVIEW_ROWS 行后立即停止

    Returns:
        [{
            "sheet_name": str,
            "row_count_estimate": int,    # 估算总行数（含表头）
            "preview_rows": [[cell, ...]],  # 前 PREVIEW_ROWS 行原始值
        }]
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheets = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            preview_rows: List[List] = []

            for row in ws.iter_rows(values_only=True):
                if len(preview_rows) >= PREVIEW_ROWS:
                    break  # 预览行足够，立即停止，不扫描剩余行
                preview_rows.append([
                    str(cell) if cell is not None else ""
                    for cell in row
                ])

            sheets.append({
                "sheet_name": sheet_name,
                "row_count_estimate": ws.max_row or 0,  # O(1) 元数据，无需全量遍历
                "preview_rows": preview_rows,
            })
    finally:
        wb.close()

    return sheets


# ─────────────────────────────────────────────────────────────────────────────
# 4. 核心导入逻辑
# ─────────────────────────────────────────────────────────────────────────────

def _env_int(name: str, default: int) -> int:
    """读取整数型环境变量，缺失或非法时回退默认值。"""
    raw = os.getenv(name)
    if raw is None or str(raw).strip() == "":
        return default
    try:
        return int(str(raw).strip())
    except ValueError:
        logger.warning("环境变量 %s=%r 不是整数，回退默认值 %d", name, raw, default)
        return default


def _rows_to_values_clause(rows: List[Tuple]) -> str:
    """将行列表转换为 INSERT VALUES 子句"""
    def _fmt(v) -> str:
        if v is None:
            return "NULL"
        if isinstance(v, bool):
            return "1" if v else "0"
        if isinstance(v, (int, float)):
            return str(v)
        # 字符串：转义单引号
        return "'" + str(v).replace("'", "\\'") + "'"

    parts = []
    for row in rows:
        cols = ", ".join(_fmt(cell) for cell in row)
        parts.append(f"({cols})")
    return ", ".join(parts)


async def run_import_job(job_id: str, config: Dict[str, Any]) -> None:
    """
    后台协程：逐 Sheet 分批读取 Excel 并插入 ClickHouse。

    config 结构：
    {
        "file_path": str,
        "connection_env": str,
        "batch_size": int,
        "sheets": [
            {
                "sheet_name": str,
                "database": str,
                "table": str,
                "has_header": bool,
                "enabled": bool,
            }
        ],
        "db_session_factory": callable,   # SessionLocal
    }
    """
    import openpyxl
    from backend.config.database import SessionLocal
    from backend.models.import_job import ImportJob

    file_path = config["file_path"]
    env = config["connection_env"]
    batch_size = config.get("batch_size", DEFAULT_BATCH_SIZE)
    sheet_configs = [s for s in config["sheets"] if s.get("enabled", True)]

    # 每次运行时读取，便于运维改 env 后重启即生效、也便于测试注入
    retry_max = max(0, _env_int("IMPORT_TOO_MANY_RETRY_MAX", DEFAULT_TOO_MANY_RETRY_MAX))
    retry_backoff = max(0, _env_int("IMPORT_TOO_MANY_RETRY_BACKOFF", DEFAULT_TOO_MANY_RETRY_BACKOFF))

    def _get_job(db) -> Optional[ImportJob]:
        return db.query(ImportJob).filter(ImportJob.id == job_id).first()

    def _save(db, job: ImportJob):
        job.updated_at = datetime.utcnow()
        db.commit()

    # ── 标记开始 ──────────────────────────────────────────────────────────────
    db = SessionLocal()
    try:
        job = _get_job(db)
        if not job:
            logger.error("[ImportJob %s] Job not found, aborting.", job_id)
            return
        # 取消请求可能在任务启动前就到达（pending → cancelling）
        if job.status == "cancelling":
            job.status = "cancelled"
            job.finished_at = datetime.utcnow()
            _save(db, job)
            logger.info("[ImportJob %s] Cancelled before start.", job_id)
            return
        job.status = "running"
        job.started_at = datetime.utcnow()
        job.total_sheets = len(sheet_configs)
        _save(db, job)
    finally:
        db.close()

    client = _build_ch_client(
        env, timeout=_env_int("IMPORT_INSERT_TIMEOUT", DEFAULT_INSERT_TIMEOUT)
    )
    errors: List[Dict] = []
    total_imported = 0
    total_batches_all = 0
    done_batches_all = 0

    # ── 快速估算总批次（用 max_row，不全量遍历文件）────────────────────────
    try:
        import openpyxl as _ox
        wb_scan = _ox.load_workbook(file_path, read_only=True, data_only=True)
        for sc in sheet_configs:
            ws = wb_scan[sc["sheet_name"]]
            row_count = ws.max_row or 0
            data_rows = max(row_count - (1 if sc.get("has_header", True) else 0), 0)
            batches = (data_rows + batch_size - 1) // batch_size if data_rows else 0
            total_batches_all += batches
        wb_scan.close()
    except Exception as e:
        logger.warning("[ImportJob %s] Pre-scan failed: %s", job_id, e)

    db = SessionLocal()
    try:
        job = _get_job(db)
        job.total_batches = total_batches_all
        _save(db, job)
    finally:
        db.close()

    # ── 逐 Sheet 导入 ─────────────────────────────────────────────────────────
    done_sheets = 0
    abort_flag = False

    def _is_cancelling() -> bool:
        """检查任务是否被请求取消（每批次调用一次）"""
        db = SessionLocal()
        try:
            j = _get_job(db)
            return j is not None and j.status == "cancelling"
        finally:
            db.close()

    def _mark_cancelled():
        db = SessionLocal()
        try:
            j = _get_job(db)
            if j:
                j.status = "cancelled"
                j.finished_at = datetime.utcnow()
                j.imported_rows = total_imported
                j.done_batches = done_batches_all
                _save(db, j)
        finally:
            db.close()

    def _fail_job(err_msg: str) -> None:
        """按 abort 策略把任务标记为 failed 并落盘当前进度与错误列表。"""
        db = SessionLocal()
        try:
            j = _get_job(db)
            if j:
                j.status = "failed"
                j.error_message = err_msg
                j.errors = list(errors)
                j.imported_rows = total_imported
                j.done_batches = done_batches_all
                j.finished_at = datetime.utcnow()
                _save(db, j)
        finally:
            db.close()

    def _persist_errors() -> None:
        """仅落盘 errors 与进度（任务仍在运行，不改 status）。

        重试期间会原地停留 retry_max × retry_backoff 秒，前端轮询到的
        done_batches 一动不动。必须把 warning 立刻写库，否则用户分辨不出
        「正在退避重试」和「真的卡死」。
        """
        db = SessionLocal()
        try:
            j = _get_job(db)
            if j:
                j.errors = list(errors)
                j.imported_rows = total_imported
                j.done_batches = done_batches_all
                _save(db, j)
        finally:
            db.close()

    def _batch_fail_message(
        sheet: str, batch_num: int, batch_len: int, exc: BaseException
    ) -> str:
        """构造带断点信息的失败消息，让用户一眼知道已入库多少、要清理多少。"""
        attempts = getattr(exc, "attempts", 0)
        cause = getattr(exc, "cause", exc)
        retried = f"（已重试 {attempts} 次仍失败）" if attempts else ""
        if batch_num > 1:
            done_desc = f"当前 Sheet '{sheet}' 第 1~{batch_num - 1} 批已入库"
        else:
            done_desc = f"当前 Sheet '{sheet}' 尚无批次入库"
        return (
            f"Sheet '{sheet}' 第 {batch_num} 批插入失败{retried}: {cause}\n"
            f"断点信息：本任务已成功导入 {total_imported} 行（{done_desc}），"
            f"第 {batch_num} 批的 {batch_len} 行未入库。重导前需清理已入库的数据。"
        )

    for sc in sheet_configs:
        if abort_flag:
            break

        sheet_name = sc["sheet_name"]
        database = sc["database"]
        table = sc["table"]
        has_header = sc.get("has_header", True)
        import_type = sc.get("import_type", "standard")

        # 每个 sheet 开始前检查取消
        if _is_cancelling():
            logger.info("[ImportJob %s] Cancelled before sheet '%s'.", job_id, sheet_name)
            _mark_cancelled()
            return

        # 更新当前 sheet
        db = SessionLocal()
        try:
            job = _get_job(db)
            job.current_sheet = sheet_name
            _save(db, job)
        finally:
            db.close()

        logger.info("[ImportJob %s] Starting sheet '%s' → %s.%s [type=%s]",
                    job_id, sheet_name, database, table, import_type)

        # ── 批次插入辅助（根据 import_type 选择路径）────────────────────────
        cr_headers: List[str] = []   # call_record_imported 时保存表头列名

        def _do_insert(batch: List[Tuple]) -> None:
            if import_type == "call_record_imported":
                transformed = transform_call_record_batch(
                    batch, cr_headers, job_id, os.path.basename(file_path)
                )
                client.insert_json_rows(database, table, _CR_TARGET_COLS, transformed)
            else:
                client.insert_tsv(database, table, batch)

        async def _insert_with_retry(batch: List[Tuple], batch_num: int) -> int:
            """
            插入一批数据；**仅**对 ClickHouse Code 202
            (TOO_MANY_SIMULTANEOUS_QUERIES) 做固定间隔退避重试。

            重发的是同一批同一份数据，openpyxl 行迭代器不回退，游标停在原地，
            因此天然是断点续传：已成功的批次不会重发，失败的批次不会跳过。

            超时（TimeoutError）与连接错误（ConnectionError）**刻意不重试**：
            客户端超时只是单方面放弃等待，服务端可能正在成功写入，重发会产生
            无法事后察觉的重复行。

            Returns:
                实际重试次数（0 表示首次即成功）
            Raises:
                _ImportCancelled:   重试等待期间任务被请求取消
                _BatchInsertFailed: 非 202 错误，或 202 但重试次数已用尽
            """
            attempt = 0
            while True:
                try:
                    _do_insert(batch)
                    return attempt
                except Exception as exc:
                    if not is_ch_too_many_queries_error(exc) or attempt >= retry_max:
                        raise _BatchInsertFailed(exc, attempt)

                    attempt += 1
                    logger.warning(
                        "[ImportJob %s] Sheet '%s' 第 %d 批遇 ClickHouse 并发上限，"
                        "%ds 后重试 (%d/%d): %s",
                        job_id, sheet_name, batch_num, retry_backoff,
                        attempt, retry_max, exc,
                    )
                    errors.append({
                        "sheet": sheet_name,
                        "batch": batch_num,
                        "level": "warning",
                        "message": (
                            f"ClickHouse 并发上限，{retry_backoff}s 后重试 "
                            f"({attempt}/{retry_max}): {exc}"
                        ),
                    })
                    _persist_errors()

                    # 必须 await asyncio.sleep 而非 time.sleep：本协程与 FastAPI
                    # 共用 event loop，阻塞式 sleep 会冻住整个后端。
                    # 按秒切片是为了让「取消」最多 1 秒生效，而不是等满
                    # retry_max × retry_backoff 秒。
                    for _ in range(retry_backoff):
                        await asyncio.sleep(1)
                        if _is_cancelling():
                            raise _ImportCancelled()

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            batch_rows: List[Tuple] = []
            sheet_imported = 0
            sheet_batch_num = 0
            is_first_row = True
            _row_iter = ws.iter_rows(values_only=True)

            for row in _row_iter:
                # 处理表头行
                if is_first_row and has_header:
                    is_first_row = False
                    if import_type == "call_record_imported":
                        cr_headers = [str(c) if c is not None else "" for c in row]
                    continue
                is_first_row = False

                batch_rows.append(row)

                if len(batch_rows) >= batch_size:
                    sheet_batch_num += 1
                    try:
                        await _insert_with_retry(batch_rows, sheet_batch_num)
                        sheet_imported += len(batch_rows)
                        total_imported += len(batch_rows)
                        done_batches_all += 1
                    except _ImportCancelled:
                        # 重试等待期间被取消：按 cancelled 收尾，不是 failed
                        logger.info(
                            "[ImportJob %s] Cancelled during retry wait, sheet '%s' "
                            "batch %d, after %d rows.",
                            job_id, sheet_name, sheet_batch_num, sheet_imported,
                        )
                        _row_iter.close()
                        wb.close()
                        _mark_cancelled()
                        return
                    except Exception as e:
                        err_msg = _batch_fail_message(
                            sheet_name, sheet_batch_num, len(batch_rows), e
                        )
                        logger.error("[ImportJob %s] %s", job_id, err_msg)
                        errors.append({
                            "sheet": sheet_name,
                            "batch": sheet_batch_num,
                            "level": "error",
                            "message": str(getattr(e, "cause", e)),
                        })
                        abort_flag = True
                        # abort 策略：立即终止
                        _fail_job(err_msg)
                        _row_iter.close()  # 显式关闭生成器，释放 Windows 文件锁
                        wb.close()
                        break

                    batch_rows = []

                    # 每 10 批更新一次进度（减少 PostgreSQL round-trip）
                    if done_batches_all % 10 == 0:
                        db = SessionLocal()
                        try:
                            job = _get_job(db)
                            job.imported_rows = total_imported
                            job.done_batches = done_batches_all
                            _save(db, job)
                        finally:
                            db.close()

                    # 让出事件循环，避免阻塞
                    await asyncio.sleep(0)

                    # 每批检查是否被取消
                    if _is_cancelling():
                        logger.info("[ImportJob %s] Cancelled mid-sheet '%s' after %d rows.",
                                    job_id, sheet_name, sheet_imported)
                        _row_iter.close()
                        wb.close()
                        _mark_cancelled()
                        return

            if abort_flag:
                break

            # 尾部剩余行
            if batch_rows:
                sheet_batch_num += 1
                try:
                    await _insert_with_retry(batch_rows, sheet_batch_num)
                    sheet_imported += len(batch_rows)
                    total_imported += len(batch_rows)
                    done_batches_all += 1
                except _ImportCancelled:
                    logger.info(
                        "[ImportJob %s] Cancelled during retry wait, sheet '%s' "
                        "tail batch %d, after %d rows.",
                        job_id, sheet_name, sheet_batch_num, sheet_imported,
                    )
                    wb.close()
                    _mark_cancelled()
                    return
                except Exception as e:
                    err_msg = _batch_fail_message(
                        sheet_name, sheet_batch_num, len(batch_rows), e
                    )
                    logger.error("[ImportJob %s] %s", job_id, err_msg)
                    errors.append({
                        "sheet": sheet_name,
                        "batch": sheet_batch_num,
                        "level": "error",
                        "message": str(getattr(e, "cause", e)),
                    })
                    abort_flag = True
                    _fail_job(err_msg)
                    wb.close()
                    break

            wb.close()

        except _ImportCancelled:
            # 兜底：两处批次调用点已各自处理，此分支正常不会命中，
            # 但必须在 except Exception 之前，避免取消被误报成「解析失败」
            logger.info("[ImportJob %s] Cancelled during retry wait (sheet '%s').",
                        job_id, sheet_name)
            _mark_cancelled()
            return
        except Exception as e:
            err_msg = f"Sheet '{sheet_name}' 解析失败: {e}"
            logger.error("[ImportJob %s] %s", job_id, err_msg)
            errors.append({
                "sheet": sheet_name,
                "batch": 0,
                "level": "error",
                "message": str(e),
            })
            abort_flag = True
            _fail_job(err_msg)
            break

        if not abort_flag:
            done_sheets += 1
            logger.info("[ImportJob %s] Sheet '%s' done: %d rows", job_id, sheet_name, sheet_imported)
            db = SessionLocal()
            try:
                job = _get_job(db)
                job.done_sheets = done_sheets
                job.imported_rows = total_imported
                job.done_batches = done_batches_all
                _save(db, job)
            finally:
                db.close()

    # ── 最终状态 ──────────────────────────────────────────────────────────────
    if not abort_flag:
        db = SessionLocal()
        try:
            job = _get_job(db)
            job.status = "completed"
            job.done_sheets = done_sheets
            job.imported_rows = total_imported
            job.done_batches = done_batches_all
            job.finished_at = datetime.utcnow()
            job.errors = errors if errors else None
            _save(db, job)
        finally:
            db.close()
        logger.info("[ImportJob %s] Completed: %d rows imported.", job_id, total_imported)

    # 清理临时文件
    try:
        os.unlink(file_path)
        logger.info("[ImportJob %s] Temp file removed: %s", job_id, file_path)
    except Exception:
        pass
