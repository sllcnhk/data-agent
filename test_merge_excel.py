"""
小工具 - 合并Excel文件 单元测试

A · 列结构一致性校验（validate_files）
B · 数据行读取（_iter_data_rows）
C · run_merge_job 协程（正常完成、分 Sheet、文件排序、取消、失败）
D · REST API 端点（权限、状态码、上传、下载）
E · RBAC 权限种子
F · 坏 <dimension> 元数据导致列截断（真实线上故障场景回归）
G · 日期格式误套到整行导致数值列显示成荒谬日期（真实线上故障场景回归）

运行：
    python -m pytest test_merge_excel.py -v -s
"""
import asyncio
import datetime
import os
import re
import sys
import uuid
import zipfile
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent))

os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_mx_{uuid.uuid4().hex[:6]}_"


def _write_xlsx(path: Path, rows, header=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    r = 1
    if header is not None:
        ws.append(header)
        r += 1
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def _make_bad_dimension_xlsx(path: Path, header, rows, fake_dimension: str) -> Path:
    """
    构造一份 <dimension> 元数据谎报列范围、但 sheetData 实际写入更多列的 xlsx，
    复现线上真实故障（21 个源文件的 dimension 全部只声明 1 列，实际有 6 列）。

    做法：先用 openpyxl 正常写一份完整数据的文件，再直接改写 zip 包内
    xl/worksheets/sheet1.xml 里的 <dimension ref="..."/> 字符串，
    不改动 sheetData 本身 —— 精确模拟"元数据与实际内容不符"这一种故障，
    而不是简单地少写几列数据。
    """
    _write_xlsx(path, rows, header=header)

    with zipfile.ZipFile(path, "r") as zin:
        entries = {name: zin.read(name) for name in zin.namelist()}

    sheet_xml = entries["xl/worksheets/sheet1.xml"].decode("utf-8")
    patched_xml, n = re.subn(
        r'<dimension ref="[^"]+"\s*/>',
        f'<dimension ref="{fake_dimension}"/>',
        sheet_xml,
        count=1,
    )
    assert n == 1, "未找到 <dimension> 标签，openpyxl 输出格式可能已变化"
    entries["xl/worksheets/sheet1.xml"] = patched_xml.encode("utf-8")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data)

    return path


# =============================================================================
# A · 列结构一致性校验
# =============================================================================

class TestValidateFiles:
    def test_a1_matching_columns_no_warnings(self, tmp_path):
        """A1: 列数/表头完全一致 → valid, 无 warning"""
        from backend.services.excel_merge_service import validate_files

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y"]], header=["id", "name"])
        files = [
            {"filename": "a_1.xlsx", "file_path": str(f1)},
            {"filename": "b_2.xlsx", "file_path": str(f2)},
        ]
        is_valid, err, warnings, header, colcount = validate_files(files, has_header=True)
        assert is_valid is True
        assert err is None
        assert warnings == []
        assert header == ["id", "name"]
        assert colcount == 2

    def test_a2_mismatched_colcount_fails(self, tmp_path):
        """A2: 列数不同 → invalid，错误信息含问题文件名"""
        from backend.services.excel_merge_service import validate_files

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y", "z"]], header=["id", "name", "extra"])
        files = [
            {"filename": "a_1.xlsx", "file_path": str(f1)},
            {"filename": "b_2.xlsx", "file_path": str(f2)},
        ]
        is_valid, err, warnings, _, _ = validate_files(files, has_header=True)
        assert is_valid is False
        assert "b_2.xlsx" in err

    def test_a3_matching_colcount_diff_header_warns(self, tmp_path):
        """A3: 列数相同但表头文字不同 → valid + warning（非阻断）"""
        from backend.services.excel_merge_service import validate_files

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y"]], header=["ID", "NAME"])
        files = [
            {"filename": "a_1.xlsx", "file_path": str(f1)},
            {"filename": "b_2.xlsx", "file_path": str(f2)},
        ]
        is_valid, err, warnings, _, _ = validate_files(files, has_header=True)
        assert is_valid is True
        assert len(warnings) == 1
        assert "b_2.xlsx" in warnings[0]

    def test_a4_no_header_ignores_text_diff(self, tmp_path):
        """A4: has_header=False 时不比较首行文字，只比较列数"""
        from backend.services.excel_merge_service import validate_files

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"], ["2", "y"]])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["foo", "bar"], ["3", "z"]])
        files = [
            {"filename": "a_1.xlsx", "file_path": str(f1)},
            {"filename": "b_2.xlsx", "file_path": str(f2)},
        ]
        is_valid, err, warnings, _, _ = validate_files(files, has_header=False)
        assert is_valid is True
        assert warnings == []

    def test_a5_empty_file_list_invalid(self):
        """A5: 空文件列表 → invalid"""
        from backend.services.excel_merge_service import validate_files
        is_valid, err, warnings, _, _ = validate_files([], has_header=True)
        assert is_valid is False
        assert err


# =============================================================================
# B · 数据行读取
# =============================================================================

class TestIterDataRows:
    def test_b1_has_header_skips_first_row(self, tmp_path):
        from backend.services.excel_merge_service import _iter_data_rows

        f = _write_xlsx(tmp_path / "f.xlsx", [["1", "x"], ["2", "y"]], header=["id", "name"])
        rows = list(_iter_data_rows(str(f), has_header=True, colcount=2))
        assert rows == [("1", "x"), ("2", "y")]

    def test_b2_no_header_keeps_all_rows(self, tmp_path):
        from backend.services.excel_merge_service import _iter_data_rows

        f = _write_xlsx(tmp_path / "f.xlsx", [["1", "x"], ["2", "y"]])
        rows = list(_iter_data_rows(str(f), has_header=False, colcount=2))
        assert rows == [("1", "x"), ("2", "y")]

    def test_b3_skips_trailing_blank_rows(self, tmp_path):
        from backend.services.excel_merge_service import _iter_data_rows

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["id"])
        ws.append(["1"])
        ws.append([None])  # 全空行（openpyxl 已用区域内可能出现）
        f = tmp_path / "f.xlsx"
        wb.save(f)
        rows = list(_iter_data_rows(str(f), has_header=True, colcount=1))
        assert rows == [("1",)]

    def test_b4_bom_prefix_stripped_from_string_cells(self, tmp_path):
        """B4: 字符串单元格内容自带的 BOM 前缀被清理（不影响数值/日期类型）"""
        from backend.services.excel_merge_service import _iter_data_rows

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["﻿header"])
        ws.append(["﻿value"])
        f = tmp_path / "f.xlsx"
        wb.save(f)
        rows = list(_iter_data_rows(str(f), has_header=True, colcount=1))
        assert rows == [("value",)]

    def test_b5_max_col_bypasses_bad_dimension(self, tmp_path):
        """
        B5: 复现真实故障场景 —— 源文件 <dimension> 元数据只声明 1 列，
        但 sheetData 实际写入了 6 列。read_only 默认信任 dimension 会截断，
        显式传入真实 colcount 作为 max_col 应读到完整 6 列。
        """
        from backend.services.excel_merge_service import _iter_data_rows

        f = _make_bad_dimension_xlsx(
            tmp_path / "bad_dim.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["1", "x", "a", "b", "c", "d"]],
            fake_dimension="A1:A2",
        )

        # 先验：不传 colcount（旧行为，依赖 dimension 默认值）会被截断成 1 列，
        # 证明测试数据确实复现了线上故障场景。
        import openpyxl as _opx
        wb = _opx.load_workbook(str(f), read_only=True, data_only=True)
        legacy_row = next(wb.active.iter_rows(min_row=1, max_row=1, values_only=True))
        wb.close()
        assert len(legacy_row) == 1, "测试数据未复现坏 dimension 截断场景"

        # 修复后：显式传入真实列数 6，应读到完整数据
        rows = list(_iter_data_rows(str(f), has_header=True, colcount=6))
        assert rows == [("1", "x", "a", "b", "c", "d")]


# =============================================================================
# C · run_merge_job 协程
# =============================================================================

class TestRunMergeJob:
    def _make_job(self, db, username, source_files, has_header=True, status="pending"):
        from backend.models.merge_excel_job import MergeExcelJob
        job = MergeExcelJob(
            user_id="uid", username=username,
            has_header=has_header, source_files=source_files,
            status=status, total_files=len(source_files),
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)

    def _run(self, coro):
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()

    def test_c1_completed_merge_two_files(self, tmp_path):
        """C1: 2 个含表头文件合并 → 表头 1 行 + 数据行之和"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"], ["2", "y"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["3", "z"]], header=["id", "name"])
        source_files = [
            {"filename": "a_1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}c1"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "completed"
        assert job.merged_rows == 3
        assert job.total_sheets == 1
        assert Path(job.file_path).exists()

        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        assert rows[0] == ("id", "name")
        assert rows[1:] == [("1", "x"), ("2", "y"), ("3", "z")]

    def test_c2_file_order_by_filename(self, tmp_path):
        """C2: 传入乱序文件，按文件名排序后合并（首文件表头生效）"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f_b = _write_xlsx(tmp_path / "b_second.xlsx", [["2", "second"]], header=["id", "name"])
        f_a = _write_xlsx(tmp_path / "a_first.xlsx", [["1", "first"]], header=["id", "name"])
        # 故意乱序传入（b 在前）
        source_files = [
            {"filename": "b_second.xlsx", "file_path": str(f_b), "size": 1},
            {"filename": "a_first.xlsx", "file_path": str(f_a), "size": 1},
        ]

        username = f"{_PREFIX}c2"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        # a_first 排在前面 → 数据顺序 first, second
        assert rows[1:] == [("1", "first"), ("2", "second")]

    def test_c3_no_header_keeps_all_rows(self, tmp_path):
        """C3: has_header=False → 所有行都保留（不跳表头）"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"], ["2", "y"]])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["3", "z"]])
        source_files = [
            {"filename": "a_1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}c3"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files, has_header=False)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.merged_rows == 3
        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        assert rows == [("1", "x"), ("2", "y"), ("3", "z")]

    def test_c4_auto_split_sheet_when_exceeding_limit(self, tmp_path):
        """C4: 超过 MAX_ROWS_PER_SHEET(mock 为 2) → 自动分 Sheet2"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(
            tmp_path / "a_1.xlsx",
            [["1", "r1"], ["2", "r2"], ["3", "r3"]],
            header=["id", "name"],
        )
        source_files = [{"filename": "a_1.xlsx", "file_path": str(f1), "size": 1}]

        username = f"{_PREFIX}c4"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ), patch("backend.services.excel_merge_service.MAX_ROWS_PER_SHEET", 2):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.total_sheets == 2
        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        assert wb.sheetnames == ["Sheet1", "Sheet2"]
        sheet1_rows = list(wb["Sheet1"].iter_rows(values_only=True))
        sheet2_rows = list(wb["Sheet2"].iter_rows(values_only=True))
        assert sheet1_rows == [("id", "name"), ("1", "r1"), ("2", "r2")]
        assert sheet2_rows == [("id", "name"), ("3", "r3")]

    def test_c5_mismatched_columns_marks_failed(self, tmp_path):
        """C5: 列数不一致 → 任务标记 failed，不生成输出文件"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y", "z"]], header=["id", "name", "extra"])
        source_files = [
            {"filename": "a_1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}c5"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "failed"
        assert "b_2.xlsx" in job.error_message
        assert job.file_path is None

    def test_c6_cancelled_before_start(self, tmp_path):
        """C6: 启动前已 cancelling → 直接 cancelled"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        source_files = [{"filename": "a_1.xlsx", "file_path": str(f1), "size": 1}]

        username = f"{_PREFIX}c6"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files, status="cancelling")
        db.close()

        self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "cancelled"

    def test_c7_cancel_mid_run(self, tmp_path):
        """C7: 运行中检测到 cancelling → 提前终止并标记 cancelled"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(
            tmp_path / "a_1.xlsx", [["1", "x"], ["2", "y"]], header=["id", "name"],
        )
        source_files = [{"filename": "a_1.xlsx", "file_path": str(f1), "size": 1}]

        username = f"{_PREFIX}c7"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        call_count = {"n": 0}

        def _fake_is_cancelling(jid):
            call_count["n"] += 1
            # 第一次(启动检查, merged_rows=0)放行；第二次(写完第一行后)触发取消
            return call_count["n"] >= 2

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ), patch("backend.services.excel_merge_service.CANCEL_CHECK_EVERY_ROWS", 1), \
             patch("backend.services.excel_merge_service._is_cancelling", side_effect=_fake_is_cancelling):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "cancelled"
        assert job.merged_rows == 1

    def test_c8_job_not_found_exits_gracefully(self):
        """C8: job_id 不存在时协程安全退出"""
        import backend.services.excel_merge_service as svc
        self._run(svc.run_merge_job(str(uuid.uuid4())))
        # 不应抛出任何异常


# =============================================================================
# F · 坏 <dimension> 元数据导致列截断（真实线上故障场景回归）
#
# 背景：用户合并 21 个源文件后输出只有 1 列（预期 6 列）。根因是这批源文件的
# <dimension> 元数据全部谎报只有 1 列，而 openpyxl read_only 模式默认信任该
# 元数据来决定每行读几列，导致 validate_files 误判"列数一致"放行，
# 合并主循环又逐行截断成 1 列。见 excel_merge_service.py 里
# _EXCEL_MAX_COLS / read_header_and_colcount / _iter_data_rows 的修复。
# =============================================================================

class TestBadDimensionRegression:
    def _make_job(self, db, username, source_files, has_header=True):
        from backend.models.merge_excel_job import MergeExcelJob
        job = MergeExcelJob(
            user_id="uid", username=username,
            has_header=has_header, source_files=source_files,
            status="pending", total_files=len(source_files),
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)

    def _run(self, coro):
        loop = asyncio.new_event_loop()
        try:
            return loop.run_until_complete(coro)
        finally:
            loop.close()

    def test_f1_read_header_and_colcount_ignores_bad_dimension(self, tmp_path):
        """F1: read_header_and_colcount 对坏 dimension 文件仍返回真实 6 列"""
        from backend.services.excel_merge_service import read_header_and_colcount

        f = _make_bad_dimension_xlsx(
            tmp_path / "bad_dim.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["1", "x", "a", "b", "c", "d"]],
            fake_dimension="A1:A2",
        )
        header, colcount = read_header_and_colcount(str(f))
        assert colcount == 6
        assert header == ["id", "name", "c3", "c4", "c5", "c6"]

    def test_f2_validate_files_uses_real_colcount(self, tmp_path):
        """
        F2: 两个坏 dimension 文件，谎报列数碰巧一致（都是 1）但真实列数一致
        （都是 6）——修复前会用谎报值判定"一致"（表面结果一样但后续截断成 1 列），
        修复后应基于真实探测出的 6 列判定一致，且返回的 baseline_colcount 必须是 6，
        而不是被谎报值污染的 1。
        """
        from backend.services.excel_merge_service import validate_files

        f1 = _make_bad_dimension_xlsx(
            tmp_path / "a_bad1.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["1", "x", "a", "b", "c", "d"]],
            fake_dimension="A1:A2",
        )
        f2 = _make_bad_dimension_xlsx(
            tmp_path / "b_bad2.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["2", "y", "e", "f", "g", "h"]],
            fake_dimension="A1:A2",
        )
        files = [
            {"filename": "a_bad1.xlsx", "file_path": str(f1)},
            {"filename": "b_bad2.xlsx", "file_path": str(f2)},
        ]
        is_valid, err, warnings, header, colcount = validate_files(files, has_header=True)
        assert is_valid is True
        assert colcount == 6, "必须使用真实探测列数，而不是被坏 dimension 污染的谎报值"

    def test_f3_end_to_end_merge_keeps_all_columns(self, tmp_path):
        """
        F3: 端到端复现 —— 2 个坏 dimension 文件（每个声明谎报为 1 列，
        实际 6 列）跑完整 run_merge_job，输出必须是完整 6 列，且每行每列
        的值都正确落地（不是列数对但值错位）。
        """
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _make_bad_dimension_xlsx(
            tmp_path / "a_bad1.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["1", "x", "a", "b", "c", "d"]],
            fake_dimension="A1:A2",
        )
        f2 = _make_bad_dimension_xlsx(
            tmp_path / "b_bad2.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["2", "y", "e", "f", "g", "h"]],
            fake_dimension="A1:A2",
        )
        source_files = [
            {"filename": "a_bad1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_bad2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}f3"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "completed", job.error_message
        assert job.merged_rows == 2

        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        assert rows[0] == ("id", "name", "c3", "c4", "c5", "c6")
        assert rows[1] == ("1", "x", "a", "b", "c", "d")
        assert rows[2] == ("2", "y", "e", "f", "g", "h")

    def test_f4_bom_header_cleaned_end_to_end(self, tmp_path):
        """F4: 表头自带 BOM 前缀（源文件字符串内容里混入 \\ufeff）在合并结果中被清理"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(
            tmp_path / "a_1.xlsx", [["1", "x"]], header=["﻿enterprise_name", "name"],
        )
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y"]], header=["enterprise_name", "name"])
        source_files = [
            {"filename": "a_1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}f4"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "completed"

        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        assert rows[0] == ("enterprise_name", "name")

    def test_f5_mixed_good_and_bad_dimension_files(self, tmp_path):
        """
        F5: 混合场景（更贴近真实：21 个文件里只是"恰好都命中同一个上游 bug"，
        不代表所有情况都会出问题）—— 1 个 dimension 正常的文件 + 1 个
        dimension 谎报的文件，真实列数一致（6 列），应正常合并成功，不因为
        "部分正常、部分异常"而在列数比对上产生误判。
        """
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        f1 = _write_xlsx(
            tmp_path / "a_good.xlsx",
            [["1", "x", "a", "b", "c", "d"]],
            header=["id", "name", "c3", "c4", "c5", "c6"],
        )  # dimension 正常
        f2 = _make_bad_dimension_xlsx(
            tmp_path / "b_bad.xlsx",
            header=["id", "name", "c3", "c4", "c5", "c6"],
            rows=[["2", "y", "e", "f", "g", "h"]],
            fake_dimension="A1:A2",
        )  # dimension 谎报为 1 列
        source_files = [
            {"filename": "a_good.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_bad.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}f5"
        db = SessionLocal()
        job_id = self._make_job(db, username, source_files)
        db.close()

        with patch(
            "backend.config.settings.settings.allowed_directories",
            [str(tmp_path / "customer_data")],
        ):
            self._run(svc.run_merge_job(job_id))

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "completed", job.error_message

        wb = openpyxl.load_workbook(job.file_path, read_only=True, data_only=True)
        rows = list(wb["Sheet1"].iter_rows(values_only=True))
        assert rows[1] == ("1", "x", "a", "b", "c", "d")
        assert rows[2] == ("2", "y", "e", "f", "g", "h")

    def test_f6_iter_data_rows_still_streams_lazily(self, tmp_path):
        """
        F6: 性能基线 —— 确认修复没有把 read_only 流式模式退化成整份加载。
        colcount 参数应传入校验后的真实列数（本例为 2），而不是探测阶段用的
        16384 上限，否则每行会拖出上万个 None 造成体积/性能退化。
        """
        import inspect
        from backend.services.excel_merge_service import _iter_data_rows

        rows_data = [[str(i), f"name-{i}"] for i in range(5000)]
        f = _write_xlsx(tmp_path / "big.xlsx", rows_data, header=["id", "name"])

        gen = _iter_data_rows(str(f), has_header=True, colcount=2)
        assert inspect.isgenerator(gen)

        first_row = next(gen)
        assert first_row == ("0", "name-0")
        assert len(first_row) == 2, "colcount 必须是真实列数，不能是探测上限 16384"


# =============================================================================
# G · 日期格式误套到整行导致数值列显示成荒谬日期（真实线上故障场景回归）
#
# 背景：用户反馈合并结果里，日期列多出了 "00:00:00"，后面本应是数值的统计列
# （total_calls 等）全部变成了类似 "3156-01-20 00:00:00" 的乱码日期。根因是
# _run_merge_sync 原来用 ws.write_row(row, 0, values, date_format) 写一整行，
# 该调用会把 date_format（yyyy-mm-dd hh:mm:ss 的 num_format）当作 cell_format
# 套到这一行的每一个单元格，而不仅仅是日期列——Excel 会把数值列的原始整数当成
# 日期序列号来解释显示。修复为 _write_data_row 逐列按值类型选择性套用格式。
# =============================================================================

class TestDateFormatLeakRegression:
    def test_g1_write_data_row_only_formats_date_cells(self, tmp_path):
        """
        G1: 单测 _write_data_row —— 同一行里日期列和数值列混合，
        数值列写入后必须保留 General 格式（不能被套上日期 num_format），
        日期列（午夜时刻）应使用纯日期格式（不带时分秒）。
        """
        import xlsxwriter
        from backend.services.excel_merge_service import _write_data_row

        out = tmp_path / "g1.xlsx"
        wb = xlsxwriter.Workbook(str(out))
        ws = wb.add_worksheet()
        date_only_format = wb.add_format({"num_format": "yyyy-mm-dd"})
        datetime_format = wb.add_format({"num_format": "yyyy-mm-dd hh:mm:ss"})

        row = ("mx_cc_pds", datetime.datetime(2026, 6, 1), 234020, 73576)
        _write_data_row(ws, 0, row, date_only_format, datetime_format)
        wb.close()

        wb2 = openpyxl.load_workbook(out)
        ws2 = wb2.active
        cells = list(ws2[1])

        assert cells[0].value == "mx_cc_pds"
        assert cells[0].number_format == "General"

        assert cells[1].value == datetime.datetime(2026, 6, 1)
        assert cells[1].number_format == "yyyy-mm-dd"

        # 核心断言：数值列必须保留原始数值 + General 格式，不能被日期格式污染
        assert cells[2].value == 234020
        assert cells[2].number_format == "General"
        assert cells[3].value == 73576
        assert cells[3].number_format == "General"

    def test_g2_write_data_row_uses_datetime_format_for_non_midnight(self, tmp_path):
        """G2: 带具体时分秒的日期值应使用完整 datetime 格式（而非纯日期格式）"""
        import xlsxwriter
        from backend.services.excel_merge_service import _write_data_row

        out = tmp_path / "g2.xlsx"
        wb = xlsxwriter.Workbook(str(out))
        ws = wb.add_worksheet()
        date_only_format = wb.add_format({"num_format": "yyyy-mm-dd"})
        datetime_format = wb.add_format({"num_format": "yyyy-mm-dd hh:mm:ss"})

        row = (datetime.datetime(2026, 6, 1, 13, 45, 0),)
        _write_data_row(ws, 0, row, date_only_format, datetime_format)
        wb.close()

        wb2 = openpyxl.load_workbook(out)
        cell = wb2.active["A1"]
        assert cell.value == datetime.datetime(2026, 6, 1, 13, 45, 0)
        assert cell.number_format == "yyyy-mm-dd hh:mm:ss"

    def test_g3_end_to_end_merge_keeps_numeric_columns_intact(self, tmp_path):
        """
        G3: 端到端复现 —— 源文件含"日期列 + 多个数值列"（贴近用户真实数据结构：
        enterprise_name/call_date/total_calls/...），合并后数值列必须保持原始
        数值且不带日期格式；日期列必须是纯日期显示（不带 00:00:00）。
        """
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob
        import backend.services.excel_merge_service as svc

        def _write_source(path, rows):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["enterprise_name", "call_date", "total_calls", "connected_mins"])
            for r in rows:
                ws.append(r)
            wb.save(path)
            return path

        f1 = _write_source(
            tmp_path / "a_1.xlsx",
            [["mx_cc_pds", datetime.datetime(2026, 6, 1), 234020, 73576]],
        )
        f2 = _write_source(
            tmp_path / "b_2.xlsx",
            [["mx_cl_pds", datetime.datetime(2026, 6, 1), 428750, 119746]],
        )
        source_files = [
            {"filename": "a_1.xlsx", "file_path": str(f1), "size": 1},
            {"filename": "b_2.xlsx", "file_path": str(f2), "size": 1},
        ]

        username = f"{_PREFIX}g3"
        db = SessionLocal()
        job = MergeExcelJob(
            user_id="uid", username=username,
            has_header=True, source_files=source_files,
            status="pending", total_files=2,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        job_id = str(job.id)
        db.close()

        loop = asyncio.new_event_loop()
        try:
            with patch(
                "backend.config.settings.settings.allowed_directories",
                [str(tmp_path / "customer_data")],
            ):
                loop.run_until_complete(svc.run_merge_job(job_id))
        finally:
            loop.close()

        db = SessionLocal()
        job = db.query(MergeExcelJob).filter(MergeExcelJob.id == job_id).first()
        db.close()
        assert job.status == "completed", job.error_message

        wb = openpyxl.load_workbook(job.file_path)
        ws = wb["Sheet1"]
        row2 = list(ws[2])  # 首个数据行（第 1 行是表头）

        # 数值列必须还原成原始整数，且不能带日期 num_format
        assert row2[2].value == 234020
        assert "y" not in row2[2].number_format  # 不含 y/m/d 之类的日期占位符
        assert row2[3].value == 73576
        assert "y" not in row2[3].number_format

        # 日期列必须是纯日期（无 00:00:00），且值仍是正确的日期
        assert row2[1].value == datetime.datetime(2026, 6, 1)
        assert row2[1].number_format == "yyyy-mm-dd"


# =============================================================================
# D · REST API 端点
# =============================================================================

class TestMergeExcelAPI:
    @pytest.fixture
    def client(self):
        os.environ["ENABLE_AUTH"] = "False"
        from fastapi.testclient import TestClient
        sys.path.insert(0, str(Path(__file__).parent / "backend"))
        from main import app
        with TestClient(app) as c:
            yield c

    def test_d1_upload_rejects_non_excel(self, client):
        """D1: 非 xlsx/xls 文件 → 400"""
        resp = client.post(
            "/api/v1/tools/merge-excel/upload",
            files={"file": ("test.txt", b"hello", "text/plain")},
        )
        assert resp.status_code == 400

    def test_d2_upload_accepts_xlsx(self, client, tmp_path):
        """D2: 合法 xlsx 上传成功，返回 upload_id"""
        f = _write_xlsx(tmp_path / "up.xlsx", [["1", "x"]], header=["id", "name"])
        with open(f, "rb") as fp:
            resp = client.post(
                "/api/v1/tools/merge-excel/upload",
                files={"file": ("up.xlsx", fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["upload_id"]
        assert data["filename"] == "up.xlsx"
        # 清理
        try:
            Path(data["file_path"]).unlink(missing_ok=True)
        except Exception:
            pass

    def test_d3_execute_with_unknown_file_id_404(self, client):
        """D3: file_ids 指向不存在的上传 → 404"""
        resp = client.post(
            "/api/v1/tools/merge-excel/execute",
            json={"file_ids": [str(uuid.uuid4()), str(uuid.uuid4())], "has_header": True},
        )
        assert resp.status_code == 404

    def test_d4_execute_requires_at_least_two_files(self, client):
        """D4: file_ids 少于 2 个 → 422 参数校验失败"""
        resp = client.post(
            "/api/v1/tools/merge-excel/execute",
            json={"file_ids": [str(uuid.uuid4())], "has_header": True},
        )
        assert resp.status_code == 422

    def test_d5_list_jobs_200(self, client):
        """D5: GET /jobs → 200 带分页字段"""
        resp = client.get("/api/v1/tools/merge-excel/jobs")
        assert resp.status_code == 200
        data = resp.json()["data"]
        assert "total" in data
        assert "items" in data

    def test_d6_get_nonexistent_job_404(self, client):
        resp = client.get(f"/api/v1/tools/merge-excel/jobs/{uuid.uuid4()}")
        assert resp.status_code == 404

    def test_d7_cancel_nonexistent_job_404(self, client):
        resp = client.post(f"/api/v1/tools/merge-excel/jobs/{uuid.uuid4()}/cancel")
        assert resp.status_code == 404

    def test_d8_delete_nonexistent_job_404(self, client):
        resp = client.delete(f"/api/v1/tools/merge-excel/jobs/{uuid.uuid4()}")
        assert resp.status_code == 404

    def test_d9_download_nonexistent_job_404(self, client):
        resp = client.get(f"/api/v1/tools/merge-excel/jobs/{uuid.uuid4()}/download")
        assert resp.status_code == 404

    def test_d10_download_uncompleted_job_400(self, client):
        """D10: 未完成任务下载 → 400"""
        from backend.config.database import SessionLocal
        from backend.models.merge_excel_job import MergeExcelJob

        db = SessionLocal()
        job = MergeExcelJob(
            user_id="uid", username=f"{_PREFIX}d10",
            has_header=True, source_files=[], status="pending",
        )
        db.add(job)
        db.commit()
        job_id = str(job.id)
        db.close()

        resp = client.get(f"/api/v1/tools/merge-excel/jobs/{job_id}/download")
        assert resp.status_code == 400

    def test_d11_end_to_end_upload_execute_download(self, client, tmp_path):
        """D11: 端到端 — 上传 2 文件 → 提交合并 → 轮询完成 → 下载校验内容"""
        import time

        f1 = _write_xlsx(tmp_path / "a_1.xlsx", [["1", "x"]], header=["id", "name"])
        f2 = _write_xlsx(tmp_path / "b_2.xlsx", [["2", "y"]], header=["id", "name"])

        file_ids = []
        for f in (f1, f2):
            with open(f, "rb") as fp:
                resp = client.post(
                    "/api/v1/tools/merge-excel/upload",
                    files={"file": (f.name, fp, "application/octet-stream")},
                )
            assert resp.status_code == 200
            file_ids.append(resp.json()["data"]["upload_id"])

        resp = client.post(
            "/api/v1/tools/merge-excel/execute",
            json={"file_ids": file_ids, "has_header": True, "job_name": f"{_PREFIX}e2e"},
        )
        assert resp.status_code == 200
        job_id = resp.json()["data"]["job_id"]

        job = None
        for _ in range(50):
            resp = client.get(f"/api/v1/tools/merge-excel/jobs/{job_id}")
            job = resp.json()["data"]
            if job["status"] in ("completed", "failed"):
                break
            time.sleep(0.1)

        assert job["status"] == "completed", job.get("error_message")
        assert job["merged_rows"] == 2

        resp = client.get(f"/api/v1/tools/merge-excel/jobs/{job_id}/download")
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        resp = client.delete(f"/api/v1/tools/merge-excel/jobs/{job_id}")
        assert resp.status_code == 200


# =============================================================================
# E · RBAC 权限种子
# =============================================================================

class TestRBACSeed:
    def test_e1_permission_exists(self):
        """E1: tools:merge_excel 权限已入库"""
        from backend.config.database import SessionLocal
        from backend.models.permission import Permission

        db = SessionLocal()
        try:
            perm = (
                db.query(Permission)
                .filter(Permission.resource == "tools", Permission.action == "merge_excel")
                .first()
            )
            assert perm is not None
        finally:
            db.close()

    def test_e2_assigned_to_superadmin(self):
        """E2: tools:merge_excel 已分配给 superadmin 角色"""
        from backend.config.database import SessionLocal
        from backend.models.permission import Permission
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission

        db = SessionLocal()
        try:
            perm = (
                db.query(Permission)
                .filter(Permission.resource == "tools", Permission.action == "merge_excel")
                .first()
            )
            role = db.query(Role).filter(Role.name == "superadmin").first()
            assert perm is not None and role is not None
            rp = (
                db.query(RolePermission)
                .filter(RolePermission.role_id == role.id, RolePermission.permission_id == perm.id)
                .first()
            )
            assert rp is not None
        finally:
            db.close()
