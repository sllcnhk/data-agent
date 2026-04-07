"""
test_export_e2e_comprehensive.py — Code 160 分批提取功能综合端到端测试

作为资深测试工程师，针对 ClickHouse max_execution_time 限制绕过方案进行全面测试：

  F (4)  — count_rows 在 chunked 重试链路中的异常处理
           F1: Code 160 触发 → count_rows 正常 → chunked 成功（完整链路）
           F2: Code 160 触发 → count_rows 抛 RuntimeError → 任务 failed，错误信息记录
           F3: Code 160 触发 → count_rows 抛 TimeoutError → 任务 failed，错误信息记录
           F4: Code 160 触发 → count_rows 返回 0 → 任务 completed（空结果集）

  G (3)  — chunked 模式下 DB 进度字段的准确性
           G1: total_rows / total_batches 在 chunked 开始前写入 DB
           G2: done_batches 随批次推进（按 SQL 分批，非行批）
           G3: completed 时 done_batches == total_sql_chunks（精准闭合）

  H (2)  — chunked 模式下取消任务
           H1: chunked 首批开始前检测到 cancelling → 任务 cancelled，部分文件被保存
           H2: 非 chunked（stream）模式取消行为不受 chunked 改动影响

  I (3)  — SQL 安全：LIMIT/OFFSET 包裹前 SQL 清洗
           I1: SQL 末尾有分号 → 被去除，chunk_sql 正确生成
           I2: SQL 末尾有空格和分号混合 → 同样正确清洗
           I3: count_rows 的 count_sql 也正确去除尾部分号

  J (2)  — 配置项绑定与默认值验证
           J1: export_query_max_execution_time 默认 300，可从 env 读取
           J2: export_chunk_size / export_auto_chunk_threshold 默认值正确

  K (3)  — _parse_tsv_cell：TabSeparated 单元格解析
           K1: \\N 解析为 None（SQL NULL）
           K2: 转义序列正确还原（\\t \\n \\\\ \\r）
           K3: 普通字符串直接返回，无副作用

  RBAC (1) — 新功能不引入未授权菜单
           R1: chunked 提取为服务层变更，AppLayout.tsx data-export 菜单仍仅需 data:export

共计: 17 个测试用例

运行：
    /d/ProgramData/Anaconda3/envs/dataagent/python.exe -m pytest test_export_e2e_comprehensive.py -v -s
"""
from __future__ import annotations

import asyncio
import os
import sys
import tempfile
import unittest
import uuid
from pathlib import Path
from unittest.mock import Mock, patch

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(1, os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_e2ec_{uuid.uuid4().hex[:6]}_"


# ─────────────────────────────────────────────────────────────────────────────
# 共用工具
# ─────────────────────────────────────────────────────────────────────────────

def _db():
    from backend.config.database import SessionLocal
    return SessionLocal()


_g_db = _db()


def _make_job(db, suffix="", status="pending"):
    from backend.models.export_job import ExportJob
    job = ExportJob(
        user_id="test-uid",
        username=f"{_PREFIX}{suffix}",
        query_sql="SELECT id, name FROM t",
        connection_env="test",
        connection_type="clickhouse",
        status=status,
        output_filename=f"{_PREFIX}{suffix}.xlsx",
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def teardown_module(_=None):
    from backend.models.export_job import ExportJob
    try:
        _g_db.query(ExportJob).filter(
            ExportJob.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        _g_db.commit()
    finally:
        _g_db.close()


# ─────────────────────────────────────────────────────────────────────────────
# 核心 helper：运行 run_export_job（可控 stream / chunked / count_rows 行为）
# ─────────────────────────────────────────────────────────────────────────────

def _run_export(
    job_id: str,
    output_path: str,
    stream_effect,          # Exception 或 批次列表（如 [[row,...], ...]）
    count_rows_effect=3,    # int 或 Exception
    chunked_effect=None,    # 批次列表，默认 [[(1,"a"),(2,"b"),(3,"c")]]
    cancelling_after=None,  # None 或 批次索引（该批次开始前将任务设为 cancelling）
):
    """
    用 mock ClickHouseExportClient 运行 run_export_job。

    stream_effect:
      - 如果是 Exception，stream_batches 第一次迭代时 raise
      - 否则作为批次列表按序 yield

    count_rows_effect:
      - 如果是 Exception，count_rows 时 raise
      - 否则返回该整数

    chunked_effect:
      - chunked 模式下 stream_batches_chunked yield 的批次列表
    """
    from backend.services.data_export_service import run_export_job
    from backend.services.export_clients.clickhouse import ClickHouseExportClient

    fake_cols = [
        type("C", (), {"name": "id", "type": "Int32"})(),
        type("C", (), {"name": "name", "type": "String"})(),
    ]
    _chunked_effect = chunked_effect if chunked_effect is not None else [[(1, "a"), (2, "b"), (3, "c")]]

    batch_call_count = [0]

    def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
        def _gen():
            if isinstance(stream_effect, Exception):
                raise stream_effect
            yield from stream_effect
        return _gen()

    def fake_count_rows(sql, timeout=300):
        if isinstance(count_rows_effect, Exception):
            raise count_rows_effect
        return count_rows_effect

    def fake_stream_batches_chunked(sql, chunk_size, total_rows, batch_size=50000, extra_settings=None):
        for i, batch in enumerate(_chunked_effect):
            if cancelling_after is not None and i >= cancelling_after:
                # 模拟任务被外部设为 cancelling
                _set_cancelling(job_id)
            yield batch

    with patch.object(ClickHouseExportClient, "get_columns", return_value=fake_cols), \
         patch.object(ClickHouseExportClient, "stream_batches", side_effect=fake_stream_batches), \
         patch.object(ClickHouseExportClient, "count_rows", side_effect=fake_count_rows), \
         patch.object(ClickHouseExportClient, "stream_batches_chunked", side_effect=fake_stream_batches_chunked), \
         patch("backend.services.data_export_service._build_export_client") as mock_build:

        mock_build.return_value = ClickHouseExportClient("localhost", 8123, "default", "", "test")
        asyncio.run(run_export_job(job_id, {
            "query_sql": "SELECT id, name FROM t",
            "connection_env": "test",
            "connection_type": "clickhouse",
            "batch_size": 50000,
            "output_path": output_path,
            "output_filename": "test.xlsx",
        }))


def _set_cancelling(job_id: str):
    """将任务状态置为 cancelling（模拟外部取消请求）"""
    from backend.models.export_job import ExportJob
    db = _db()
    try:
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        if j and j.status == "running":
            j.status = "cancelling"
            db.commit()
    finally:
        db.close()


# ═════════════════════════════════════════════════════════════════════════════
# Section F — count_rows 在 chunked 重试链路中的异常处理
# ═════════════════════════════════════════════════════════════════════════════

class TestCountRowsInChunkedRetryChain(unittest.TestCase):
    """F1-F4: Code 160 触发后 count_rows 各种结果对任务状态的影响"""

    @classmethod
    def setUpClass(cls):
        cls.tmp_dir = tempfile.mkdtemp()

    def _load_job(self, job_id):
        from backend.models.export_job import ExportJob
        db = _db()
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        return j

    def test_F1_code160_then_count_rows_ok_then_chunked_completes(self):
        """F1: Code 160 → count_rows 成功 → chunked 导出 → completed"""
        job = _make_job(_g_db, "f1")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}f1.xlsx")

        code160 = RuntimeError(
            "ClickHouse 错误 500: Code: 160, e.displayText() = DB::Exception: "
            "Estimated query execution time (65.0 seconds) is too long. Maximum: 60."
        )
        _run_export(str(job.id), out, stream_effect=code160, count_rows_effect=3)

        j = self._load_job(job.id)
        self.assertEqual(j.status, "completed",
                         f"F1 应 completed，实际: {j.status} | {j.error_message}")
        self.assertTrue(Path(out).exists(), "F1 Excel 文件应存在")

    def test_F2_code160_then_count_rows_runtime_error_marks_failed(self):
        """F2: Code 160 触发 chunked → count_rows 抛 RuntimeError → failed，错误信息包含描述"""
        job = _make_job(_g_db, "f2")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}f2.xlsx")

        code160 = RuntimeError("Code: 160, Estimated query execution time too long.")
        cnt_err = RuntimeError("count 查询超时：内存溢出")

        _run_export(str(job.id), out, stream_effect=code160, count_rows_effect=cnt_err)

        j = self._load_job(job.id)
        self.assertEqual(j.status, "failed",
                         f"F2 应 failed，实际: {j.status}")
        self.assertIsNotNone(j.error_message, "F2 error_message 不应为空")
        self.assertIn("预扫描", j.error_message,
                      f"F2 错误信息应包含预扫描说明，实际: {j.error_message}")
        self.assertFalse(Path(out).exists(), "F2 失败后文件不应存在")

    def test_F3_code160_then_count_rows_timeout_marks_failed(self):
        """F3: Code 160 触发 chunked → count_rows 抛 TimeoutError → failed"""
        job = _make_job(_g_db, "f3")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}f3.xlsx")

        code160 = RuntimeError("ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED")
        cnt_timeout = TimeoutError("count_rows 超时 (>300s)")

        _run_export(str(job.id), out, stream_effect=code160, count_rows_effect=cnt_timeout)

        j = self._load_job(job.id)
        self.assertEqual(j.status, "failed", f"F3 应 failed，实际: {j.status}")
        self.assertIsNotNone(j.error_message)

    def test_F4_code160_then_count_rows_zero_completes_empty(self):
        """F4: Code 160 → count_rows=0（空结果）→ completed，exported_rows=0"""
        import openpyxl
        job = _make_job(_g_db, "f4")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}f4.xlsx")

        code160 = RuntimeError("Code: 160, Estimated query execution time too long.")

        _run_export(
            str(job.id), out,
            stream_effect=code160,
            count_rows_effect=0,
            chunked_effect=[],   # total_rows=0 → stream_batches_chunked 不产生任何批次
        )

        j = self._load_job(job.id)
        self.assertEqual(j.status, "completed",
                         f"F4 空结果集应 completed，实际: {j.status} | {j.error_message}")
        self.assertEqual(j.exported_rows, 0, "F4 exported_rows 应为 0")
        self.assertTrue(Path(out).exists(), "F4 应生成（含仅表头的）Excel")

        # 验证 Excel 仅含表头行
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        self.assertEqual(ws.max_row, 1, f"F4 仅表头，应为 1 行，实际: {ws.max_row}")


# ═════════════════════════════════════════════════════════════════════════════
# Section G — chunked 模式下 DB 进度字段准确性
# ═════════════════════════════════════════════════════════════════════════════

class TestChunkedProgressTracking(unittest.TestCase):
    """G1-G3: total_rows / total_batches / done_batches 在 DB 中的准确性"""

    @classmethod
    def setUpClass(cls):
        cls.tmp_dir = tempfile.mkdtemp()

    def _load_job(self, job_id):
        from backend.models.export_job import ExportJob
        db = _db()
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        return j

    def test_G1_total_rows_and_total_batches_written_to_db(self):
        """G1: chunked 模式启动后 DB 记录 total_rows 和 total_batches"""
        from backend.config.settings import settings as app_settings

        job = _make_job(_g_db, "g1")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}g1.xlsx")

        code160 = RuntimeError("Code: 160, time too long.")
        # count_rows 返回 500（chunk_size=200000，应分 1 个 SQL chunk）
        total_rows = 500
        expected_chunks = 1  # ceil(500 / 200000) = 1

        _run_export(
            str(job.id), out,
            stream_effect=code160,
            count_rows_effect=total_rows,
            chunked_effect=[[(i, f"v{i}") for i in range(total_rows)]],
        )

        j = self._load_job(job.id)
        self.assertEqual(j.status, "completed",
                         f"G1 应 completed，实际: {j.status}")
        self.assertEqual(j.total_rows, total_rows,
                         f"G1 total_rows 应为 {total_rows}，实际: {j.total_rows}")
        self.assertEqual(j.total_batches, expected_chunks,
                         f"G1 total_batches 应为 {expected_chunks}，实际: {j.total_batches}")

    def test_G2_done_batches_maps_to_sql_chunks_not_row_batches(self):
        """G2: chunked 模式下 done_batches 按 SQL 分批计，而非行批次计"""
        from backend.config.settings import settings as app_settings
        import math

        job = _make_job(_g_db, "g2")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}g2.xlsx")

        code160 = RuntimeError("Code: 160, time too long.")
        total_rows = 1000
        chunk_size = app_settings.export_chunk_size  # 默认 200000
        expected_sql_chunks = max(1, math.ceil(total_rows / chunk_size))

        _run_export(
            str(job.id), out,
            stream_effect=code160,
            count_rows_effect=total_rows,
            chunked_effect=[[(i, f"r{i}") for i in range(total_rows)]],
        )

        j = self._load_job(job.id)
        self.assertEqual(j.status, "completed", f"G2 应 completed，实际: {j.status}")
        # 完成时 done_batches == total_batches（SQL 分批数）
        self.assertEqual(j.done_batches, j.total_batches,
                         f"G2 完成时 done_batches({j.done_batches}) 应等于 total_batches({j.total_batches})")

    def test_G3_completed_job_done_batches_equals_total_sql_chunks(self):
        """G3: 完成时 done_batches == total_sql_chunks，进度条精准闭合到 100%

        chunk_size=200000（默认）：count_rows=600000 → 3 个 SQL 分批。
        chunked_effect 提供 3 批各 10 行，completed 时 done_batches 应等于 3。
        """
        import math
        from backend.config.settings import settings as app_settings

        job = _make_job(_g_db, "g3")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}g3.xlsx")

        chunk_size = app_settings.export_chunk_size     # 200000
        total_rows = chunk_size * 3                      # 600000 → 3 chunks
        expected_chunks = 3

        code160 = RuntimeError(
            "Code: 160, Estimated query execution time too long. Maximum: 60."
        )
        # 3 批，每批 10 行（用 _run_export 的 chunked_effect）
        chunked_batches = [
            [(i * 10 + j, f"v{j}") for j in range(10)]
            for i in range(expected_chunks)
        ]

        _run_export(
            str(job.id), out,
            stream_effect=code160,
            count_rows_effect=total_rows,
            chunked_effect=chunked_batches,
        )

        j = self._load_job(job.id)
        self.assertEqual(j.status, "completed", f"G3 应 completed，实际: {j.status}")
        self.assertEqual(j.done_batches, expected_chunks,
                         f"G3 done_batches 应等于 total_sql_chunks={expected_chunks}，实际: {j.done_batches}")
        self.assertEqual(j.done_batches, j.total_batches,
                         "G3 done_batches 必须等于 total_batches（进度 100%）")


# ═════════════════════════════════════════════════════════════════════════════
# Section H — chunked 模式下取消任务
# ═════════════════════════════════════════════════════════════════════════════

class TestCancelDuringChunked(unittest.TestCase):
    """H1-H2: 取消行为在 chunked / stream 两种模式下均正确"""

    @classmethod
    def setUpClass(cls):
        cls.tmp_dir = tempfile.mkdtemp()

    def _load_job(self, job_id):
        from backend.models.export_job import ExportJob
        db = _db()
        j = db.query(ExportJob).filter(ExportJob.id == job_id).first()
        db.close()
        return j

    def test_H1_cancel_detected_during_chunked_export(self):
        """H1: chunked 模式中检测到 cancelling → 任务 cancelled，保留已导出行"""
        job = _make_job(_g_db, "h1")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}h1.xlsx")

        code160 = RuntimeError("Code: 160, time too long.")

        # cancelling_after=1 → 第 2 批开始前将任务设为 cancelling
        # chunked_effect 有 2 批，第 1 批成功，第 2 批前检测到取消
        _run_export(
            str(job.id), out,
            stream_effect=code160,
            count_rows_effect=6,
            chunked_effect=[
                [(1, "a"), (2, "b"), (3, "c")],   # 第 1 批（正常处理）
                [(4, "d"), (5, "e"), (6, "f")],   # 第 2 批（取消前）
            ],
            cancelling_after=1,  # 第 2 批前设为 cancelling
        )

        j = self._load_job(job.id)
        self.assertEqual(j.status, "cancelled",
                         f"H1 应 cancelled，实际: {j.status} | {j.error_message}")

    def test_H2_cancel_in_stream_mode_not_affected_by_chunked_changes(self):
        """H2: stream 模式（非 chunked）取消行为不受 chunked 改动影响"""
        from backend.services.data_export_service import run_export_job
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        job = _make_job(_g_db, "h2", status="pending")
        out = os.path.join(self.tmp_dir, f"{_PREFIX}h2.xlsx")

        fake_cols = [
            type("C", (), {"name": "id", "type": "Int32"})(),
            type("C", (), {"name": "val", "type": "String"})(),
        ]

        cancelling_set = [False]

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            """第 1 批正常，yield 前将任务设为 cancelling，第 2 批检测到取消"""
            def _gen():
                yield [(1, "x"), (2, "y")]
                # 第 1 批 yield 后设置取消
                _set_cancelling(str(job.id))
                yield [(3, "z")]
            return _gen()

        with patch.object(ClickHouseExportClient, "get_columns", return_value=fake_cols), \
             patch.object(ClickHouseExportClient, "stream_batches", side_effect=fake_stream_batches), \
             patch("backend.services.data_export_service._build_export_client") as mock_build:

            mock_build.return_value = ClickHouseExportClient("localhost", 8123, "default", "", "test")
            asyncio.run(run_export_job(str(job.id), {
                "query_sql": "SELECT id, val FROM t",
                "connection_env": "test",
                "connection_type": "clickhouse",
                "batch_size": 2,
                "output_path": out,
                "output_filename": "test.xlsx",
            }))

        j = self._load_job(job.id)
        self.assertEqual(j.status, "cancelled",
                         f"H2 stream 模式取消应 cancelled，实际: {j.status}")


# ═════════════════════════════════════════════════════════════════════════════
# Section I — SQL 安全：LIMIT/OFFSET 包裹前的 SQL 清洗
# ═════════════════════════════════════════════════════════════════════════════

class TestSQLSafetyInChunkedWrappers(unittest.TestCase):
    """I1-I3: SQL 清洗保证 LIMIT/OFFSET 包裹不引入语法错误"""

    def _make_client(self):
        from backend.services.export_clients.clickhouse import ClickHouseExportClient
        return ClickHouseExportClient(
            host="localhost", port=8123,
            user="default", password="", database="test",
        )

    def test_I1_sql_with_trailing_semicolon_stripped_in_chunk_sql(self):
        """I1: SQL 末尾有分号 → 被去除，chunk_sql LIMIT/OFFSET 语法正确"""
        client = self._make_client()
        emitted_sqls = []

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            emitted_sqls.append(sql)
            return iter([])

        client.stream_batches = fake_stream_batches

        list(client.stream_batches_chunked(
            "SELECT id, name FROM users;",
            chunk_size=100,
            total_rows=100,
        ))

        self.assertEqual(len(emitted_sqls), 1)
        chunk_sql = emitted_sqls[0]

        # 分号不应出现在 chunk_sql 的子查询中（导致语法错误）
        # 期望格式：SELECT * FROM (SELECT id, name FROM users) AS _chunk_0 LIMIT 100 OFFSET 0
        self.assertNotIn(";;", chunk_sql, "I1 chunk_sql 不应含双分号")
        self.assertIn("LIMIT 100 OFFSET 0", chunk_sql, "I1 LIMIT/OFFSET 结构应正确")
        # 子查询括号应闭合在分号前
        inner = chunk_sql[chunk_sql.index("(") + 1:chunk_sql.rindex(")")]
        self.assertFalse(inner.rstrip().endswith(";"),
                         f"I1 子查询内不应有分号，实际内层 SQL: {inner!r}")

    def test_I2_sql_with_trailing_spaces_and_semicolons_cleaned(self):
        """I2: SQL 末尾含多个空格和分号混合 → 正确去除"""
        client = self._make_client()
        emitted_sqls = []

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            emitted_sqls.append(sql)
            return iter([])

        client.stream_batches = fake_stream_batches

        # 末尾为 ";   " 或 "  ; " 等混合形式
        list(client.stream_batches_chunked(
            "SELECT a, b FROM tbl  ;  ",
            chunk_size=50,
            total_rows=50,
        ))

        self.assertEqual(len(emitted_sqls), 1)
        inner = emitted_sqls[0][emitted_sqls[0].index("(") + 1:emitted_sqls[0].rindex(")")]
        self.assertFalse(inner.rstrip().endswith(";"),
                         f"I2 子查询内不应有分号: {inner!r}")

    def test_I3_count_rows_sql_also_strips_trailing_semicolon(self):
        """I3: count_rows 内部拼接的 count_sql 也正确去除分号"""
        client = self._make_client()
        captured_sqls = []

        mock_resp = Mock(status_code=200, text="42\n")

        def capture_post(*args, **kwargs):
            captured_sqls.append(kwargs.get("data", b"").decode("utf-8"))
            return mock_resp

        with patch("requests.post", side_effect=capture_post):
            result = client.count_rows("SELECT * FROM big_table;  ")

        self.assertEqual(result, 42)
        self.assertEqual(len(captured_sqls), 1)
        count_sql = captured_sqls[0]

        # count_sql 格式：SELECT count() FROM (SELECT * FROM big_table) AS _cnt_q
        self.assertIn("SELECT count()", count_sql, "I3 应含 count()")
        # 子查询内不应有分号
        if "(" in count_sql and ")" in count_sql:
            inner = count_sql[count_sql.index("(") + 1:count_sql.rindex(")")]
            self.assertFalse(inner.rstrip().endswith(";"),
                             f"I3 count_sql 子查询不应有分号: {inner!r}")


# ═════════════════════════════════════════════════════════════════════════════
# Section J — 配置项绑定与默认值
# ═════════════════════════════════════════════════════════════════════════════

class TestSettingsBindings(unittest.TestCase):
    """J1-J2: 新增配置项默认值正确，可从环境变量绑定"""

    def test_J1_export_query_max_execution_time_default_300(self):
        """J1: export_query_max_execution_time 默认值为 300（5 分钟）"""
        from backend.config.settings import settings
        self.assertEqual(
            settings.export_query_max_execution_time, 300,
            f"J1 默认值应为 300，实际: {settings.export_query_max_execution_time}"
        )

    def test_J2_export_chunk_size_and_threshold_defaults(self):
        """J2: export_chunk_size=200000，export_auto_chunk_threshold=500000"""
        from backend.config.settings import settings
        self.assertEqual(
            settings.export_chunk_size, 200_000,
            f"J2 export_chunk_size 默认应为 200000，实际: {settings.export_chunk_size}"
        )
        self.assertEqual(
            settings.export_auto_chunk_threshold, 500_000,
            f"J2 export_auto_chunk_threshold 默认应为 500000，实际: {settings.export_auto_chunk_threshold}"
        )


# ═════════════════════════════════════════════════════════════════════════════
# Section K — _parse_tsv_cell：TabSeparated 单元格解析
# ═════════════════════════════════════════════════════════════════════════════

class TestParseTsvCell(unittest.TestCase):
    """K1-K3: TSV 单元格转义与 NULL 还原"""

    def setUp(self):
        from backend.services.export_clients.clickhouse import _parse_tsv_cell
        self.parse = _parse_tsv_cell

    def test_K1_backslash_N_returns_none(self):
        r"""K1: \N（ClickHouse TabSeparated NULL 表示）→ Python None"""
        result = self.parse(r"\N")
        self.assertIsNone(result, r"K1 \N 应解析为 None")

    def test_K2_escape_sequences_correctly_unescaped(self):
        r"""K2: \t \n \r \\ 各自还原为真实字符"""
        cases = [
            (r"hello\tworld", "hello\tworld"),    # \t → tab
            (r"line1\nline2", "line1\nline2"),     # \n → newline
            (r"cr\rtest", "cr\rtest"),              # \r → carriage return
            (r"back\\slash", "back\\slash"),        # \\ → single backslash
        ]
        for raw, expected in cases:
            with self.subTest(raw=raw):
                result = self.parse(raw)
                self.assertEqual(result, expected,
                                 f"K2 {raw!r} 应还原为 {expected!r}，实际: {result!r}")

    def test_K3_plain_string_returned_unchanged(self):
        """K3: 普通字符串（无转义）直接返回，无副作用"""
        cases = [
            "hello",
            "12345",
            "2024-01-01",
            "",     # 空字符串
            "中文内容",
        ]
        for s in cases:
            with self.subTest(s=s):
                result = self.parse(s)
                self.assertEqual(result, s, f"K3 {s!r} 应原样返回，实际: {result!r}")


# ═════════════════════════════════════════════════════════════════════════════
# Section R — RBAC 菜单与权限范围验证
# ═════════════════════════════════════════════════════════════════════════════

class TestRBACMenuScope(unittest.TestCase):
    """R1: chunked 提取为服务层变更，不应引入新菜单或新权限"""

    def test_R1_data_export_menu_still_controlled_by_data_export_perm(self):
        """R1: AppLayout.tsx data-export 菜单项权限字段仍为 data:export（未变化）"""
        layout_path = Path(__file__).parent / "frontend/src/components/AppLayout.tsx"
        self.assertTrue(layout_path.exists(),
                        f"R1 AppLayout.tsx 不存在: {layout_path}")

        content = layout_path.read_text(encoding="utf-8")

        # 验证 data-export 路由存在
        self.assertIn("/data-export", content,
                      "R1 AppLayout.tsx 中 /data-export 菜单项不存在")

        # 验证权限字段为 data:export
        self.assertIn("data:export", content,
                      "R1 data-export 菜单项未配置 data:export 权限")

        # 验证不存在任何 chunked 相关的新菜单（chunked 是服务层透明实现）
        self.assertNotIn("chunked", content.lower(),
                         "R1 AppLayout.tsx 不应出现 chunked 相关菜单")
        self.assertNotIn("分批", content,
                         "R1 AppLayout.tsx 不应出现分批相关菜单")


# ══════════════════════════════════════════════════════════════════════════════
# 运行入口
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    unittest.main(verbosity=2)
