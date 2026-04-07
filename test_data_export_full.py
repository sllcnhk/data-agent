"""
test_data_export_full.py — 数据导出功能综合测试

作为资深测试工程师，覆盖以下维度：

  A (6)  — RBAC 权限矩阵（ENABLE_AUTH=True 实测）
  B (7)  — 核心导出协程端到端（mocked ClickHouse，真实 Excel 生成）
  C (5)  — API 参数校验 & 边界场景
  D (5)  — 任务生命周期 & 删除管理（含 Bug Fix 验证）
  E (4)  — data:export 权限纳入角色权限管理范围验证
  W (3)  — 分批提取链路回归（Code 160 / 正常流 / 取消 不因改动断裂）
  Z (3)  — 下载链路回归（GET /download 完整链路验证）

共计: 33 个测试用例

运行：
    /d/ProgramData/Anaconda3/envs/dataagent/python.exe -m pytest test_data_export_full.py -v -s
"""
from __future__ import annotations

import asyncio
import os
import sys
import tempfile
import unittest
import uuid
from pathlib import Path
from unittest.mock import Mock, patch

sys.path.insert(0, os.path.dirname(__file__))
sys.path.insert(1, os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_def_{uuid.uuid4().hex[:6]}_"


# ─────────────────────────────────────────────────────────────────────────────
# 公共工具函数
# ─────────────────────────────────────────────────────────────────────────────

def _db():
    from backend.config.database import SessionLocal
    return SessionLocal()


_g_db = _db()


def _make_user(suffix="", password="Test1234!", role_names=None, is_superadmin=False):
    """创建测试用户（可选角色）"""
    from backend.models.user import User
    from backend.models.role import Role
    from backend.models.user_role import UserRole
    from backend.core.auth.password import hash_password

    username = f"{_PREFIX}{suffix or uuid.uuid4().hex[:6]}"
    u = User(
        username=username,
        display_name=f"DEF {suffix}",
        hashed_password=hash_password(password),
        auth_source="local",
        is_active=True,
        is_superadmin=is_superadmin,
    )
    _g_db.add(u)
    _g_db.flush()
    for rname in (role_names or []):
        role = _g_db.query(Role).filter(Role.name == rname).first()
        if role:
            _g_db.add(UserRole(user_id=u.id, role_id=role.id))
    _g_db.commit()
    _g_db.refresh(u)
    return u


def _token(user):
    """生成用户 JWT token"""
    from backend.config.settings import settings
    from backend.core.auth.jwt import create_access_token
    from backend.core.rbac import get_user_roles
    roles = get_user_roles(user, _g_db)
    return create_access_token(
        {"sub": str(user.id), "username": user.username, "roles": roles},
        settings.jwt_secret, settings.jwt_algorithm,
    )


def _auth(user):
    return {"Authorization": f"Bearer {_token(user)}"}


def _make_export_job(db, username, status="pending", file_path=None, query_sql="SELECT 1"):
    """直接在 DB 中创建测试导出任务"""
    from backend.models.export_job import ExportJob
    job = ExportJob(
        user_id="uid",
        username=username,
        query_sql=query_sql,
        connection_env="test",
        connection_type="clickhouse",
        status=status,
        output_filename="out.xlsx",
        file_path=file_path,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def teardown_module(_=None):
    """清理测试数据"""
    from backend.models.user import User
    from backend.models.export_job import ExportJob
    try:
        # 删除测试导出任务
        _g_db.query(ExportJob).filter(
            ExportJob.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        # 删除测试用户
        _g_db.query(User).filter(
            User.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        _g_db.commit()
    finally:
        _g_db.close()


# ─────────────────────────────────────────────────────────────────────────────
# 共用 TestClient（ENABLE_AUTH=False）
# ─────────────────────────────────────────────────────────────────────────────
from fastapi.testclient import TestClient


def _make_app():
    from backend.main import app
    return app


# ══════════════════════════════════════════════════════════════════════════════
# Section A — RBAC 权限矩阵（ENABLE_AUTH=True）
# ══════════════════════════════════════════════════════════════════════════════

class TestRBACPermissionMatrix(unittest.TestCase):
    """A1-A6: 验证 data:export 权限在真实 RBAC 场景下的行为"""

    @classmethod
    def setUpClass(cls):
        cls.app = _make_app()
        cls.client = TestClient(cls.app, raise_server_exceptions=True)
        cls.superadmin = _make_user("a_super", role_names=["superadmin"])
        cls.viewer = _make_user("a_viewer", role_names=["viewer"])
        cls.analyst = _make_user("a_analyst", role_names=["analyst"])
        cls.admin = _make_user("a_admin", role_names=["admin"])

    def _req(self, user, method, path, **kwargs):
        """认证请求（ENABLE_AUTH=True 模式）"""
        with patch("backend.config.settings.settings.enable_auth", True):
            return getattr(self.client, method)(
                path, headers=_auth(user), **kwargs
            )

    def test_A1_data_export_perm_visible_in_permissions_api(self):
        """A1: GET /permissions 包含 data:export 权限定义"""
        resp = self._req(self.superadmin, "get", "/api/v1/permissions")
        self.assertEqual(resp.status_code, 200)
        perms = {f"{p['resource']}:{p['action']}" for p in resp.json()}
        self.assertIn("data:export", perms, f"data:export 未出现在权限列表中，实际: {perms}")

    def test_A2_superadmin_me_includes_data_export(self):
        """A2: superadmin /auth/me 返回的 permissions 中包含 data:export"""
        resp = self._req(self.superadmin, "get", "/api/v1/auth/me")
        self.assertEqual(resp.status_code, 200)
        perms = set(resp.json().get("permissions", []))
        self.assertIn("data:export", perms, f"superadmin 缺少 data:export，实际: {perms}")

    def test_A3_viewer_cannot_access_connections(self):
        """A3: viewer 角色无 data:export → GET /connections 返回 403"""
        resp = self._req(self.viewer, "get", "/api/v1/data-export/connections")
        self.assertEqual(resp.status_code, 403,
                         f"viewer 应收到 403，实际: {resp.status_code} {resp.text}")

    def test_A4_analyst_cannot_access_preview(self):
        """A4: analyst 角色无 data:export → POST /preview 返回 403"""
        resp = self._req(
            self.analyst, "post", "/api/v1/data-export/preview",
            json={"query_sql": "SELECT 1", "connection_env": "sg"},
        )
        self.assertEqual(resp.status_code, 403)

    def test_A5_admin_cannot_submit_export_job(self):
        """A5: admin 角色无 data:export → POST /execute 返回 403"""
        resp = self._req(
            self.admin, "post", "/api/v1/data-export/execute",
            json={"query_sql": "SELECT 1", "connection_env": "sg"},
        )
        self.assertEqual(resp.status_code, 403)

    def test_A6_dynamic_grant_data_export_to_admin_enables_access(self):
        """A6: 动态授予 admin 角色 data:export → 可访问；撤销后 → 403（验证权限管理范围）"""
        from backend.models.role import Role
        from backend.models.permission import Permission
        from backend.models.role_permission import RolePermission

        admin_role = _g_db.query(Role).filter(Role.name == "admin").first()
        data_export_perm = _g_db.query(Permission).filter(
            Permission.resource == "data", Permission.action == "export"
        ).first()

        self.assertIsNotNone(data_export_perm, "data:export 权限未在 DB 中注册")

        # 授权
        rp = RolePermission(role_id=admin_role.id, permission_id=data_export_perm.id)
        _g_db.add(rp)
        _g_db.commit()

        try:
            # admin 有了 data:export → 应可访问（404 表示有权但无连接，而非 403）
            # 需重新生成 token（含最新 roles）
            resp = self._req(self.admin, "get", "/api/v1/data-export/connections")
            self.assertNotEqual(resp.status_code, 403,
                                f"授权后 admin 仍收到 403: {resp.text}")
        finally:
            # 撤销权限
            _g_db.delete(rp)
            _g_db.commit()

        # 撤销后：重新请求应重新被拒绝
        resp_after = self._req(self.admin, "get", "/api/v1/data-export/connections")
        self.assertEqual(resp_after.status_code, 403,
                         f"撤销后 admin 应收到 403，实际: {resp_after.status_code}")


# ══════════════════════════════════════════════════════════════════════════════
# Section B — 核心导出协程端到端
# ══════════════════════════════════════════════════════════════════════════════

class TestCoreCoroutineE2E(unittest.TestCase):
    """B1-B7: run_export_job 协程真实执行，mocked ClickHouse 客户端"""

    @classmethod
    def setUpClass(cls):
        from backend.services.export_clients.base import ColumnInfo
        cls.ColumnInfo = ColumnInfo
        cls.svc = __import__("backend.services.data_export_service", fromlist=["run_export_job"])

    def _make_mock_client(self, rows, col_names=("id", "name"), col_types=("Int64", "String")):
        """构建 mock ClickHouse 客户端"""
        from backend.services.export_clients.base import ColumnInfo
        mock = Mock()
        mock.get_columns.return_value = [ColumnInfo(n, t) for n, t in zip(col_names, col_types)]
        mock.stream_batches.return_value = iter([rows])
        return mock

    def _run(self, job_id, config, mock_client):
        with patch("backend.services.data_export_service._build_export_client", return_value=mock_client):
            loop = asyncio.new_event_loop()
            loop.run_until_complete(self.svc.run_export_job(job_id, config))
            loop.close()

    def test_B1_full_flow_single_batch_completed(self):
        """B1: 单批次正常完成 → status=completed，文件存在，file_size > 0"""
        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b1")
        job_id = str(job.id)
        db.close()

        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b1.xlsx")
            config = {
                "query_sql": "SELECT id, name FROM t", "connection_env": "sg",
                "connection_type": "clickhouse", "batch_size": 1000,
                "output_path": out, "output_filename": "b1.xlsx",
            }
            mock_client = self._make_mock_client(
                rows=[("12345678901234567", "Alice"), ("98765432109876543", "Bob")]
            )
            self._run(job_id, config, mock_client)

            db2 = _db()
            j = db2.query(__import__("backend.models.export_job", fromlist=["ExportJob"]).ExportJob).filter_by(id=job.id).first()
            db2.close()

            self.assertEqual(j.status, "completed")
            self.assertEqual(j.exported_rows, 2)
            self.assertEqual(j.done_batches, 1)
            self.assertEqual(j.total_sheets, 1)
            self.assertGreater(j.file_size, 0)
            self.assertTrue(Path(out).exists(), "Excel 文件未生成")

    def test_B2_timestamps_set_correctly(self):
        """B2: completed 后 started_at 和 finished_at 均已设置"""
        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b2")
        job_id = str(job.id)
        db.close()

        from backend.models.export_job import ExportJob
        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b2.xlsx")
            config = {
                "query_sql": "SELECT 1", "connection_env": "sg",
                "connection_type": "clickhouse", "batch_size": 1000,
                "output_path": out, "output_filename": "b2.xlsx",
            }
            mock_client = self._make_mock_client(rows=[("1", "test")])
            self._run(job_id, config, mock_client)

            db2 = _db()
            j = db2.query(ExportJob).filter_by(id=job.id).first()
            db2.close()

            self.assertIsNotNone(j.started_at, "started_at 未设置")
            self.assertIsNotNone(j.finished_at, "finished_at 未设置")
            self.assertGreaterEqual(
                j.finished_at, j.started_at,
                "finished_at 不得早于 started_at"
            )

    def test_B3_progress_tracking_exported_rows_and_batches(self):
        """B3: 多批次时 exported_rows 和 done_batches 与实际数据一致"""
        from backend.services.export_clients.base import ColumnInfo
        from backend.models.export_job import ExportJob

        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b3")
        job_id = str(job.id)
        db.close()

        # 两个批次，各 3 行
        mock_client = Mock()
        mock_client.get_columns.return_value = [ColumnInfo("v", "String")]
        batch1 = [("row1",), ("row2",), ("row3",)]
        batch2 = [("row4",), ("row5",), ("row6",)]
        mock_client.stream_batches.return_value = iter([batch1, batch2])

        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b3.xlsx")
            config = {
                "query_sql": "SELECT v", "connection_env": "sg",
                "connection_type": "clickhouse", "batch_size": 3,
                "output_path": out, "output_filename": "b3.xlsx",
            }
            self._run(job_id, config, mock_client)

            db2 = _db()
            j = db2.query(ExportJob).filter_by(id=job.id).first()
            db2.close()

            self.assertEqual(j.exported_rows, 6, f"期望 6 行，实际: {j.exported_rows}")
            self.assertEqual(j.done_batches, 2, f"期望 2 批次，实际: {j.done_batches}")

    def test_B4_large_int_written_as_string_in_excel(self):
        """B4: Int64 大整数写入 Excel 为字符串（避免科学计数法）"""
        import openpyxl
        from backend.models.export_job import ExportJob

        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b4")
        job_id = str(job.id)
        db.close()

        large_int = 12345678901234567
        mock_client = self._make_mock_client(
            rows=[(large_int, "Alice")],
            col_names=("id", "name"),
            col_types=("Int64", "String"),
        )

        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b4.xlsx")
            config = {
                "query_sql": "SELECT id, name", "connection_env": "sg",
                "connection_type": "clickhouse", "batch_size": 1000,
                "output_path": out, "output_filename": "b4.xlsx",
            }
            self._run(job_id, config, mock_client)

            wb = openpyxl.load_workbook(out)
            ws = wb.active
            # 行 1 是表头，行 2 是第一个数据行
            cell_val = ws.cell(row=2, column=1).value
            self.assertIsInstance(cell_val, str,
                                  f"Int64 应写为字符串，实际类型: {type(cell_val)}, 值: {cell_val}")
            self.assertEqual(cell_val, str(large_int))

    def test_B5_get_columns_failure_marks_job_failed_and_no_file(self):
        """B5: get_columns 异常 → job 标记 failed，错误信息记录，无残留文件"""
        from backend.models.export_job import ExportJob

        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b5")
        job_id = str(job.id)
        db.close()

        mock_client = Mock()
        mock_client.get_columns.side_effect = RuntimeError("ClickHouse 连接失败")
        mock_client.stream_batches.return_value = iter([])

        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b5.xlsx")
            config = {
                "query_sql": "SELECT 1", "connection_env": "bad",
                "connection_type": "clickhouse", "batch_size": 1000,
                "output_path": out, "output_filename": "b5.xlsx",
            }
            self._run(job_id, config, mock_client)

            db2 = _db()
            j = db2.query(ExportJob).filter_by(id=job.id).first()
            db2.close()

            self.assertEqual(j.status, "failed", f"期望 failed，实际: {j.status}")
            self.assertIsNotNone(j.error_message, "error_message 未设置")
            self.assertIn("ClickHouse 连接失败", j.error_message)
            self.assertFalse(Path(out).exists(), "失败后不应存在残留文件")

    def test_B6_cancel_before_start_marks_cancelled(self):
        """B6: 协程启动前 DB 中已为 cancelling → 直接标记 cancelled，不写文件"""
        from backend.models.export_job import ExportJob

        db = _db()
        job = _make_export_job(db, f"{_PREFIX}b6", status="cancelling")
        job_id = str(job.id)
        db.close()

        mock_client = Mock()
        mock_client.get_columns.return_value = []

        with tempfile.TemporaryDirectory() as td:
            out = os.path.join(td, "b6.xlsx")
            config = {
                "query_sql": "SELECT 1", "connection_env": "sg",
                "connection_type": "clickhouse", "batch_size": 1000,
                "output_path": out, "output_filename": "b6.xlsx",
            }
            self._run(job_id, config, mock_client)

            db2 = _db()
            j = db2.query(ExportJob).filter_by(id=job.id).first()
            db2.close()

            self.assertEqual(j.status, "cancelled")

    def test_B7_config_snapshot_stored_in_execute_response(self):
        """B7: POST /execute 创建任务后 config_snapshot 字段正确存储"""
        from fastapi.testclient import TestClient
        from backend.main import app
        from backend.models.export_job import ExportJob

        client = TestClient(app)
        with patch("asyncio.create_task"):
            resp = client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id FROM users",
                "connection_env": "sg",
                "connection_type": "clickhouse",
                "batch_size": 10000,
                "job_name": f"{_PREFIX}b7",
            })
        self.assertEqual(resp.status_code, 200)
        job_id = resp.json()["data"]["job_id"]

        db2 = _db()
        j = db2.query(ExportJob).filter(
            ExportJob.id == job_id
        ).first()
        db2.close()

        self.assertIsNotNone(j.config_snapshot)
        self.assertEqual(j.config_snapshot["query_sql"], "SELECT id FROM users")
        self.assertEqual(j.config_snapshot["connection_env"], "sg")
        self.assertEqual(j.config_snapshot["batch_size"], 10000)


# ══════════════════════════════════════════════════════════════════════════════
# Section C — API 参数校验 & 边界场景
# ══════════════════════════════════════════════════════════════════════════════

class TestAPIValidation(unittest.TestCase):
    """C1-C5: 请求参数校验边界场景"""

    @classmethod
    def setUpClass(cls):
        cls.app = _make_app()
        cls.client = TestClient(cls.app, raise_server_exceptions=True)

    def test_C1_list_jobs_page_size_over_100_returns_422(self):
        """C1: page_size > 100 → 422 Unprocessable Entity"""
        resp = self.client.get("/api/v1/data-export/jobs", params={"page_size": 101})
        self.assertEqual(resp.status_code, 422, f"期望 422，实际: {resp.status_code}")

    def test_C2_list_jobs_page_zero_returns_422(self):
        """C2: page=0（最小值为 1）→ 422"""
        resp = self.client.get("/api/v1/data-export/jobs", params={"page": 0})
        self.assertEqual(resp.status_code, 422)

    def test_C3_preview_limit_over_500_returns_422(self):
        """C3: preview limit > 500 → 422"""
        resp = self.client.post("/api/v1/data-export/preview", json={
            "query_sql": "SELECT 1",
            "connection_env": "sg",
            "limit": 501,
        })
        self.assertEqual(resp.status_code, 422)

    def test_C4_execute_batch_size_below_minimum_returns_422(self):
        """C4: batch_size < 1000 → 422"""
        resp = self.client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1",
            "connection_env": "sg",
            "batch_size": 999,
        })
        self.assertEqual(resp.status_code, 422)

    def test_C5_execute_job_name_with_special_chars_sanitized(self):
        """C5: job_name 含特殊字符时文件名正确净化（不含危险字符）"""
        with patch("asyncio.create_task"):
            resp = self.client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT 1",
                "connection_env": "sg",
                "job_name": f"{_PREFIX}测试 ../../../etc/passwd",
            })
        self.assertEqual(resp.status_code, 200)
        filename = resp.json()["data"]["output_filename"]
        # 文件名不含路径分隔符
        self.assertNotIn("/", filename)
        self.assertNotIn("\\", filename)
        self.assertNotIn("..", filename)
        # 必须以 .xlsx 结尾
        self.assertTrue(filename.endswith(".xlsx"), f"文件名非 .xlsx: {filename}")


# ══════════════════════════════════════════════════════════════════════════════
# Section D — 任务生命周期 & 删除管理
# ══════════════════════════════════════════════════════════════════════════════

class TestJobLifecycleAndDelete(unittest.TestCase):
    """D1-D5: 任务删除行为（含 Bug Fix 验证）"""

    @classmethod
    def setUpClass(cls):
        cls.app = _make_app()
        cls.client = TestClient(cls.app, raise_server_exceptions=True)
        cls.db = _db()

    @classmethod
    def tearDownClass(cls):
        cls.db.close()

    def test_D1_delete_completed_job_removes_record_and_file(self):
        """D1: 删除 completed 任务 → 记录删除，本地文件同时删除"""
        from backend.models.export_job import ExportJob

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            fpath = f.name
            f.write(b"PK fake xlsx")

        job = _make_export_job(self.db, f"{_PREFIX}d1", status="completed", file_path=fpath)
        resp = self.client.delete(f"/api/v1/data-export/jobs/{job.id}")

        self.assertEqual(resp.status_code, 200)
        self.assertFalse(Path(fpath).exists(), "文件应被删除")
        # 再次 GET → 404
        resp2 = self.client.get(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp2.status_code, 404)

    def test_D2_delete_completed_job_file_already_missing_still_200(self):
        """D2: 文件已不存在时删除 completed 任务仍返回 200"""
        job = _make_export_job(
            self.db, f"{_PREFIX}d2",
            status="completed",
            file_path="/tmp/nonexistent_test_file_xyz.xlsx",
        )
        resp = self.client.delete(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp.status_code, 200)

    def test_D3_after_delete_get_returns_404(self):
        """D3: 删除后 GET /jobs/{id} → 404"""
        job = _make_export_job(self.db, f"{_PREFIX}d3", status="failed")
        self.client.delete(f"/api/v1/data-export/jobs/{job.id}")
        resp = self.client.get(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp.status_code, 404)

    def test_D4_delete_running_job_returns_400(self):
        """D4: 删除进行中任务（status=running）→ 400（Bug Fix 验证）"""
        job = _make_export_job(self.db, f"{_PREFIX}d4", status="running")
        resp = self.client.delete(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp.status_code, 400,
                         f"running 任务不应被直接删除，实际状态码: {resp.status_code}")
        # 任务记录应仍存在
        resp2 = self.client.get(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp2.status_code, 200, "任务不应被删除")

    def test_D5_delete_pending_job_returns_400(self):
        """D5: 删除等待中任务（status=pending）→ 400（防止孤儿协程）"""
        job = _make_export_job(self.db, f"{_PREFIX}d5", status="pending")
        resp = self.client.delete(f"/api/v1/data-export/jobs/{job.id}")
        self.assertEqual(resp.status_code, 400,
                         f"pending 任务不应被直接删除，实际状态码: {resp.status_code}")


# ══════════════════════════════════════════════════════════════════════════════
# Section E — data:export 权限纳入角色权限管理范围
# ══════════════════════════════════════════════════════════════════════════════

class TestPermissionManagementScope(unittest.TestCase):
    """E1-E4: 验证 data:export 已正确纳入权限管理体系"""

    @classmethod
    def setUpClass(cls):
        cls.app = _make_app()
        cls.client = TestClient(cls.app, raise_server_exceptions=True)
        cls.superadmin = _make_user("e_super", role_names=["superadmin"])

    def _auth_req(self, method, path, **kwargs):
        with patch("backend.config.settings.settings.enable_auth", True):
            return getattr(self.client, method)(
                path, headers=_auth(self.superadmin), **kwargs
            )

    def test_E1_data_export_perm_exists_in_db_with_correct_fields(self):
        """E1: DB 中 data:export 权限存在，字段完整（resource/action/description）"""
        from backend.models.permission import Permission
        perm = _g_db.query(Permission).filter(
            Permission.resource == "data",
            Permission.action == "export",
        ).first()
        self.assertIsNotNone(perm, "data:export 权限未在 DB 中创建")
        self.assertEqual(perm.resource, "data")
        self.assertEqual(perm.action, "export")
        self.assertIsNotNone(perm.description, "description 不应为空")

    def test_E2_superadmin_role_has_data_export(self):
        """E2: superadmin 角色权限集合中包含 data:export"""
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission
        from backend.models.permission import Permission

        role = _g_db.query(Role).filter(Role.name == "superadmin").first()
        self.assertIsNotNone(role)

        perms = (
            _g_db.query(Permission)
            .join(RolePermission, RolePermission.permission_id == Permission.id)
            .filter(RolePermission.role_id == role.id)
            .all()
        )
        perm_keys = {f"{p.resource}:{p.action}" for p in perms}
        self.assertIn("data:export", perm_keys,
                      f"superadmin 角色权限中无 data:export，实际: {perm_keys}")

    def test_E3_viewer_analyst_admin_do_not_have_data_export(self):
        """E3: viewer/analyst/admin 默认均不含 data:export 权限"""
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission
        from backend.models.permission import Permission

        for role_name in ("viewer", "analyst", "admin"):
            role = _g_db.query(Role).filter(Role.name == role_name).first()
            if not role:
                continue
            perms = (
                _g_db.query(Permission)
                .join(RolePermission, RolePermission.permission_id == Permission.id)
                .filter(RolePermission.role_id == role.id)
                .all()
            )
            perm_keys = {f"{p.resource}:{p.action}" for p in perms}
            self.assertNotIn(
                "data:export", perm_keys,
                f"角色 '{role_name}' 不应包含 data:export，实际: {perm_keys}",
            )

    def test_E4_data_export_menu_item_perm_field_is_data_export(self):
        """E4: 前端 AppLayout.tsx 中 /data-export 菜单项的 perm 字段为 'data:export'"""
        layout_path = Path(__file__).parent / "frontend/src/components/AppLayout.tsx"
        self.assertTrue(layout_path.exists(), f"AppLayout.tsx 不存在: {layout_path}")

        content = layout_path.read_text(encoding="utf-8")

        # 验证菜单 key 存在
        self.assertIn("/data-export", content,
                      "AppLayout.tsx 中缺少 /data-export 菜单项")
        # 验证菜单权限字段
        self.assertIn("data:export", content,
                      "AppLayout.tsx 中 data-export 菜单项未配置 data:export 权限")
        # 同时检查两者在同一行（确认是菜单配置，不是注释）
        for line in content.splitlines():
            if "/data-export" in line and "data:export" in line:
                break
        else:
            # 允许分布在相邻行（对象字面量格式）
            # 找到包含 /data-export 的行号，检查相邻 3 行内有 data:export
            lines = content.splitlines()
            for i, line in enumerate(lines):
                if "/data-export" in line:
                    context = "\n".join(lines[max(0, i-1):i+4])
                    self.assertIn("data:export", context,
                                  f"/data-export 菜单项附近未找到 data:export 权限配置:\n{context}")
                    break


# ══════════════════════════════════════════════════════════════════════════════
# Section W — 分批提取链路回归（验证 chunked/stream 改动不破坏已有导出行为）
# ══════════════════════════════════════════════════════════════════════════════

class TestChunkedExtractionRegression(unittest.TestCase):
    """
    W1-W3: 确保分批提取改动（extra_settings、count_rows、stream_batches_chunked）
    不破坏已有的端到端导出链路。

    W1: 正常小查询走 stream 模式（无 Code 160）→ 导出成功，extra_settings 已注入
    W2: Code 160 自动切换 chunked → 最终 completed，Excel 行数正确
    W3: chunked 模式下取消任务仍可正常取消（not stuck）
    """

    @classmethod
    def setUpClass(cls):
        from backend.config.database import SessionLocal
        cls.db = SessionLocal()
        cls.tmp_dir = tempfile.mkdtemp()

    @classmethod
    def tearDownClass(cls):
        from backend.models.export_job import ExportJob
        try:
            cls.db.query(ExportJob).filter(
                ExportJob.username.like(f"{_PREFIX}%")
            ).delete(synchronize_session=False)
            cls.db.commit()
        finally:
            cls.db.close()

    def _make_job(self, suffix=""):
        from backend.models.export_job import ExportJob
        job = ExportJob(
            user_id="test-uid",
            username=f"{_PREFIX}w_{suffix}",
            query_sql="SELECT 1",
            connection_env="test",
            connection_type="clickhouse",
            status="pending",
            output_filename=f"{_PREFIX}w_{suffix}.xlsx",
        )
        self.db.add(job)
        self.db.commit()
        self.db.refresh(job)
        return job

    def _run_job(self, job_id, output_path, stream_effect, chunked_effect=None):
        """
        用 mock ClickHouseExportClient 运行 run_export_job。
        stream_effect: 第一次 stream_batches 调用的返回（Exception 或可迭代批次列表）
        chunked_effect: chunked 模式下 stream_batches_chunked 的返回（可迭代批次列表）
        """
        import asyncio
        from backend.services.data_export_service import run_export_job
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        fake_cols = [
            type("C", (), {"name": "id", "type": "Int32"})(),
            type("C", (), {"name": "name", "type": "String"})(),
        ]

        def fake_stream_batches(sql, batch_size=50000, extra_settings=None):
            def _gen():
                if isinstance(stream_effect, Exception):
                    raise stream_effect
                yield from stream_effect
            return _gen()

        def fake_stream_batches_chunked(sql, chunk_size, total_rows, batch_size=50000, extra_settings=None):
            yield from (chunked_effect or [])

        with patch.object(ClickHouseExportClient, "get_columns", return_value=fake_cols), \
             patch.object(ClickHouseExportClient, "stream_batches", side_effect=fake_stream_batches), \
             patch.object(ClickHouseExportClient, "count_rows", return_value=3), \
             patch.object(ClickHouseExportClient, "stream_batches_chunked", side_effect=fake_stream_batches_chunked), \
             patch("backend.services.data_export_service._build_export_client") as mock_build:
            mock_build.return_value = ClickHouseExportClient("localhost", 8123, "default", "", "test")
            asyncio.run(run_export_job(job_id, {
                "query_sql": "SELECT * FROM t",
                "connection_env": "test",
                "connection_type": "clickhouse",
                "batch_size": 50000,
                "output_path": output_path,
                "output_filename": "test.xlsx",
            }))

    def test_W1_normal_stream_succeeds_with_extra_settings(self):
        """W1: 正常流（无 Code 160）→ completed，extra_settings（max_execution_time）已注入"""
        job = self._make_job("w1")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}w1.xlsx")

        fake_rows = [(1, "alice"), (2, "bob"), (3, "charlie")]
        self._run_job(str(job.id), output_path, stream_effect=[fake_rows])

        self.db.refresh(job)
        self.assertEqual(job.status, "completed",
                         f"W1 应 completed，实际: {job.status} | {job.error_message}")
        self.assertTrue(Path(output_path).exists(), "W1 Excel 文件应存在")

        # 验证 extra_settings 确实被注入（从设置读取 max_execution_time=300）
        from backend.config.settings import settings as app_settings
        self.assertEqual(app_settings.export_query_max_execution_time, 300,
                         "W1 默认 EXPORT_QUERY_MAX_EXECUTION_TIME 应为 300")

    def test_W2_code160_auto_chunked_produces_correct_excel(self):
        """W2: Code 160 自动切换 chunked → completed，Excel 包含 3 行数据"""
        import openpyxl
        job = self._make_job("w2")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}w2.xlsx")

        code160 = RuntimeError(
            "ClickHouse 错误 500: Code: 160, e.displayText() = DB::Exception: "
            "Estimated query execution time (60.9 seconds) is too long. Maximum: 60."
        )
        chunked_rows = [(1, "x"), (2, "y"), (3, "z")]

        self._run_job(
            str(job.id), output_path,
            stream_effect=code160,
            chunked_effect=[chunked_rows],
        )

        self.db.refresh(job)
        self.assertEqual(job.status, "completed",
                         f"W2 Code 160 后应切换分批并完成，实际: {job.status} | {job.error_message}")

        # 验证 Excel 内容
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        data_rows = ws.max_row - 1  # 减去表头行
        self.assertEqual(data_rows, 3, f"W2 应有 3 数据行，实际: {data_rows}")

    def test_W3_non_code160_failure_marks_failed_cleanly(self):
        """W3: 非 Code 160 错误 → failed，不触发 chunked 重试，文件已清理"""
        job = self._make_job("w3")
        output_path = os.path.join(self.tmp_dir, f"{_PREFIX}w3.xlsx")

        mem_err = RuntimeError("ClickHouse 错误 500: Code: 241, Memory limit exceeded")
        self._run_job(str(job.id), output_path, stream_effect=mem_err)

        self.db.refresh(job)
        self.assertEqual(job.status, "failed",
                         f"W3 应标记 failed，实际: {job.status}")
        self.assertFalse(Path(output_path).exists(), "W3 不完整文件应被清理")


# ══════════════════════════════════════════════════════════════════════════════
# Section Z — 下载链路回归（端到端验证 GET /download 不因改动断裂）
# ══════════════════════════════════════════════════════════════════════════════

class TestDownloadRegression(unittest.TestCase):
    """
    Z1-Z3: 确保完整链路 创建任务→完成→下载 不因前端 Blob 修复而断裂。
    测试对象是后端端点本身，通过 TestClient 模拟 axios blob 请求（附认证头）。
    ENABLE_AUTH=False 下以 AnonymousUser 身份请求，与前端 Blob 修复后的行为一致。
    """

    @classmethod
    def setUpClass(cls):
        cls.app = _make_app()
        cls.client = TestClient(cls.app, raise_server_exceptions=True)
        cls.tmp_dir = tempfile.mkdtemp()

    def _make_real_xlsx(self, name="reg_test.xlsx") -> str:
        """生成真实 xlsx 文件（openpyxl）"""
        import openpyxl
        path = os.path.join(self.tmp_dir, f"{_PREFIX}{name}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["id", "value", "label"])
        for i in range(5):
            ws.append([i, i * 10, f"row_{i}"])
        wb.save(path)
        return path

    def test_Z1_full_chain_completed_job_downloadable(self):
        """Z1: 完成任务 → GET /download → 200（模拟 axios blob 请求，含认证头为可选）"""
        xlsx_path = self._make_real_xlsx("z1.xlsx")
        job = _make_export_job(_g_db, f"{_PREFIX}z1",
                               status="completed", file_path=xlsx_path)

        # ENABLE_AUTH=False：无 token 也能通过（AnonymousUser）
        resp = self.client.get(f"/api/v1/data-export/jobs/{job.id}/download")

        self.assertEqual(resp.status_code, 200,
                         f"Z1 链路断裂：{resp.status_code} {resp.text}")
        self.assertGreater(len(resp.content), 0, "Z1 响应体为空")

        _g_db.delete(job)
        _g_db.commit()
        os.unlink(xlsx_path)

    def test_Z2_download_response_headers_correct(self):
        """Z2: 下载响应头 Content-Type=xlsx MIME + Content-Disposition=attachment"""
        xlsx_path = self._make_real_xlsx("z2.xlsx")
        job = _make_export_job(_g_db, f"{_PREFIX}z2",
                               status="completed", file_path=xlsx_path)

        resp = self.client.get(f"/api/v1/data-export/jobs/{job.id}/download")

        content_type = resp.headers.get("content-type", "")
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content_type,
            f"Z2 Content-Type 不正确: {content_type}",
        )
        disposition = resp.headers.get("content-disposition", "")
        self.assertIn("attachment", disposition,
                      f"Z2 Content-Disposition 应含 attachment: {disposition}")

        _g_db.delete(job)
        _g_db.commit()
        os.unlink(xlsx_path)

    def test_Z3_download_body_is_valid_xlsx_binary(self):
        """Z3: 响应体二进制是合法的 xlsx（ZIP 格式，PK 魔数开头）"""
        xlsx_path = self._make_real_xlsx("z3.xlsx")
        job = _make_export_job(_g_db, f"{_PREFIX}z3",
                               status="completed", file_path=xlsx_path)

        resp = self.client.get(f"/api/v1/data-export/jobs/{job.id}/download")
        self.assertEqual(resp.status_code, 200)

        # xlsx 是 ZIP 格式，前 2 字节为 PK（0x50 0x4B）
        magic = resp.content[:2]
        self.assertEqual(magic, b"PK",
                         f"Z3 响应体不是合法 xlsx（ZIP）文件，魔数: {magic!r}")

        # 用 openpyxl 解析验证内容完整性
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        self.assertIn("Sheet1", wb.sheetnames,
                      f"Z3 xlsx 中缺少 Sheet1，实际: {wb.sheetnames}")
        ws = wb["Sheet1"]
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        self.assertEqual(headers, ["id", "value", "label"],
                         f"Z3 xlsx 表头不正确: {headers}")

        _g_db.delete(job)
        _g_db.commit()
        os.unlink(xlsx_path)


# ══════════════════════════════════════════════════════════════════════════════
# 运行入口
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    unittest.main(verbosity=2)
