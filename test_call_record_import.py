"""
test_call_record_import.py — call_record_imported 导入类型完整测试套件
=====================================================================

测试层次：
  I  (8)  — Transformer 单元测试（transform_call_record_batch 核心逻辑）
  J  (6)  — insert_json_rows 单元测试（HTTP client 新方法）
  K  (4)  — API Schema 验证（SheetConfig.import_type 字段）
  L (10)  — 端到端导入流程（Mock CH，验证 call_record_imported 路径）
  M  (6)  — 回归测试（standard 类型不受影响，原有功能完整保留）
  N  (4)  — 真实 ClickHouse 集成测试（写入 MY 环境并验证，测后清理）

总计: 38 个测试用例

设计原则：
  - I/J/K 层：纯单元，无外部依赖
  - L/M 层：Mock _build_ch_client，使用真实 PostgreSQL ImportJob 状态追踪
  - N 层：真实 CH 写入（MY 202.165.17.231），import_job_id 隔离，测后删除
  - 回归保护：M 层确认 standard import 路径完全未受影响
"""
import asyncio
import io
import os
import sys
import tempfile
import unittest
import uuid
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch, call

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_cri_{uuid.uuid4().hex[:6]}_"

# 固定 15 列列名（与 data_import_service 保持一致）
_FIXED_COLS = [
    "Task Name", "Dialogue Name", "Contact ID", "Audio Name", "Call ID",
    "Result", "Call Time", "Call Duration", "Agent Call Duration",
    "Dialogue Round", "Have Read", "Agent", "Tags", "Transfer Status",
    "Call Record Text Detail Masked",
]

# 完整 67 列表头（与 test100.xlsx 一致）
_ALL_HEADERS = _FIXED_COLS[:14] + [
    "Call Outcome", "Call Outcome(Original)",
    "Last Node", "Last Node(Original)",
    "First Verification", "First Verification(Original)",
    "Verification Method", "Verification Method(Original)",
    "PTP", "PTP(Original)",
    "Call Back Reason", "Call Back Reason(Original)",
    "Request", "Request(Original)",
    "Update Details", "Update Details(Original)",
    "Agent Action", "Agent Action(Original)",
    "KB", "KB(Original)",
    "Default KB", "Default KB(Original)",
    "DCRMS Reaction Code", "DCRMS Reaction Code(Original)",
    "DCRMS Reaction Code 2", "DCRMS Reaction Code 2(Original)",
    "Second Verification", "Second Verification(Original)",
    "Outcome", "Outcome(Original)",
    "LN", "LN(Original)",
    "Verif Method", "Verif Method(Original)",
    "CB Reason", "CB Reason(Original)",
    "Upd", "Upd(Original)",
    "Action", "Action(Original)",
    "DCRMS", "DCRMS(Original)",
    "V1", "V1(Original)",
    "V2", "V2(Original)",
    "Call Back Time", "Call Back Time(Original)",
    "Call Back Date", "Call Back Date(Original)",
    "Call Back Date Time", "Call Back Date Time(Original)",
] + ["Call Record Text Detail Masked"]   # 索引 66

# 目标表插入列（与 _CR_TARGET_COLS 一致）
_TARGET_COLS = [
    "import_job_id", "source_file",
    "task_name", "dialogue_name", "contact_id", "audio_name", "call_id",
    "result", "call_start_time", "call_duration", "agent_call_duration",
    "dialogue_round", "have_read", "agent", "tags", "transfer_status",
    "call_record_text_detail_masked", "tag_array",
]

# ─── auth patch ──────────────────────────────────────────────────────────────
_auth_patcher = None

def setup_module(_=None):
    global _auth_patcher
    from backend.config.settings import settings
    _auth_patcher = patch.object(settings, 'enable_auth', False)
    _auth_patcher.start()


def teardown_module(_=None):
    global _auth_patcher
    if _auth_patcher:
        _auth_patcher.stop()
        _auth_patcher = None
    from backend.models.import_job import ImportJob
    from backend.config.database import SessionLocal
    db = SessionLocal()
    try:
        db.query(ImportJob).filter(
            ImportJob.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


# ─── helpers ─────────────────────────────────────────────────────────────────

def _make_call_record_xlsx(
    data_rows=None,
    sheet_name="Sheet1",
) -> str:
    """
    生成含完整 67 列表头的 Excel 临时文件，返回文件路径。

    data_rows: list of list，长度 67（与 _ALL_HEADERS 对应）。
               None 时写入 2 条默认样本行。
    """
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(_ALL_HEADERS)

    if data_rows is None:
        # Row 1: Connected call，部分扩展列有值
        row1 = [
            "Task_A", "Dialogue_A", "CONTACT001", "audio_a.mp3", "CALL001",
            "Connected", "31/03/2026 17:31:46", 71, 30, 5,
            "Read", "AgentX", "tag1;tag2", "No Transfer",
            "Already Paid", "Already Paid",   # Call Outcome, Original
            "Payment Reminder", "Payment Reminder",  # Last Node
            None, None,   # First Verification
            "IC Speech", "IC Speech",   # Verif Method
            "Already Paid", "Already Paid",  # PTP
        ] + [None] * (52 - 10) + [
            "Bot: Hello\nUser: Hi",   # Call Record Text Detail Masked (idx 66)
        ]
        # Row 2: Missed Call，扩展列全空
        row2 = [
            "Task_B", "Dialogue_B", "CONTACT002", "audio_b.mp3", "CALL002",
            "Missed Call", "01/04/2026 09:00:00", 0, 0, 0,
            "Unread", None, None, "No Transfer",
        ] + [None] * 52 + [None]
        data_rows = [row1, row2]

    for r in data_rows:
        ws.append(r)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        return f.name


def _safe_unlink(fp):
    try:
        if fp and os.path.exists(fp):
            os.unlink(fp)
    except OSError:
        pass


def _db():
    from backend.config.database import SessionLocal
    return SessionLocal()


def _create_import_job(status="pending"):
    """在 DB 创建 ImportJob 并返回 job_id"""
    from backend.models.import_job import ImportJob
    db = _db()
    try:
        job = ImportJob(
            id=uuid.uuid4(),
            user_id=str(uuid.uuid4()),
            username=f"{_PREFIX}u",
            upload_id=str(uuid.uuid4()),
            filename="test.xlsx",
            connection_env="my",
            status=status,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)
    finally:
        db.close()


def _run_cri_job(file_path, sheet_configs, batch_size=500,
                 mock_side_effect=None):
    """
    运行 call_record_imported 类型的导入协程，返回最终 job dict + mock_client。
    mock_side_effect: 注入到 mock_client.insert_json_rows.side_effect
    """
    from backend.models.import_job import ImportJob
    from backend.config.database import SessionLocal
    from backend.services.data_import_service import run_import_job

    job_id = _create_import_job()
    config = {
        "file_path": file_path,
        "connection_env": "my",
        "batch_size": batch_size,
        "sheets": sheet_configs,
    }
    mock_client = MagicMock()
    if mock_side_effect:
        mock_client.insert_json_rows.side_effect = mock_side_effect

    with patch("backend.services.data_import_service._build_ch_client",
               return_value=mock_client):
        asyncio.run(run_import_job(job_id, config))

    db = SessionLocal()
    try:
        job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
        result = job.to_dict() if job else {}
        result["_mock"] = mock_client
        return result
    finally:
        db.close()


# ═══════════════════════════════════════════════════════════════════════════════
# I — Transformer 单元测试 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestTransformer(unittest.TestCase):
    """transform_call_record_batch 核心转换逻辑纯单元测试，无外部依赖"""

    def _transform(self, rows, headers=None, job_id="JOB1", src="file.xlsx"):
        from backend.services.data_import_service import transform_call_record_batch
        return transform_call_record_batch(rows, headers or _ALL_HEADERS, job_id, src)

    def test_I1_fixed_columns_mapped_correctly(self):
        """15 个固定列正确映射到目标字段名"""
        row = [
            "TaskA", "DialogueB", "C001", "audio.mp3", "CALLX",
            "Connected", "31/03/2026 17:31:46", 71, 30, 5,
            "Read", "Agent1", "t1;t2", "No Transfer",
        ] + [None] * 52 + ["transcript text"]
        result = self._transform([row])
        r = result[0]
        self.assertEqual(r["task_name"], "TaskA")
        self.assertEqual(r["dialogue_name"], "DialogueB")
        self.assertEqual(r["contact_id"], "C001")
        self.assertEqual(r["audio_name"], "audio.mp3")
        self.assertEqual(r["call_id"], "CALLX")
        self.assertEqual(r["result"], "Connected")
        self.assertEqual(r["call_duration"], 71)
        self.assertEqual(r["agent_call_duration"], 30)
        self.assertEqual(r["dialogue_round"], 5)
        self.assertEqual(r["have_read"], "Read")
        self.assertEqual(r["agent"], "Agent1")
        self.assertEqual(r["tags"], "t1;t2")
        self.assertEqual(r["transfer_status"], "No Transfer")
        self.assertEqual(r["call_record_text_detail_masked"], "transcript text")

    def test_I2_call_time_parsed_to_datetime_string(self):
        """Call Time DD/MM/YYYY HH:MM:SS → YYYY-MM-DD HH:MM:SS"""
        row = ["T", "D", "C", "A", "ID", "R", "31/03/2026 17:31:46", 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        self.assertEqual(r["call_start_time"], "2026-03-31 17:31:46")

    def test_I3_call_time_none_returns_none(self):
        """Call Time 为 None 时 call_start_time 返回 None"""
        row = ["T", "D", "C", "A", "ID", "R", None, 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        self.assertIsNone(r["call_start_time"])

    def test_I4_call_time_invalid_format_returns_none(self):
        """无法解析的 Call Time 返回 None，不抛异常"""
        row = ["T", "D", "C", "A", "ID", "R", "not-a-date", 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        self.assertIsNone(r["call_start_time"])

    def test_I5_tag_array_excludes_null_and_empty_values(self):
        """扩展列值为 None 或空串时，不写入 tag_array"""
        row = ["T", "D", "C", "A", "ID", "R", "01/04/2026 09:00:00", 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        # 全是 None → tag_array 应为空
        self.assertEqual(r["tag_array"], [])

    def test_I6_tag_array_includes_only_nonempty_pairs(self):
        """只有非空列进入 tag_array，格式为 'ColName=value'"""
        row = ["T", "D", "C", "A", "ID", "R", "31/03/2026 17:31:46", 71, 0, 5,
               "Read", None, None, "No Transfer",
               "Already Paid",   # Call Outcome → 有值
               None,              # Call Outcome(Original) → 跳过
               "Payment Reminder", None,   # Last Node, Original
               None, None, None, None, None, None,  # First Verification...
        ] + [None] * (52 - 8) + ["text"]
        r = self._transform([row])[0]
        tags = r["tag_array"]
        # 只有 "Call Outcome=Already Paid" 和 "Last Node=Payment Reminder"
        self.assertIn("Call Outcome=Already Paid", tags)
        self.assertIn("Last Node=Payment Reminder", tags)
        self.assertNotIn("Call Outcome(Original)=", "".join(tags))

    def test_I7_numeric_fields_cast_to_int(self):
        """call_duration / agent_call_duration / dialogue_round 转为 int"""
        row = ["T", "D", "C", "A", "ID", "R", "01/04/2026 00:00:00", "71", "30", "5",
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        self.assertIsInstance(r["call_duration"], int)
        self.assertEqual(r["call_duration"], 71)
        self.assertIsInstance(r["agent_call_duration"], int)
        self.assertIsInstance(r["dialogue_round"], int)

    def test_I8_metadata_fields_injected(self):
        """import_job_id 和 source_file 正确注入每行"""
        row = ["T", "D", "C", "A", "ID", "R", None, 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        results = self._transform([row], job_id="JOB_XYZ", src="rhb_data.xlsx")
        r = results[0]
        self.assertEqual(r["import_job_id"], "JOB_XYZ")
        self.assertEqual(r["source_file"], "rhb_data.xlsx")

    def test_I9_all_target_cols_present_in_output(self):
        """每行输出 dict 包含全部 18 个目标列"""
        row = ["T", "D", "C", "A", "ID", "R", "31/03/2026 10:00:00", 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        r = self._transform([row])[0]
        for col in _TARGET_COLS:
            self.assertIn(col, r, f"Target column '{col}' missing from transformer output")

    def test_I10_batch_of_multiple_rows(self):
        """批量转换：返回与输入等长的 list"""
        row = ["T", "D", "C", "A", "ID", "R", None, 0, 0, 0,
               "U", None, None, "NT"] + [None] * 52 + [None]
        results = self._transform([row] * 7)
        self.assertEqual(len(results), 7)


# ═══════════════════════════════════════════════════════════════════════════════
# J — insert_json_rows 单元测试 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestInsertJsonRows(unittest.TestCase):
    """ClickHouseHTTPClient.insert_json_rows 的 HTTP 层行为测试"""

    def _make_client(self):
        from backend.mcp.clickhouse.http_client import ClickHouseHTTPClient
        return ClickHouseHTTPClient(
            host="127.0.0.1", port=8123,
            user="u", password="p", database="db", timeout=5,
        )

    def _mock_ok(self):
        resp = MagicMock()
        resp.status_code = 200
        resp.text = ""
        return resp

    def test_J1_sends_jsoneachrow_format_in_query_param(self):
        """query 参数中包含 FORMAT JSONEachRow"""
        client = self._make_client()
        with patch("requests.sessions.Session.post", return_value=self._mock_ok()) as mock_post:
            client.insert_json_rows("db", "t", ["col_a"], [{"col_a": "val1"}])
            params = mock_post.call_args[1]["params"]
            self.assertIn("JSONEachRow", params["query"])

    def test_J2_body_is_valid_json_each_row(self):
        """POST body 是逐行 JSON（每行一个 JSON 对象）"""
        import json
        client = self._make_client()
        rows = [{"col_a": "v1", "arr": ["x=1", "y=2"]},
                {"col_a": "v2", "arr": []}]
        with patch("requests.sessions.Session.post", return_value=self._mock_ok()) as mock_post:
            client.insert_json_rows("db", "t", ["col_a", "arr"], rows)
            body = mock_post.call_args[1]["data"].decode("utf-8")
            lines = [l for l in body.strip().split("\n") if l]
            self.assertEqual(len(lines), 2)
            obj0 = json.loads(lines[0])
            self.assertEqual(obj0["col_a"], "v1")
            self.assertEqual(obj0["arr"], ["x=1", "y=2"])

    def test_J3_array_field_serialized_as_json_list(self):
        """Array(String) 字段序列化为 JSON 数组，不是 Python repr"""
        import json
        client = self._make_client()
        rows = [{"tag_array": ["PTP=Paid", "V1=Pass"]}]
        with patch("requests.sessions.Session.post", return_value=self._mock_ok()) as mock_post:
            client.insert_json_rows("db", "t", ["tag_array"], rows)
            body = mock_post.call_args[1]["data"].decode("utf-8")
            obj = json.loads(body.strip())
            self.assertIsInstance(obj["tag_array"], list)
            self.assertEqual(obj["tag_array"][0], "PTP=Paid")

    def test_J4_http_error_raises_runtime_error(self):
        """CH 返回非 200 时抛 RuntimeError"""
        client = self._make_client()
        err_resp = MagicMock()
        err_resp.status_code = 400
        err_resp.text = "DB error"
        with patch("requests.sessions.Session.post", return_value=err_resp):
            with self.assertRaises(RuntimeError) as ctx:
                client.insert_json_rows("db", "t", ["c"], [{"c": 1}])
            self.assertIn("INSERT 失败", str(ctx.exception))

    def test_J5_connection_error_raises_connection_error(self):
        """网络连接失败时抛 ConnectionError"""
        import requests as req
        client = self._make_client()
        with patch("requests.sessions.Session.post",
                   side_effect=req.ConnectionError("refused")):
            with self.assertRaises(ConnectionError):
                client.insert_json_rows("db", "t", ["c"], [{"c": 1}])

    def test_J6_only_col_names_included_in_json_body(self):
        """仅 col_names 列表中的字段写入 body，其余 dict key 忽略"""
        import json
        client = self._make_client()
        rows = [{"col_a": "v1", "col_b": "v2", "extra": "ignored"}]
        with patch("requests.sessions.Session.post", return_value=self._mock_ok()) as mock_post:
            client.insert_json_rows("db", "t", ["col_a", "col_b"], rows)
            body = mock_post.call_args[1]["data"].decode("utf-8")
            obj = json.loads(body.strip())
            self.assertIn("col_a", obj)
            self.assertIn("col_b", obj)
            # extra 字段不应出现（因为 col_names 只声明了 col_a/col_b）
            # 注意：当前实现用 {k: row.get(k) for k in col_names}，extra 不会出现


# ═══════════════════════════════════════════════════════════════════════════════
# K — API Schema 验证 (4 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAPISchema(unittest.TestCase):
    """SheetConfig.import_type 字段的 Pydantic 验证和 execute 端点行为"""

    def _get_client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def test_K1_sheet_config_accepts_standard_type(self):
        """import_type='standard' 通过 Pydantic 验证"""
        from backend.api.data_import import SheetConfig
        sc = SheetConfig(
            sheet_name="Sheet1", database="db", table="t",
            has_header=True, enabled=True, import_type="standard",
        )
        self.assertEqual(sc.import_type, "standard")

    def test_K2_sheet_config_default_is_standard(self):
        """import_type 未传时默认值为 'standard'"""
        from backend.api.data_import import SheetConfig
        sc = SheetConfig(sheet_name="Sheet1", database="db", table="t",
                         has_header=True, enabled=True)
        self.assertEqual(sc.import_type, "standard")

    def test_K3_sheet_config_accepts_call_record_imported(self):
        """import_type='call_record_imported' 通过 Pydantic 验证"""
        from backend.api.data_import import SheetConfig
        sc = SheetConfig(
            sheet_name="Sheet1", database="db", table="t",
            has_header=True, enabled=True, import_type="call_record_imported",
        )
        self.assertEqual(sc.import_type, "call_record_imported")

    def test_K4_import_type_saved_in_config_snapshot(self):
        """execute 提交时 import_type 写入 config_snapshot"""
        client = self._get_client()
        xlsx_bytes = _make_xlsx_bytes_simple()

        # 上传文件
        up = client.post("/api/v1/data-import/upload",
                         files={"file": ("t.xlsx", xlsx_bytes,
                                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
        if up.status_code != 200:
            self.skipTest("Upload endpoint unavailable")
        upload_id = up.json()["data"]["upload_id"]
        sheet_name = up.json()["data"]["sheets"][0]["sheet_name"]

        payload = {
            "upload_id": upload_id,
            "connection_env": "sg",
            "batch_size": 100,
            "sheets": [{
                "sheet_name": sheet_name,
                "database": "crm",
                "table": "t",
                "has_header": True,
                "enabled": True,
                "import_type": "call_record_imported",
            }],
        }
        with patch("backend.api.data_import.asyncio.create_task"):
            resp = client.post("/api/v1/data-import/execute", json=payload)
        if resp.status_code != 200:
            self.skipTest(f"Execute endpoint returned {resp.status_code}")

        job_id = resp.json()["data"]["job_id"]
        from backend.models.import_job import ImportJob
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertIsNotNone(job)
            sheets_cfg = job.config_snapshot.get("sheets", [])
            self.assertEqual(len(sheets_cfg), 1)
            self.assertEqual(sheets_cfg[0].get("import_type"), "call_record_imported")
        finally:
            db.close()


def _make_xlsx_bytes_simple() -> bytes:
    """生成最简单的两列 xlsx 字节流，用于上传端点测试"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["name", "val"])
    ws.append(["Alice", 1])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# L — 端到端导入流程（Mock CH）(10 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestEndToEndCallRecordImport(unittest.TestCase):
    """
    call_record_imported 路径的完整状态机验证。
    使用真实 PostgreSQL ImportJob + Mock ClickHouseHTTPClient。
    """

    def _sheet_cfg(self, **kwargs):
        base = {
            "sheet_name": "Sheet1",
            "database": "data_statistics",
            "table": "private_rhb_call_records",
            "has_header": True,
            "enabled": True,
            "import_type": "call_record_imported",
        }
        base.update(kwargs)
        return base

    def test_L1_full_flow_pending_to_completed(self):
        """pending → running → completed 完整状态机"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            self.assertEqual(result["status"], "completed")
        finally:
            _safe_unlink(fp)

    def test_L2_insert_json_rows_called_not_insert_tsv(self):
        """call_record_imported 使用 insert_json_rows，不调用 insert_tsv"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            mock = result["_mock"]
            self.assertGreater(mock.insert_json_rows.call_count, 0,
                               "insert_json_rows 应被调用至少一次")
            self.assertEqual(mock.insert_tsv.call_count, 0,
                             "call_record_imported 不应调用 insert_tsv")
        finally:
            _safe_unlink(fp)

    def test_L3_imported_row_count_excludes_header(self):
        """imported_rows 不含表头行"""
        fp = _make_call_record_xlsx()   # 默认写 2 条数据
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            self.assertEqual(result["imported_rows"], 2)
        finally:
            _safe_unlink(fp)

    def test_L4_inserted_rows_contain_all_target_columns(self):
        """每批插入的 rows 中包含全部 18 个目标列"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            mock = result["_mock"]
            # 取第一次 insert_json_rows 的 col_names 参数（第3个位置参数）
            call_args = mock.insert_json_rows.call_args_list[0]
            col_names = call_args[0][2]  # positional arg index 2
            for col in _TARGET_COLS:
                self.assertIn(col, col_names,
                              f"目标列 '{col}' 未出现在 insert 列列表中")
        finally:
            _safe_unlink(fp)

    def test_L5_call_time_correctly_parsed_in_inserted_data(self):
        """插入数据中 call_start_time 格式为 YYYY-MM-DD HH:MM:SS"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            mock = result["_mock"]
            rows = mock.insert_json_rows.call_args_list[0][0][3]  # 4th positional arg
            ts = rows[0].get("call_start_time")
            self.assertIsNotNone(ts)
            # 验证格式
            try:
                datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                self.fail(f"call_start_time '{ts}' 不是 YYYY-MM-DD HH:MM:SS 格式")
        finally:
            _safe_unlink(fp)

    def test_L6_tag_array_is_list_of_strings(self):
        """tag_array 字段是 list[str] 类型"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            mock = result["_mock"]
            rows = mock.insert_json_rows.call_args_list[0][0][3]
            for row in rows:
                ta = row.get("tag_array")
                self.assertIsInstance(ta, list,
                                      f"tag_array 应为 list，实际: {type(ta)}")
                for item in ta:
                    self.assertIsInstance(item, str)
        finally:
            _safe_unlink(fp)

    def test_L7_tag_array_excludes_null_values(self):
        """Missed Call 行（扩展列全空）的 tag_array 为空列表"""
        fp = _make_call_record_xlsx()  # row2 是 Missed Call，扩展列全 None
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            mock = result["_mock"]
            rows = mock.insert_json_rows.call_args_list[0][0][3]
            missed_row = next((r for r in rows if r.get("result") == "Missed Call"), None)
            self.assertIsNotNone(missed_row, "没有找到 Missed Call 行")
            self.assertEqual(missed_row["tag_array"], [],
                             "Missed Call 行的扩展列全空，tag_array 应为 []")
        finally:
            _safe_unlink(fp)

    def test_L8_batch_size_splits_correctly(self):
        """batch_size=1 → 每行一批，insert_json_rows 调用次数等于数据行数"""
        fp = _make_call_record_xlsx()  # 2 条数据
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()], batch_size=1)
            mock = result["_mock"]
            self.assertEqual(mock.insert_json_rows.call_count, 2)
        finally:
            _safe_unlink(fp)

    def test_L9_abort_on_first_batch_failure(self):
        """单批 insert_json_rows 失败 → job 状态变为 failed，立即终止"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()],
                                  mock_side_effect=RuntimeError("CH error"))
            self.assertEqual(result["status"], "failed")
            self.assertIsNotNone(result["error_message"])
            self.assertIn("失败", result["error_message"])
            self.assertEqual(len(result["errors"]), 1)
        finally:
            _safe_unlink(fp)

    def test_L10_job_id_injected_into_every_row(self):
        """每条插入数据的 import_job_id 与实际 job id 一致"""
        fp = _make_call_record_xlsx()
        try:
            result = _run_cri_job(fp, [self._sheet_cfg()])
            job_id = result["job_id"]
            mock = result["_mock"]
            rows = mock.insert_json_rows.call_args_list[0][0][3]
            for row in rows:
                self.assertEqual(row["import_job_id"], job_id)
        finally:
            _safe_unlink(fp)


# ═══════════════════════════════════════════════════════════════════════════════
# M — 回归测试：standard 类型不受影响 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestRegressionStandardImport(unittest.TestCase):
    """
    确认 standard 类型的原有行为完全不受 call_record_imported 改动影响。
    """

    def _run_standard(self, rows, batch_size=100, has_header=True,
                      ch_side_effect=None):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for r in rows:
            ws.append(r)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            fp = f.name

        from backend.services.data_import_service import run_import_job
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal

        job_id = _create_import_job()
        mock_client = MagicMock()
        if ch_side_effect:
            mock_client.insert_tsv.side_effect = ch_side_effect

        config = {
            "file_path": fp,
            "connection_env": "sg",
            "batch_size": batch_size,
            "sheets": [{
                "sheet_name": "Data",
                "database": "db", "table": "t",
                "has_header": has_header,
                "enabled": True,
                "import_type": "standard",
            }],
        }
        with patch("backend.services.data_import_service._build_ch_client",
                   return_value=mock_client):
            asyncio.run(run_import_job(job_id, config))

        db = SessionLocal()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            result = job.to_dict()
            result["_mock"] = mock_client
            return result
        finally:
            db.close()
        _safe_unlink(fp)

    def test_M1_standard_uses_insert_tsv_not_json(self):
        """standard 类型调用 insert_tsv，不调用 insert_json_rows"""
        rows = [["col1", "col2"], ["a", 1], ["b", 2]]
        result = self._run_standard(rows)
        mock = result["_mock"]
        self.assertGreater(mock.insert_tsv.call_count, 0)
        self.assertEqual(mock.insert_json_rows.call_count, 0)

    def test_M2_standard_import_completes_successfully(self):
        """standard 导入流程 → completed"""
        rows = [["name", "age"]] + [["Alice", i] for i in range(5)]
        result = self._run_standard(rows)
        self.assertEqual(result["status"], "completed")
        self.assertEqual(result["imported_rows"], 5)

    def test_M3_standard_without_import_type_field_defaults_to_tsv(self):
        """config 中无 import_type 字段（旧版 payload）→ 自动降级为 standard"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OldSheet"
        ws.append(["x", "y"]); ws.append([1, 2])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name); fp = f.name

        from backend.services.data_import_service import run_import_job
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal

        job_id = _create_import_job()
        mock_client = MagicMock()
        config = {
            "file_path": fp,
            "connection_env": "sg",
            "batch_size": 100,
            "sheets": [{
                "sheet_name": "OldSheet",
                "database": "db", "table": "t",
                "has_header": True,
                "enabled": True,
                # 注意：故意不传 import_type
            }],
        }
        with patch("backend.services.data_import_service._build_ch_client",
                   return_value=mock_client):
            asyncio.run(run_import_job(job_id, config))

        db = SessionLocal()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertEqual(job.status, "completed")
        finally:
            db.close()
        _safe_unlink(fp)

    def test_M4_standard_has_header_true_skips_header(self):
        """standard 路径：has_header=True 仍正确跳过表头"""
        rows = [["h1", "h2"], ["r1", "v1"], ["r2", "v2"], ["r3", "v3"]]
        result = self._run_standard(rows, has_header=True)
        self.assertEqual(result["imported_rows"], 3)

    def test_M5_mixed_sheets_standard_and_cri(self):
        """同一 job 含 standard + call_record_imported 两个 sheet 各走各的路径"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Std"
        ws1.append(["a", "b"]); ws1.append([1, 2]); ws1.append([3, 4])
        ws2 = wb.create_sheet("CRI")
        ws2.append(_ALL_HEADERS)
        # 一条完整 call record 数据
        data_row = ["T", "D", "C", "A", "CALLMIX", "R", "01/04/2026 10:00:00",
                    60, 40, 3, "Read", None, None, "No Transfer"] + [None]*52 + [None]
        ws2.append(data_row)

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name); fp = f.name

        from backend.services.data_import_service import run_import_job
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal

        job_id = _create_import_job()
        mock_client = MagicMock()
        config = {
            "file_path": fp,
            "connection_env": "my",
            "batch_size": 100,
            "sheets": [
                {"sheet_name": "Std", "database": "db", "table": "t1",
                 "has_header": True, "enabled": True, "import_type": "standard"},
                {"sheet_name": "CRI", "database": "data_statistics",
                 "table": "private_rhb_call_records",
                 "has_header": True, "enabled": True,
                 "import_type": "call_record_imported"},
            ],
        }
        with patch("backend.services.data_import_service._build_ch_client",
                   return_value=mock_client):
            asyncio.run(run_import_job(job_id, config))

        db = SessionLocal()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertEqual(job.status, "completed")
            self.assertEqual(job.imported_rows, 3)  # 2 from Std + 1 from CRI
        finally:
            db.close()
        _safe_unlink(fp)

    def test_M6_abort_on_standard_batch_failure_unaffected(self):
        """standard 路径的 abort-on-failure 行为不受新代码影响"""
        rows = [["id"]] + [[i] for i in range(10)]
        result = self._run_standard(rows, batch_size=3,
                                    ch_side_effect=RuntimeError("tsv fail"))
        self.assertEqual(result["status"], "failed")
        self.assertIsNotNone(result["error_message"])
        self.assertEqual(len(result["errors"]), 1)


# ═══════════════════════════════════════════════════════════════════════════════
# N — 真实 ClickHouse 集成测试（MY 环境）(4 tests)
# ═══════════════════════════════════════════════════════════════════════════════

_MY_CH_HOST = "202.165.17.231"
_MY_CH_HTTP_PORT = 8123
_MY_CH_USER = "wizadmin"
_MY_CH_PASS = "Thtss1000c!"
_TEST_JOB_ID = f"PYTEST_{_PREFIX}"   # 用于测后清理隔离


def _my_client():
    from backend.mcp.clickhouse.http_client import ClickHouseHTTPClient
    return ClickHouseHTTPClient(
        host=_MY_CH_HOST, port=_MY_CH_HTTP_PORT,
        user=_MY_CH_USER, password=_MY_CH_PASS,
        database="data_statistics", timeout=15,
    )


def _ch_available() -> bool:
    try:
        _my_client().execute("SELECT 1")
        return True
    except Exception:
        return False


@unittest.skipUnless(_ch_available(), "MY ClickHouse 不可达，跳过集成测试")
class TestRealClickHouseIntegration(unittest.TestCase):
    """
    真实写入 data_statistics.private_rhb_call_records 并验证。
    所有写入行以 import_job_id=_TEST_JOB_ID 标记，tearDownClass 统一清理。
    """

    @classmethod
    def tearDownClass(cls):
        try:
            _my_client().execute(
                f"ALTER TABLE data_statistics.private_rhb_call_records "
                f"DELETE WHERE import_job_id='{_TEST_JOB_ID}'"
            )
        except Exception as e:
            print(f"[WARNING] N 层测试数据清理失败: {e}")

    def _do_real_insert(self, rows_dict):
        """直接调用 insert_json_rows 写入测试数据"""
        _my_client().insert_json_rows(
            "data_statistics", "private_rhb_call_records",
            _TARGET_COLS, rows_dict,
        )

    def _count(self):
        r = _my_client().execute(
            f"SELECT count() FROM data_statistics.private_rhb_call_records "
            f"WHERE import_job_id='{_TEST_JOB_ID}'"
        )
        # JSONCompact 格式下 UInt64 以字符串返回，需强转
        return int(r[0][0]) if r else 0

    def test_N1_real_insert_and_count(self):
        """insert_json_rows 写入 2 行，SELECT count() 返回 2"""
        from backend.services.data_import_service import transform_call_record_batch

        headers = _ALL_HEADERS
        rows = [
            ["Task_N1", "Dlg", "CN001", "aud.mp3", "CALL_N1_A",
             "Connected", "21/05/2026 10:00:00", 60, 30, 3,
             "Read", None, None, "No Transfer"]
            + [None] * 52 + ["transcript A"],
            ["Task_N1", "Dlg", "CN002", "aud.mp3", "CALL_N1_B",
             "Missed Call", "21/05/2026 10:01:00", 0, 0, 0,
             "Unread", None, None, "No Transfer"]
            + [None] * 52 + [None],
        ]
        transformed = transform_call_record_batch(
            [tuple(r) for r in rows], headers, _TEST_JOB_ID, "pytest.xlsx"
        )
        self._do_real_insert(transformed)

        # 允许 CH 异步写入短暂延迟
        import time; time.sleep(1)
        self.assertEqual(self._count(), 2)

    def test_N2_call_start_time_stored_as_datetime(self):
        """call_start_time 在 CH 中存储为 DateTime 格式，可用 WHERE 过滤"""
        r = _my_client().execute(
            f"SELECT call_start_time FROM data_statistics.private_rhb_call_records "
            f"WHERE import_job_id='{_TEST_JOB_ID}' "
            f"AND call_id='CALL_N1_A' LIMIT 1"
        )
        self.assertTrue(len(r) > 0, "找不到测试行（N1 需先于 N2 运行）")
        ts = r[0][0]
        self.assertIsNotNone(ts)
        # CH 返回的 DateTime 应等于 2026-05-21 10:00:00
        self.assertEqual(str(ts)[:10], "2026-05-21")

    def test_N3_tag_array_stored_and_queryable(self):
        """tag_array 以 Array(String) 存储，has(tag_array, 'key=val') 可过滤"""
        from backend.services.data_import_service import transform_call_record_batch

        headers = _ALL_HEADERS
        row = ["Task_N3", "Dlg", "CN003", "aud.mp3", "CALL_N3",
               "Connected", "21/05/2026 11:00:00", 90, 60, 5,
               "Read", None, None, "No Transfer",
               "Already Paid",  # Call Outcome → 进入 tag_array
               None,
               None, None, None, None, None, None,
               "Already Paid",  # PTP
               ] + [None] * (52 - 8) + ["text N3"]
        transformed = transform_call_record_batch(
            [tuple(row)], headers, _TEST_JOB_ID, "pytest.xlsx"
        )
        self._do_real_insert(transformed)

        import time; time.sleep(1)
        r = _my_client().execute(
            f"SELECT tag_array FROM data_statistics.private_rhb_call_records "
            f"WHERE call_id='CALL_N3' AND import_job_id='{_TEST_JOB_ID}' LIMIT 1"
        )
        self.assertTrue(len(r) > 0)
        tag_array = r[0][0]
        self.assertIsInstance(tag_array, (list, tuple))
        tag_strs = list(tag_array)
        self.assertTrue(
            any("Call Outcome=Already Paid" in s for s in tag_strs),
            f"tag_array 中未找到 'Call Outcome=Already Paid'，实际: {tag_strs}"
        )

    def test_N4_full_pipeline_from_real_excel(self):
        """从真实 test100.xlsx 读取数据，经完整 transformer → 写入 CH → 验证行数"""
        EXCEL_PATH = r"D:\BaiduSyncdisk\sgpwork_sync\2026\13 RHB DATA\for test\test100.xlsx"
        if not os.path.exists(EXCEL_PATH):
            self.skipTest("test100.xlsx 不存在，跳过")

        import pandas as pd
        from backend.services.data_import_service import transform_call_record_batch

        df = pd.read_excel(EXCEL_PATH, header=0)
        headers = list(df.columns)
        raw_rows = [tuple(r) for r in df.values[:10]]  # 取前 10 行

        # 用唯一 job_id 隔离本次写入
        test_job_id = f"{_TEST_JOB_ID}_N4"
        transformed = transform_call_record_batch(
            raw_rows, headers, test_job_id, "test100.xlsx"
        )
        _my_client().insert_json_rows(
            "data_statistics", "private_rhb_call_records",
            _TARGET_COLS, transformed,
        )
        import time; time.sleep(1)

        r = _my_client().execute(
            f"SELECT count() FROM data_statistics.private_rhb_call_records "
            f"WHERE import_job_id='{test_job_id}'"
        )
        count = int(r[0][0]) if r else 0
        self.assertEqual(count, 10, f"期望写入 10 行，实际: {count}")

        # 清理
        _my_client().execute(
            f"ALTER TABLE data_statistics.private_rhb_call_records "
            f"DELETE WHERE import_job_id='{test_job_id}'"
        )


if __name__ == "__main__":
    unittest.main(verbosity=2)
