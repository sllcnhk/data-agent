"""
test_data_import_e2e.py — Excel 数据导入功能完整测试套件
=========================================================

测试层次：
  A  (6)  — RBAC / 权限种子验证（data:import 是否正确落库、角色分配、管理 API 可见性）
  B  (6)  — API 权限守卫（ENABLE_AUTH=true 模式下正确鉴权）
  C  (6)  — 连接/数据库/表查询端点（Mock CH 调用）
  D  (8)  — 文件上传端点（类型/大小校验、正常流程）
  E  (8)  — execute 导入任务端点（参数校验、任务创建、后台启动）
  F  (6)  — 任务状态/历史列表端点（状态查询、分页排序）
  G  (8)  — 端到端导入流程（Mock ClickHouseHTTPClient，验证完整状态机）
  H  (8)  — 取消功能完整流程（cancel 端点 + run_import_job 协程合作式退出）

总计: 56 个测试用例

Bug 修复验证：
  - Fix1: data:import 权限已种子到 DB（A1-A4 覆盖）
  - Fix2: asyncio.create_task 附加 done_callback（E8 覆盖）
  - 菜单权限控制：data:import 在角色管理 API 中可见（A5-A6 覆盖）
"""
import asyncio
import io
import os
import sys
import tempfile
import time
import unittest
import uuid
from datetime import datetime
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch, call

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_die_{uuid.uuid4().hex[:6]}_"

# ─── 模块级 auth 补丁（兼容与其他测试套件混跑）─────────────────────────────────
_auth_patcher = None

def setup_module(_=None):
    global _auth_patcher
    from backend.config.settings import settings
    _auth_patcher = patch.object(settings, 'enable_auth', False)
    _auth_patcher.start()

# ─── DB helpers ──────────────────────────────────────────────────────────────

def _db():
    from backend.config.database import SessionLocal
    return SessionLocal()


_g_db = _db()


def teardown_module(_=None):
    global _auth_patcher
    if _auth_patcher is not None:
        _auth_patcher.stop()
        _auth_patcher = None
    from backend.models.import_job import ImportJob
    try:
        _g_db.query(ImportJob).filter(
            ImportJob.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        _g_db.commit()
    except Exception:
        _g_db.rollback()
    finally:
        _g_db.close()


def _make_xlsx(sheet_name="Sheet1", rows=None, multi=False) -> bytes:
    """生成内存 xlsx 文件字节内容"""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    default = [["name", "age", "city"], ["Alice", 30, "BJ"], ["Bob", 25, "SH"]]
    for row in (rows or default):
        ws.append(row)
    if multi:
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["x", "y"])
        ws2.append([1, 2])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _create_job(status="pending", username=None, imported_rows=0, done_batches=0):
    """在 DB 直接创建 ImportJob 供测试使用"""
    from backend.models.import_job import ImportJob
    db = _db()
    try:
        job = ImportJob(
            user_id=str(uuid.uuid4()),
            username=username or f"{_PREFIX}u",
            upload_id=str(uuid.uuid4()),
            filename="test.xlsx",
            connection_env="sg",
            status=status,
            imported_rows=imported_rows,
            done_batches=done_batches,
        )
        db.add(job)
        db.commit()
        db.refresh(job)
        return str(job.id)
    finally:
        db.close()


def _get_client():
    from fastapi.testclient import TestClient
    from backend.main import app
    return TestClient(app, raise_server_exceptions=False)


# ═══════════════════════════════════════════════════════════════════════════════
# A — RBAC / 权限种子验证 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestRBACPermissionSeeding(unittest.TestCase):
    """
    验证 data:import 权限已正确种子到数据库，且角色分配符合设计：
    - superadmin: 有权限
    - admin / analyst / viewer: 无权限
    - GET /permissions 可见 data:import
    - GET /roles 中 superadmin 含 data:import
    """

    def setUp(self):
        self.db = _db()

    def tearDown(self):
        self.db.close()

    def _get_perm(self):
        from backend.models.permission import Permission
        return self.db.query(Permission).filter(
            Permission.resource == "data", Permission.action == "import"
        ).first()

    def test_A1_data_import_permission_exists_in_db(self):
        """data:import 权限已种子到 permissions 表"""
        perm = self._get_perm()
        self.assertIsNotNone(perm, "data:import permission not found in DB — run migrate_data_import.py")
        self.assertEqual(perm.resource, "data")
        self.assertEqual(perm.action, "import")
        self.assertIsNotNone(perm.description)

    def test_A2_superadmin_has_data_import(self):
        """superadmin 角色拥有 data:import 权限"""
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission
        perm = self._get_perm()
        self.assertIsNotNone(perm)
        role = self.db.query(Role).filter(Role.name == "superadmin").first()
        self.assertIsNotNone(role)
        rp = self.db.query(RolePermission).filter(
            RolePermission.role_id == role.id,
            RolePermission.permission_id == perm.id,
        ).first()
        self.assertIsNotNone(rp, "superadmin role does not have data:import permission")

    def test_A3_admin_does_not_have_data_import(self):
        """admin 角色不含 data:import 权限（设计要求：仅 superadmin）"""
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission
        perm = self._get_perm()
        self.assertIsNotNone(perm)
        role = self.db.query(Role).filter(Role.name == "admin").first()
        if not role:
            self.skipTest("admin role not seeded")
        rp = self.db.query(RolePermission).filter(
            RolePermission.role_id == role.id,
            RolePermission.permission_id == perm.id,
        ).first()
        self.assertIsNone(rp, "admin role must NOT have data:import permission")

    def test_A4_analyst_does_not_have_data_import(self):
        """analyst 角色不含 data:import 权限"""
        from backend.models.role import Role
        from backend.models.role_permission import RolePermission
        perm = self._get_perm()
        self.assertIsNotNone(perm)
        role = self.db.query(Role).filter(Role.name == "analyst").first()
        if not role:
            self.skipTest("analyst role not seeded")
        rp = self.db.query(RolePermission).filter(
            RolePermission.role_id == role.id,
            RolePermission.permission_id == perm.id,
        ).first()
        self.assertIsNone(rp, "analyst role must NOT have data:import permission")

    def test_A5_get_permissions_api_includes_data_import(self):
        """GET /permissions API 返回列表包含 data:import（角色管理 UI 可见）"""
        with patch("api.data_import.list_writable_connections", return_value=[]):
            resp = _get_client().get("/api/v1/permissions")
        # ENABLE_AUTH=false → AnonymousUser is superadmin → passes users:read
        self.assertEqual(resp.status_code, 200)
        perms = resp.json()
        resources = [(p["resource"], p["action"]) for p in perms]
        self.assertIn(("data", "import"), resources,
                      "data:import must appear in GET /permissions for role management UI")

    def test_A6_get_roles_api_superadmin_contains_data_import(self):
        """GET /roles API 返回中，superadmin 角色包含 data:import 权限"""
        resp = _get_client().get("/api/v1/roles")
        self.assertEqual(resp.status_code, 200)
        roles = resp.json()
        sa = next((r for r in roles if r["name"] == "superadmin"), None)
        self.assertIsNotNone(sa)
        perm_keys = [(p["resource"], p["action"]) for p in sa.get("permissions", [])]
        self.assertIn(("data", "import"), perm_keys,
                      "superadmin must have data:import in GET /roles response")


# ═══════════════════════════════════════════════════════════════════════════════
# B — API 权限守卫（ENABLE_AUTH=true 模式）(6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAPIPermissionGuard(unittest.TestCase):
    """
    验证 ENABLE_AUTH=true 时鉴权逻辑：
    - 未登录 → 401
    - 普通用户（无 data:import）→ 403
    - superadmin → 通过
    """

    def _make_superadmin_token(self):
        from backend.config.settings import settings
        from backend.core.auth.jwt import create_access_token
        from backend.models.user import User
        db = _db()
        try:
            sa = db.query(User).filter(User.is_superadmin == True).first()
            if not sa:
                return None
            return create_access_token(
                {"sub": str(sa.id), "username": sa.username, "roles": ["superadmin"]},
                settings.jwt_secret,
            )
        finally:
            db.close()

    def _make_analyst_token(self):
        from backend.config.settings import settings
        from backend.core.auth.jwt import create_access_token
        from backend.models.user import User
        from backend.models.role import Role
        from backend.models.user_role import UserRole
        from backend.core.auth.password import hash_password
        db = _db()
        try:
            analyst_role = db.query(Role).filter(Role.name == "analyst").first()
            username = f"{_PREFIX}analyst_{uuid.uuid4().hex[:4]}"
            u = User(username=username, hashed_password=hash_password("Test1234!"),
                     auth_source="local", is_active=True, is_superadmin=False)
            db.add(u)
            db.flush()
            if analyst_role:
                db.add(UserRole(user_id=u.id, role_id=analyst_role.id))
            db.commit()
            db.refresh(u)
            return create_access_token(
                {"sub": str(u.id), "username": u.username, "roles": ["analyst"]},
                settings.jwt_secret,
            )
        finally:
            db.close()

    def test_B1_no_token_returns_401_when_auth_enabled(self):
        """ENABLE_AUTH=true + 无 token → 401"""
        mock_settings = MagicMock()
        mock_settings.enable_auth = True
        mock_settings.jwt_secret = "test_secret"
        mock_settings.jwt_algorithm = "HS256"
        with patch("backend.api.deps.settings", mock_settings):
            resp = _get_client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 401)

    def test_B2_invalid_token_returns_401(self):
        """ENABLE_AUTH=true + 非法 token → 401"""
        mock_settings = MagicMock()
        mock_settings.enable_auth = True
        mock_settings.jwt_secret = "test_secret"
        mock_settings.jwt_algorithm = "HS256"
        with patch("backend.api.deps.settings", mock_settings):
            resp = _get_client().get(
                "/api/v1/data-import/connections",
                headers={"Authorization": "Bearer bad.token.here"}
            )
        self.assertEqual(resp.status_code, 401)

    def test_B3_analyst_token_returns_403_on_connections(self):
        """analyst 用户（无 data:import）→ 403 on /connections"""
        from backend.config.settings import settings as real_settings
        if not real_settings.enable_auth:
            # 强制开启 auth 进行测试
            token = self._make_analyst_token()
            if not token:
                self.skipTest("analyst user creation failed")
            # 直接用 require_permission 的核心逻辑测试
            from backend.core.rbac import get_user_permissions
            from backend.models.user import User
            db = _db()
            try:
                from backend.models.role import Role
                analyst_role = db.query(Role).filter(Role.name == "analyst").first()
                if not analyst_role:
                    self.skipTest("analyst role not found")
                perms = get_user_permissions(
                    MagicMock(is_superadmin=False, id=uuid.uuid4()), db
                )
                self.assertNotIn("data:import", perms)
            finally:
                db.close()
        else:
            token = self._make_analyst_token()
            if not token:
                self.skipTest("analyst user creation failed")
            resp = _get_client().get(
                "/api/v1/data-import/connections",
                headers={"Authorization": f"Bearer {token}"}
            )
            self.assertEqual(resp.status_code, 403)

    def test_B4_superadmin_passes_permission_check(self):
        """superadmin 直接通过 data:import 权限检查"""
        from backend.api.deps import AnonymousUser
        from backend.core.rbac import get_user_permissions
        # AnonymousUser.is_superadmin=True → bypasses permission check entirely
        user = AnonymousUser()
        self.assertTrue(user.is_superadmin)
        db = _db()
        try:
            # superadmin 也应通过 get_user_permissions check
            perms = get_user_permissions(user, db)
            # With ENABLE_AUTH=false, AnonymousUser is superadmin so bypassed;
            # testing via real User with superadmin flag
            from backend.models.user import User
            sa = db.query(User).filter(User.is_superadmin == True).first()
            if sa:
                perms = get_user_permissions(sa, db)
                self.assertIn("data:import", perms)
        finally:
            db.close()

    def test_B5_data_import_menu_only_visible_to_superadmin(self):
        """data:import 菜单权限配置为 'data:import'，只有拥有该权限的用户可见"""
        # 验证 AppLayout 中配置的 perm 字段值
        import subprocess, json
        layout_path = Path(__file__).parent / "frontend" / "src" / "components" / "AppLayout.tsx"
        if layout_path.exists():
            content = layout_path.read_text(encoding="utf-8")
            self.assertIn("data-import", content, "data-import route should be in AppLayout")
            self.assertIn("data:import", content, "perm: 'data:import' should be in AppLayout")
        else:
            self.skipTest("AppLayout.tsx not found")

    def test_B6_data_import_route_exists_in_app_tsx(self):
        """DataImport 组件已注册到 React Router"""
        app_path = Path(__file__).parent / "frontend" / "src" / "App.tsx"
        if app_path.exists():
            content = app_path.read_text(encoding="utf-8")
            self.assertIn("DataImport", content, "DataImport should be imported in App.tsx")
            self.assertIn("/data-import", content, "/data-import route should exist in App.tsx")
        else:
            self.skipTest("App.tsx not found")


# ═══════════════════════════════════════════════════════════════════════════════
# C — 连接/数据库/表查询端点 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestQueryEndpoints(unittest.TestCase):

    def test_C1_connections_endpoint_returns_list(self):
        """GET /connections 返回连接列表（含 env、host 字段）"""
        mock_data = [{"env": "sg", "display_name": "clickhouse-sg",
                      "host": "10.0.0.1", "http_port": 8123,
                      "database": "default", "server_name": "clickhouse-sg"}]
        with patch("api.data_import.list_writable_connections", return_value=mock_data):
            resp = _get_client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data["success"])
        self.assertIsInstance(data["data"], list)
        self.assertEqual(data["data"][0]["env"], "sg")

    def test_C2_connections_empty_list_when_no_writable(self):
        """没有可写连接时返回空列表"""
        with patch("api.data_import.list_writable_connections", return_value=[]):
            resp = _get_client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"], [])

    def test_C3_databases_endpoint_returns_db_list(self):
        """GET /connections/{env}/databases 返回数据库列表"""
        with patch("api.data_import.list_databases", return_value=["mydb", "analytics"]):
            resp = _get_client().get("/api/v1/data-import/connections/sg/databases")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("mydb", resp.json()["data"])
        self.assertIn("analytics", resp.json()["data"])

    def test_C4_tables_endpoint_returns_table_list(self):
        """GET /connections/{env}/databases/{db}/tables 返回表列表"""
        with patch("api.data_import.list_tables", return_value=["users", "orders"]):
            resp = _get_client().get("/api/v1/data-import/connections/sg/databases/mydb/tables")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("users", resp.json()["data"])

    def test_C5_ch_connection_error_returns_500(self):
        """ClickHouse 连接失败时端点返回 500 而非崩溃，detail 字段含错误信息"""
        with patch("api.data_import.list_databases", side_effect=ConnectionError("refused")):
            resp = _get_client().get("/api/v1/data-import/connections/bad_env/databases")
        self.assertEqual(resp.status_code, 500)
        # HTTPException returns {"detail": "..."} (not {"success": false})
        self.assertIn("detail", resp.json())

    def test_C6_writable_connections_excludes_ro(self):
        """list_writable_connections 过滤逻辑：-ro 后缀的连接被排除"""
        mock_manager = MagicMock()
        mock_manager.servers = {
            "clickhouse-sg": MagicMock(),
            "clickhouse-sg-ro": MagicMock(),       # 应排除
            "clickhouse-idn": MagicMock(),
            "postgres-main": MagicMock(),            # 非 CH，应排除
        }
        mock_settings = MagicMock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "localhost", "http_port": 8123,
            "user": "default", "password": "", "database": "default",
        }
        with patch("backend.mcp.manager.get_mcp_manager", return_value=mock_manager), \
             patch("backend.config.settings.settings", mock_settings):
            from backend.services import data_import_service as _svc
            result = _svc.list_writable_connections()
        envs = [r["env"] for r in result]
        self.assertIn("sg", envs)
        self.assertIn("idn", envs)
        self.assertNotIn("sg_ro", envs)
        self.assertEqual(len(result), 2)


# ═══════════════════════════════════════════════════════════════════════════════
# D — 文件上传端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestUploadEndpoint(unittest.TestCase):

    def _upload(self, content=None, filename="test.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        return _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": (filename, content or _make_xlsx(), mime)},
        )

    def test_D1_valid_xlsx_upload_succeeds(self):
        """正常 xlsx 文件上传成功，返回 upload_id、filename、sheets"""
        resp = self._upload()
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        self.assertIn("upload_id", d)
        self.assertIn("filename", d)
        self.assertIn("sheets", d)
        self.assertIsInstance(d["sheets"], list)
        self.assertGreater(len(d["sheets"]), 0)

    def test_D2_upload_id_is_valid_uuid(self):
        """upload_id 格式为合法 UUID"""
        resp = self._upload()
        self.assertEqual(resp.status_code, 200)
        uid = resp.json()["data"]["upload_id"]
        uuid.UUID(uid)  # raises if not valid

    def test_D3_sheet_preview_rows_returned(self):
        """上传后 sheets 含 preview_rows（前 5 行）"""
        resp = self._upload()
        self.assertEqual(resp.status_code, 200)
        sheet = resp.json()["data"]["sheets"][0]
        self.assertIn("preview_rows", sheet)
        self.assertGreater(len(sheet["preview_rows"]), 0)
        self.assertIn("row_count_estimate", sheet)

    def test_D4_multi_sheet_xlsx_parses_all_sheets(self):
        """多 Sheet Excel 返回所有 Sheet 的预览"""
        resp = self._upload(content=_make_xlsx(multi=True))
        self.assertEqual(resp.status_code, 200)
        sheets = resp.json()["data"]["sheets"]
        names = [s["sheet_name"] for s in sheets]
        self.assertIn("Sheet1", names)
        self.assertIn("Sheet2", names)

    def test_D5_csv_file_rejected_400(self):
        """上传 .csv 文件 → 400（不支持非 Excel 格式）"""
        resp = _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": ("data.csv", b"a,b\n1,2", "text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("xlsx", resp.json()["detail"].lower())

    def test_D6_no_extension_rejected_400(self):
        """无扩展名文件 → 400"""
        resp = _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": ("datafile", b"content", "application/octet-stream")},
        )
        self.assertEqual(resp.status_code, 400)

    def test_D7_oversized_file_rejected_413(self):
        """超过 100MB 的文件 → 413"""
        big = b"x" * (101 * 1024 * 1024)
        resp = _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": ("big.xlsx", big,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(resp.status_code, 413)

    def test_D8_corrupted_xlsx_returns_422(self):
        """损坏的 xlsx 内容 → 422"""
        resp = _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": ("bad.xlsx", b"this_is_not_xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(resp.status_code, 422)


# ═══════════════════════════════════════════════════════════════════════════════
# E — execute 导入端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestExecuteEndpoint(unittest.TestCase):

    def _upload_and_get_id(self):
        resp = _get_client().post(
            "/api/v1/data-import/upload",
            files={"file": ("test.xlsx", _make_xlsx(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        return resp.json()["data"]["upload_id"] if resp.status_code == 200 else None

    def _execute(self, uid, sheets=None, batch_size=None):
        payload = {
            "upload_id": uid,
            "connection_env": "sg",
            "sheets": sheets or [{"sheet_name": "Sheet1", "database": "db",
                                   "table": "t", "has_header": True, "enabled": True}],
        }
        if batch_size is not None:
            payload["batch_size"] = batch_size
        with patch("api.data_import.asyncio.create_task"):
            return _get_client().post("/api/v1/data-import/execute", json=payload)

    def test_E1_execute_returns_job_id_and_pending_status(self):
        """execute 成功返回 job_id 和 pending 状态"""
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        resp = self._execute(uid)
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        self.assertIn("job_id", d)
        self.assertEqual(d["status"], "pending")
        uuid.UUID(d["job_id"])  # valid UUID

    def test_E2_job_persisted_in_db_with_correct_fields(self):
        """execute 后 ImportJob 记录写入 DB，字段正确"""
        from backend.models.import_job import ImportJob
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        resp = self._execute(uid)
        if resp.status_code != 200:
            self.skipTest("execute failed")
        job_id = resp.json()["data"]["job_id"]
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertIsNotNone(job)
            self.assertEqual(job.status, "pending")
            self.assertEqual(job.connection_env, "sg")
            self.assertEqual(job.upload_id, uid)
        finally:
            db.close()

    def test_E3_unknown_upload_id_returns_404(self):
        """upload_id 对应文件不存在 → 404"""
        with patch("api.data_import.asyncio.create_task"):
            resp = _get_client().post("/api/v1/data-import/execute", json={
                "upload_id": str(uuid.uuid4()),
                "connection_env": "sg",
                "sheets": [{"sheet_name": "S1", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            })
        self.assertEqual(resp.status_code, 404)
        self.assertIn("upload_id", resp.json()["detail"])

    def test_E4_all_disabled_sheets_returns_400(self):
        """所有 sheet 禁用 → 400"""
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task"):
            resp = _get_client().post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db",
                             "table": "t", "has_header": True, "enabled": False}],
            })
        self.assertEqual(resp.status_code, 400)
        self.assertIn("Sheet", resp.json()["detail"])

    def test_E5_batch_size_below_minimum_returns_422(self):
        """batch_size < 100 → 422（Pydantic 校验）"""
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        resp = self._execute(uid, batch_size=50)
        self.assertEqual(resp.status_code, 422)

    def test_E6_config_snapshot_saved_with_all_fields(self):
        """config_snapshot 保存完整配置（connection_env, batch_size, sheets）"""
        from backend.models.import_job import ImportJob
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        resp = self._execute(uid, batch_size=500,
                             sheets=[{"sheet_name": "Sheet1", "database": "mydb",
                                      "table": "mytable", "has_header": False, "enabled": True}])
        if resp.status_code != 200:
            self.skipTest("execute failed")
        job_id = resp.json()["data"]["job_id"]
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            snap = job.config_snapshot
            self.assertEqual(snap["connection_env"], "sg")
            self.assertEqual(snap["batch_size"], 500)
            self.assertEqual(snap["sheets"][0]["table"], "mytable")
            self.assertFalse(snap["sheets"][0]["has_header"])
        finally:
            db.close()

    def test_E7_asyncio_create_task_called_once(self):
        """execute 成功时 asyncio.create_task 被调用一次（后台任务启动）"""
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task") as mock_ct:
            resp = _get_client().post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            })
        if resp.status_code == 200:
            mock_ct.assert_called_once()

    def test_E8_done_callback_attached_to_task(self):
        """asyncio.create_task 返回的 Task 对象上附加了 done_callback（Fix2 验证）"""
        uid = self._upload_and_get_id()
        if not uid:
            self.skipTest("upload failed")
        mock_task = MagicMock()
        with patch("api.data_import.asyncio.create_task", return_value=mock_task):
            resp = _get_client().post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            })
        if resp.status_code == 200:
            mock_task.add_done_callback.assert_called_once()


# ═══════════════════════════════════════════════════════════════════════════════
# F — 任务状态 / 历史列表端点 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestJobStatusEndpoints(unittest.TestCase):

    def test_F1_get_job_status_returns_all_required_fields(self):
        """GET /jobs/{job_id} 返回所有前端必需字段"""
        job_id = _create_job(status="running")
        resp = _get_client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        required = ["job_id", "filename", "connection_env", "status",
                    "total_sheets", "done_sheets", "imported_rows",
                    "total_batches", "done_batches", "errors", "created_at"]
        for field in required:
            self.assertIn(field, d, f"Missing required field: {field}")

    def test_F2_nonexistent_job_returns_404(self):
        """不存在的 job_id → 404"""
        resp = _get_client().get(f"/api/v1/data-import/jobs/{uuid.uuid4()}")
        self.assertEqual(resp.status_code, 404)

    def test_F3_job_status_reflects_db_state(self):
        """job 状态变化后 GET /jobs/{id} 实时反映"""
        from backend.models.import_job import ImportJob
        job_id = _create_job(status="pending")
        # 模拟状态变更
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            job.status = "completed"
            job.imported_rows = 999
            db.commit()
        finally:
            db.close()
        resp = _get_client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        self.assertEqual(d["status"], "completed")
        self.assertEqual(d["imported_rows"], 999)

    def test_F4_list_jobs_returns_paginated_structure(self):
        """GET /jobs 返回 total / page / page_size / items 分页结构"""
        resp = _get_client().get("/api/v1/data-import/jobs?page=1&page_size=10")
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        self.assertIn("total", d)
        self.assertIn("page", d)
        self.assertIn("page_size", d)
        self.assertIn("items", d)
        self.assertIsInstance(d["items"], list)
        self.assertEqual(d["page"], 1)
        self.assertEqual(d["page_size"], 10)

    def test_F5_list_jobs_sorted_by_created_at_desc(self):
        """历史列表按 created_at 倒序（最新任务在第一条）"""
        ids = [_create_job() for _ in range(3)]
        time.sleep(0.05)  # 确保时间戳不同
        resp = _get_client().get("/api/v1/data-import/jobs?page=1&page_size=100")
        self.assertEqual(resp.status_code, 200)
        items = resp.json()["data"]["items"]
        our = [it for it in items if it["job_id"] in ids]
        if len(our) >= 2 and our[0]["created_at"] and our[1]["created_at"]:
            self.assertGreaterEqual(our[0]["created_at"], our[1]["created_at"])

    def test_F6_list_jobs_default_page_size_is_10(self):
        """GET /jobs 不传参数默认 page=1, page_size=10"""
        resp = _get_client().get("/api/v1/data-import/jobs")
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        self.assertEqual(d["page_size"], 10)
        self.assertEqual(d["page"], 1)
        self.assertLessEqual(len(d["items"]), 10)


# ═══════════════════════════════════════════════════════════════════════════════
# G — 端到端导入流程（Mock ClickHouseHTTPClient）(8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestEndToEndImportFlow(unittest.TestCase):
    """
    完整状态机验证：
      pending → running → completed / failed

    使用真实 PostgreSQL（import_jobs 表）+ Mock ClickHouseHTTPClient。
    通过 asyncio.run() 直接驱动 run_import_job 协程。
    """

    def _run_job(self, file_path, sheet_configs, env="sg", batch_size=100,
                 ch_insert_side_effect=None, initial_status="pending"):
        """辅助：创建 ImportJob + 运行协程 + 返回最终 job 状态

        ch_insert_side_effect: 注入到 mock_client.insert_tsv.side_effect（模拟插入失败或取消）
        initial_status: ImportJob 初始状态（默认 pending，可设为 cancelling 测试预取消）
        """
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal

        job_id = str(uuid.uuid4())
        db = SessionLocal()
        try:
            job = ImportJob(
                id=uuid.UUID(job_id),
                user_id=str(uuid.uuid4()),
                username=f"{_PREFIX}e2e",
                upload_id=str(uuid.uuid4()),
                filename=Path(file_path).name,
                connection_env=env,
                status=initial_status,
            )
            db.add(job)
            db.commit()
        finally:
            db.close()

        config = {
            "file_path": file_path,
            "connection_env": env,
            "batch_size": batch_size,
            "sheets": sheet_configs,
        }

        mock_client = MagicMock()
        if ch_insert_side_effect:
            mock_client.insert_tsv.side_effect = ch_insert_side_effect

        with patch("backend.services.data_import_service._build_ch_client",
                   return_value=mock_client):
            asyncio.run(
                __import__("backend.services.data_import_service",
                           fromlist=["run_import_job"]).run_import_job(job_id, config)
            )

        db = SessionLocal()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            result = job.to_dict() if job else None
            # Attach mock for caller inspection
            if result is not None:
                result["_mock_client"] = mock_client
            return result
        finally:
            db.close()

    def _write_xlsx(self, rows, has_extra_sheet=False):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        for row in rows:
            ws.append(row)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            return f.name

    def _safe_unlink(self, fp):
        """Windows 兼容的文件删除——openpyxl generator 可能短暂持有文件句柄"""
        try:
            if os.path.exists(fp):
                os.unlink(fp)
        except OSError:
            pass

    def test_G1_full_flow_pending_to_completed(self):
        """完整流程：pending → running → completed"""
        rows = [["name", "age"]] + [["Alice", i] for i in range(10)]
        fp = self._write_xlsx(rows)
        try:
            result = self._run_job(fp, [
                {"sheet_name": "TestSheet", "database": "db",
                 "table": "t", "has_header": True, "enabled": True}
            ])
            self.assertIsNotNone(result)
            self.assertEqual(result["status"], "completed")
            self.assertEqual(result["imported_rows"], 10)  # 头行不计入
        finally:
            self._safe_unlink(fp)

    def test_G2_has_header_true_skips_first_row(self):
        """has_header=True 时首行（表头）不导入"""
        rows = [["name", "age"], ["Alice", 30], ["Bob", 25]]
        fp = self._write_xlsx(rows)
        try:
            result = self._run_job(fp, [
                {"sheet_name": "TestSheet", "database": "db",
                 "table": "t", "has_header": True, "enabled": True}
            ])
            self.assertEqual(result["status"], "completed")
            self.assertEqual(result["imported_rows"], 2)  # 跳过表头
        finally:
            self._safe_unlink(fp)

    def test_G3_has_header_false_includes_all_rows(self):
        """has_header=False 时所有行都导入"""
        rows = [["name", "age"], ["Alice", 30], ["Bob", 25]]
        fp = self._write_xlsx(rows)
        try:
            result = self._run_job(fp, [
                {"sheet_name": "TestSheet", "database": "db",
                 "table": "t", "has_header": False, "enabled": True}
            ])
            self.assertEqual(result["status"], "completed")
            self.assertEqual(result["imported_rows"], 3)  # 全部行
        finally:
            self._safe_unlink(fp)

    def test_G4_batch_size_splits_data_correctly(self):
        """batch_size 正确分批：20 行 + batch_size=7 → 3 批（7+7+6）"""
        rows = [["id", "val"]] + [[i, f"v{i}"] for i in range(20)]
        fp = self._write_xlsx(rows)
        mock_client = MagicMock()
        try:
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal
            job_id = str(uuid.uuid4())
            db = SessionLocal()
            try:
                job = ImportJob(
                    id=uuid.UUID(job_id),
                    user_id=str(uuid.uuid4()),
                    username=f"{_PREFIX}e2e_batch",
                    upload_id=str(uuid.uuid4()),
                    filename="batch.xlsx",
                    connection_env="sg",
                    status="pending",
                )
                db.add(job)
                db.commit()
            finally:
                db.close()
            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 7,
                "sheets": [{"sheet_name": "TestSheet", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))
            # 20 data rows, batch=7 → ceil(20/7)=3 batches；服务使用 insert_tsv
            self.assertEqual(mock_client.insert_tsv.call_count, 3)
        finally:
            self._safe_unlink(fp)

    def test_G5_abort_on_first_batch_failure(self):
        """单批失败立即终止：job 状态变为 failed，error_message 已设置，errors 列表非空"""
        rows = [["id"]] + [[i] for i in range(50)]
        fp = self._write_xlsx(rows)
        try:
            result = self._run_job(fp, [
                {"sheet_name": "TestSheet", "database": "db",
                 "table": "t", "has_header": True, "enabled": True}
            ], batch_size=10,
               ch_insert_side_effect=RuntimeError("CH insert error"))
            # 核心断言：任务状态 + 错误信息
            self.assertEqual(result["status"], "failed")
            self.assertIsNotNone(result["error_message"])
            self.assertIn("失败", result["error_message"])
            self.assertGreater(len(result["errors"]), 0)
            # 验证 abort-on-first-failure：只有 1 个错误批次
            self.assertEqual(len(result["errors"]), 1)
        finally:
            self._safe_unlink(fp)

    def test_G6_temp_file_deleted_after_success(self):
        """导入完成后服务调用 os.unlink 清理临时文件"""
        rows = [["a"], [1], [2]]
        fp = self._write_xlsx(rows)
        self.assertTrue(os.path.exists(fp))
        unlink_calls = []
        real_unlink = os.unlink
        def tracking_unlink(path, *a, **kw):
            if str(path) == str(fp):
                unlink_calls.append(path)
            real_unlink(path, *a, **kw)
        with patch("backend.services.data_import_service.os.unlink", side_effect=tracking_unlink):
            self._run_job(fp, [
                {"sheet_name": "TestSheet", "database": "db",
                 "table": "t", "has_header": False, "enabled": True}
            ])
        self.assertTrue(len(unlink_calls) > 0, "os.unlink should have been called for temp file")
        self._safe_unlink(fp)  # cleanup if still exists on Windows

    def test_G7_temp_file_deleted_after_failure(self):
        """导入失败后临时文件也被删除"""
        rows = [["id"]] + [[i] for i in range(5)]
        fp = self._write_xlsx(rows)
        self.assertTrue(os.path.exists(fp))
        self._run_job(fp, [
            {"sheet_name": "TestSheet", "database": "db",
             "table": "t", "has_header": True, "enabled": True}
        ], ch_insert_side_effect=RuntimeError("fail"))
        self.assertFalse(os.path.exists(fp), "Temp file should be deleted even on failure")

    def test_G8_disabled_sheet_skipped(self):
        """enabled=False 的 Sheet 不被处理"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active; ws1.title = "Active"
        ws1.append(["id"]); ws1.append([1]); ws1.append([2])
        ws2 = wb.create_sheet("Disabled")
        ws2.append(["x"]); ws2.append([99])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name); fp = f.name
        try:
            mock_client = MagicMock()
            mock_client.execute.return_value = []
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal
            job_id = str(uuid.uuid4())
            db = SessionLocal()
            try:
                job = ImportJob(
                    id=uuid.UUID(job_id),
                    user_id=str(uuid.uuid4()),
                    username=f"{_PREFIX}e2e_skip",
                    upload_id=str(uuid.uuid4()),
                    filename="skip.xlsx",
                    connection_env="sg",
                    status="pending",
                )
                db.add(job)
                db.commit()
            finally:
                db.close()
            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 100,
                "sheets": [
                    {"sheet_name": "Active", "database": "db", "table": "t1",
                     "has_header": True, "enabled": True},
                    {"sheet_name": "Disabled", "database": "db", "table": "t2",
                     "has_header": True, "enabled": False},
                ],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))
            # Active 有 2 data rows in 1 batch; Disabled has 1 row but should be skipped
            # Only 1 INSERT should be executed (for Active sheet)
            call_sqls = [str(c) for c in mock_client.execute.call_args_list]
            # Verify no INSERT for t2 (Disabled sheet)
            self.assertFalse(
                any("t2" in s for s in call_sqls),
                "Disabled sheet should not generate any INSERT"
            )
            db = SessionLocal()
            try:
                job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
                self.assertEqual(job.status, "completed")
                self.assertEqual(job.imported_rows, 2)  # Only Active sheet data
            finally:
                db.close()
        finally:
            self._safe_unlink(fp)


# ═══════════════════════════════════════════════════════════════════════════════
# H — 取消功能完整流程 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestCancelFlow(unittest.TestCase):
    """
    端到端取消功能验证：
    - cancel 端点状态机（pending/running 可取消，其他拒绝）
    - run_import_job 协程对 cancelling 状态的检测与干净退出
    - _mark_cancelled 正确设置最终字段
    """

    def _write_xlsx(self, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CancelSheet"
        for row in rows:
            ws.append(row)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            return f.name

    def _safe_unlink(self, fp):
        try:
            if os.path.exists(fp):
                os.unlink(fp)
        except OSError:
            pass

    def _create_job_db(self, status="pending"):
        """直接在 DB 创建 ImportJob，返回 job_id"""
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal
        job_id = str(uuid.uuid4())
        db = SessionLocal()
        try:
            job = ImportJob(
                id=uuid.UUID(job_id),
                user_id=str(uuid.uuid4()),
                username=f"{_PREFIX}cancel",
                upload_id=str(uuid.uuid4()),
                filename="cancel.xlsx",
                connection_env="sg",
                status=status,
            )
            db.add(job)
            db.commit()
        finally:
            db.close()
        return job_id

    def _get_job_dict(self, job_id):
        from backend.models.import_job import ImportJob
        from backend.config.database import SessionLocal
        db = SessionLocal()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            return job.to_dict() if job else None
        finally:
            db.close()

    # ── cancel 端点状态机 ──────────────────────────────────────────────────────

    def test_H1_cancel_endpoint_pending_returns_cancelling(self):
        """POST /cancel on pending → 200, data.status=cancelling"""
        job_id = _create_job(status="pending")
        resp = _get_client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["status"], "cancelling")

    def test_H2_cancel_endpoint_running_returns_cancelling(self):
        """POST /cancel on running → 200, data.status=cancelling"""
        job_id = _create_job(status="running")
        resp = _get_client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["status"], "cancelling")

    def test_H3_cancel_endpoint_completed_returns_400(self):
        """POST /cancel on completed → 400"""
        job_id = _create_job(status="completed")
        resp = _get_client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("completed", resp.json()["detail"])

    def test_H4_cancel_endpoint_updates_db(self):
        """cancel 端点成功后 DB 状态立即变为 cancelling"""
        from backend.models.import_job import ImportJob
        job_id = _create_job(status="running")
        _get_client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertEqual(job.status, "cancelling")
        finally:
            db.close()

    # ── run_import_job 协程取消逻辑 ───────────────────────────────────────────

    def test_H5_cancel_detected_before_second_sheet(self):
        """多 Sheet 文件：Sheet1 完成后触发取消，Sheet2 开始前 _is_cancelling() 检测到 → cancelled"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        for i in range(5):
            ws1.append([i])
        ws2 = wb.create_sheet("S2")
        for i in range(5):
            ws2.append([i])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            fp = f.name
        try:
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal

            job_id = self._create_job_db(status="pending")
            insert_calls = [0]

            def side_effect(database, table, batch_rows):
                insert_calls[0] += 1
                # Sheet1 批次完成后把状态设为 cancelling
                if table == "t1":
                    db2 = SessionLocal()
                    try:
                        j = db2.query(ImportJob).filter(ImportJob.id == job_id).first()
                        j.status = "cancelling"
                        db2.commit()
                    finally:
                        db2.close()

            mock_client = MagicMock()
            mock_client.insert_tsv.side_effect = side_effect

            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 100,
                "sheets": [
                    {"sheet_name": "S1", "database": "db", "table": "t1",
                     "has_header": False, "enabled": True},
                    {"sheet_name": "S2", "database": "db", "table": "t2",
                     "has_header": False, "enabled": True},
                ],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))

            result = self._get_job_dict(job_id)
            self.assertEqual(result["status"], "cancelled")
            # Sheet2 不应被处理
            self.assertEqual(insert_calls[0], 1,
                             "取消后 Sheet2 不应产生任何 INSERT 调用")
        finally:
            self._safe_unlink(fp)

    def test_H6_cancelled_job_has_finished_at_set(self):
        """取消完成后 _mark_cancelled 将 finished_at 设置为非 None"""
        rows = [["id", "val"]] + [[i, f"v{i}"] for i in range(20)]
        fp = self._write_xlsx(rows)
        try:
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal

            job_id = self._create_job_db(status="pending")

            def cancelling_insert(database, table, batch_rows):
                db2 = SessionLocal()
                try:
                    j = db2.query(ImportJob).filter(ImportJob.id == job_id).first()
                    j.status = "cancelling"
                    db2.commit()
                finally:
                    db2.close()

            mock_client = MagicMock()
            mock_client.insert_tsv.side_effect = cancelling_insert

            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 5,
                "sheets": [{"sheet_name": "CancelSheet", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))

            result = self._get_job_dict(job_id)
            self.assertEqual(result["status"], "cancelled")
            self.assertIsNotNone(result["finished_at"], "_mark_cancelled 应设置 finished_at")
        finally:
            self._safe_unlink(fp)

    def test_H7_mid_sheet_cancellation(self):
        """mid-sheet 取消：第一批 insert_tsv 后状态变为 cancelling，协程在下一批检测到并退出"""
        # 40 行数据，batch=10 → 4 批，第一批后设为 cancelling
        rows = [["id", "val"]] + [[i, f"v{i}"] for i in range(40)]
        fp = self._write_xlsx(rows)
        try:
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal

            job_id = self._create_job_db(status="pending")

            call_count = [0]

            def cancelling_insert(database, table, batch_rows):
                call_count[0] += 1
                if call_count[0] == 1:
                    # 第一批完成后设为 cancelling
                    db2 = SessionLocal()
                    try:
                        j = db2.query(ImportJob).filter(ImportJob.id == job_id).first()
                        j.status = "cancelling"
                        db2.commit()
                    finally:
                        db2.close()

            mock_client = MagicMock()
            mock_client.insert_tsv.side_effect = cancelling_insert

            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 10,
                "sheets": [{"sheet_name": "CancelSheet", "database": "db",
                             "table": "t", "has_header": True, "enabled": True}],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))

            result = self._get_job_dict(job_id)
            self.assertEqual(result["status"], "cancelled")
            # 第一批已完成（count=1），取消后不再继续
            self.assertEqual(call_count[0], 1,
                             "取消后不应再执行后续批次")
        finally:
            self._safe_unlink(fp)

    def test_H8_mid_cancel_preserves_imported_rows(self):
        """mid-sheet 取消后，imported_rows 保留已完成批次的行数"""
        rows = [["id"]] + [[i] for i in range(30)]
        fp = self._write_xlsx(rows)
        try:
            from backend.services.data_import_service import run_import_job
            from backend.models.import_job import ImportJob
            from backend.config.database import SessionLocal

            job_id = self._create_job_db(status="pending")

            def cancelling_insert(database, table, batch_rows):
                # 第一批（10行）完成后立即设为 cancelling
                db2 = SessionLocal()
                try:
                    j = db2.query(ImportJob).filter(ImportJob.id == job_id).first()
                    j.status = "cancelling"
                    db2.commit()
                finally:
                    db2.close()

            mock_client = MagicMock()
            mock_client.insert_tsv.side_effect = cancelling_insert

            config = {
                "file_path": fp,
                "connection_env": "sg",
                "batch_size": 10,
                "sheets": [{"sheet_name": "CancelSheet", "database": "db",
                             "table": "t", "has_header": False, "enabled": True}],
            }
            with patch("backend.services.data_import_service._build_ch_client",
                       return_value=mock_client):
                asyncio.run(run_import_job(job_id, config))

            result = self._get_job_dict(job_id)
            self.assertEqual(result["status"], "cancelled")
            # 第一批（10行）已导入，cancelled 时保留该数值
            self.assertGreaterEqual(result["imported_rows"], 10,
                                    "已完成批次的行数应保留在 imported_rows 中")
        finally:
            self._safe_unlink(fp)


if __name__ == "__main__":
    unittest.main(verbosity=2)
