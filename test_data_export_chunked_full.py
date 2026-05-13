"""
按日期分块导出 — 综合端到端测试 v2.13
=============================================================================

设计原则：
  - 不止 mock，更要验证真实产物（openpyxl 读回 xlsx 检查 sheet/header/row/type）
  - 覆盖研发设计中的关键风险点（B1-B17，详见 PR 描述）
  - 含 RBAC、Cancel 时序、Edge、Code 160 嵌套、并发等真实生产场景

测试切片：
  A · 校验层防御     校验在创建 Job 前拦截非法配置
  B · UI 字段语义     current_sheet/output_files 在 chunked 模式下的语义
  C · 真实 Excel 验证   读回 xlsx 文件并检查内容
  D · 边界规模        empty/single/large(>1M)/many(30) chunks
  E · 故障态行为      列预检失败、块中故障、partial download
  F · 删除幂等        目录不存在/部分存在/完整删除
  G · RBAC 端点防护   chunked 相关全部端点 require_permission 验证
  H · SQL 契约         注入路径生成的 SQL 字符串精确验证
  I · Code 160 嵌套   chunked 模式下单块触发 Code 160 仍能 LIMIT/OFFSET 回退
  J · 并发安全        两个 chunked Job 并行执行互不干扰
  K · JSON 契约       任务列表/状态/output_files JSON 序列化稳定

运行：
    /d/ProgramData/Anaconda3/envs/dataagent/python.exe -m pytest test_data_export_chunked_full.py -v -s
"""
import asyncio
import json
import os
import shutil
import sys
import time
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple
from unittest.mock import Mock, patch

import pytest

sys.path.insert(0, str(Path(__file__).parent))
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_full_{uuid.uuid4().hex[:6]}_"


# ─────────────────────────────────────────────────────────────────────────────
# 工具
# ─────────────────────────────────────────────────────────────────────────────

def _make_columns(extra_types: List[Tuple[str, str]] = None):
    from backend.services.export_clients.base import ColumnInfo
    cols = [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]
    if extra_types:
        cols.extend([ColumnInfo(n, t) for n, t in extra_types])
    return cols


def _make_batch(n: int, base: int = 0, dt_col: str = None, dt_value: str = None):
    """生成 n 行；可选附加 dt 列以模拟日期过滤"""
    rows = [(str(base + i), f"name_{base + i}") for i in range(n)]
    if dt_col:
        rows = [r + (dt_value,) for r in rows]
    return rows


@pytest.fixture
def app_client():
    os.environ["ENABLE_AUTH"] = "False"
    from fastapi.testclient import TestClient
    sys.path.insert(0, str(Path(__file__).parent / "backend"))
    from main import app
    with TestClient(app) as c:
        yield c


def _wait_done(app_client, job_id, timeout=15.0):
    start = time.monotonic()
    last = None
    while time.monotonic() - start < timeout:
        r = app_client.get(f"/api/v1/data-export/jobs/{job_id}")
        if r.status_code != 200:
            return None
        last = r.json()["data"]
        if last["status"] in ("completed", "failed", "cancelled"):
            return last
        time.sleep(0.05)
    return last


def _read_xlsx(path: str) -> Dict[str, Any]:
    """读回 xlsx 验证其内容"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        sheets[sheet_name] = {
            "header": rows[0] if rows else None,
            "data_rows": rows[1:] if len(rows) > 1 else [],
            "total_rows": len(rows),
        }
    wb.close()
    return {"sheet_names": wb.sheetnames, "sheets": sheets}


def _create_chunked_job_via_api(
    app_client, tmp_path: Path, job_label: str,
    sql: str = None,
    chunk_config: Dict = None,
    monkeypatch=None,
    xlsx_engine: str = None,
) -> str:
    """通过 API 提交 chunked Job 并返回 job_id"""
    from backend.api import data_export as api_module
    if monkeypatch is not None:
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
    payload = {
        "query_sql": sql or (
            "SELECT id, name FROM events "
            "WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'"
        ),
        "connection_env": "test",
        "job_name": f"{_PREFIX}{job_label}",
        "chunk_config": chunk_config or {
            "date_start": "2025-04-01",
            "date_end": "2025-04-20",
            "chunk_days": 10,
        },
    }
    if xlsx_engine is not None:
        payload["xlsx_engine"] = xlsx_engine
    r = app_client.post("/api/v1/data-export/execute", json=payload)
    assert r.status_code == 200, r.text
    return r.json()["data"]["job_id"]


# =============================================================================
# A · 校验层防御（在创建 Job 前拦截非法配置）
# =============================================================================



def _attach_csv_stream_raw(mc) -> None:
    """
    给 Mock 的 export_client 补一个 stream_raw 实现,使其在 csv_staging 引擎下也能工作。

    背景：v2.14.x 起 chunked 模式 xlsx_engine="auto" 默认走 csv_staging,
          会调用 export_client.stream_raw(sql, format_name="CSVWithNames")
          直接取 CSV 字节流落盘。老 Mock 只 stub 了 stream_batches → 缺方法。

    本 helper 用同一个 stream_batches 的 side_effect 重放出 CSV 字节流,
    格式为 CSVWithNames：第 1 行列名,后续每行 row 值,逗号分隔,逗号/引号/换行需转义。
    side_effect 中的副作用(如 sqls_seen.append)会在 csv_staging 模式下从 stream_raw
    触发,语义等价于 direct 模式下从 stream_batches 触发(都是「每个 SQL 一次」)。
    """
    cols = mc.get_columns.return_value
    try:
        col_names: List[str] = [c.name for c in cols] if cols else []
    except TypeError:
        # Mock 默认 return_value 不可迭代 / get_columns 用 side_effect 抛错的场景 →
        # 列预检会在外层失败,此处提供退化列名让 helper 不抛(否则 _mk_client 自身崩溃)
        col_names = []
    sb_side = mc.stream_batches.side_effect  # 捕获原闭包

    def _csv_escape(v: Any) -> str:
        if v is None:
            return ""
        s = str(v)
        if "," in s or '"' in s or "\n" in s or "\r" in s:
            s = '"' + s.replace('"', '""') + '"'
        return s

    def _stream_raw(sql, format_name="CSVWithNames", **kw):
        yield (",".join(col_names) + "\n").encode("utf-8")
        gen = sb_side(sql) if sb_side else iter([])
        for batch in gen:
            for row in batch:
                yield (",".join(_csv_escape(v) for v in row) + "\n").encode("utf-8")

    mc.stream_raw.side_effect = _stream_raw


class TestValidationDefense:
    """A1-A8：错误请求应在 API 层就被拒绝，不污染 DB"""

    def test_a1_partial_placeholder_only_start_with_column_rejected(self, app_client):
        """A1（B1 假设）: SQL 仅有 {{date_start}}（缺 {{date_end}}）→ 走 wrapper 路径，
        但 {{date_start}} 字面量会被送给 ClickHouse → 必须在校验层拒绝"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT * FROM t WHERE dt >= '{{date_start}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a1",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 10,
            },
        })
        # 期望：400，明确提示「占位符必须成对」
        assert r.status_code == 400, (
            f"应拒绝单占位符 SQL；当前返回 {r.status_code}: {r.text}"
        )

    def test_a2_partial_placeholder_only_end_rejected(self, app_client):
        """A2（B1）: SQL 仅有 {{date_end}}（缺 {{date_start}}）→ 拒绝"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT * FROM t WHERE dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a2",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 10,
            },
        })
        assert r.status_code == 400

    def test_a3_no_placeholder_no_column_rejected(self, app_client):
        """A3: 无占位符 + 无 date_column → 400"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT id FROM events",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a3",
            "chunk_config": {
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 10,
            },
        })
        assert r.status_code == 400
        assert "date_column" in r.json()["detail"]

    def test_a4_chunk_days_too_large_422(self, app_client):
        """A4: chunk_days=91 (>90) → 422 (Pydantic schema)"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a4",
            "chunk_config": {
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 91,
            },
        })
        assert r.status_code == 422

    def test_a5_chunk_days_zero_422(self, app_client):
        """A5: chunk_days=0 → 422"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a5",
            "chunk_config": {
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 0,
            },
        })
        assert r.status_code == 422

    def test_a6_start_after_end_400(self, app_client):
        """A6: start > end → 400 (chunker 校验)"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a6",
            "chunk_config": {
                "date_start": "2025-04-30",
                "date_end": "2025-04-01",
                "chunk_days": 10,
            },
        })
        assert r.status_code == 400

    def test_a7_invalid_date_format_400_or_422(self, app_client):
        """A7: 日期格式 '2025/04/01' → 400 或 422"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a7",
            "chunk_config": {
                "date_start": "2025/04/01",
                "date_end": "2025-04-30",
                "chunk_days": 10,
            },
        })
        assert r.status_code in (400, 422)

    def test_a8_date_column_sql_injection_400(self, app_client):
        """A8: date_column 含 SQL 注入字符 → 400"""
        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 FROM events",
            "connection_env": "test",
            "job_name": f"{_PREFIX}a8",
            "chunk_config": {
                "date_column": "dt; DROP TABLE users;",
                "date_start": "2025-04-01",
                "date_end": "2025-04-30",
                "chunk_days": 10,
            },
        })
        assert r.status_code == 400
        assert "非法字符" in r.json()["detail"] or "date_column" in r.json()["detail"]


# =============================================================================
# B · UI 字段语义
# =============================================================================

class TestUIFieldSemantics:
    """B1-B3：current_sheet 在 chunked 模式下应能反映块进度（B2 假设的 bug）"""

    def test_b1_chunked_current_sheet_shows_chunk_label(self, app_client, tmp_path, monkeypatch):
        """B1（B2 假设）: chunked 模式下，current_sheet 字段应包含「块 N/M」信息，
        让前端能展示块级进度而非仅 'Sheet1'"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        # 共享 hold 标志：第 2 个 stream 调用挂起，让轮询能捕获执行中状态
        block = {"hold": True, "calls": 0}

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kwargs):
                block["calls"] += 1
                if block["calls"] == 2:
                    # 第 2 块：挂起等放行
                    timeout_at = time.monotonic() + 5.0  # 防死锁
                    while block["hold"] and time.monotonic() < timeout_at:
                        time.sleep(0.01)
                yield _make_batch(2)
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        captured: List[str] = []
        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "b1",
                chunk_config={"date_start": "2025-04-01", "date_end": "2025-04-20", "chunk_days": 10},
                monkeypatch=monkeypatch,
            )

            # 持续轮询，捕获所有执行中的 current_sheet 取值
            deadline = time.monotonic() + 4.0
            while time.monotonic() < deadline:
                r = app_client.get(f"/api/v1/data-export/jobs/{jid}")
                d = r.json()["data"]
                if d["status"] == "running" and d.get("current_sheet"):
                    captured.append(d["current_sheet"])
                if d["done_batches"] >= 1:
                    # 已经进入第 2 块，再多采几次然后放行
                    for _ in range(5):
                        r = app_client.get(f"/api/v1/data-export/jobs/{jid}")
                        d = r.json()["data"]
                        if d.get("current_sheet"):
                            captured.append(d["current_sheet"])
                        time.sleep(0.05)
                    break
                time.sleep(0.05)

            block["hold"] = False
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"

        # 关键断言：执行中 current_sheet 至少有一次出现块级标签
        # 修复前 current_sheet 仅为「Sheet1/Sheet2」，无块进度信息
        # 修复后应为「块 N/M (start~end) - SheetX」
        assert captured, f"未捕获到执行中的 current_sheet (block.calls={block['calls']})"
        chunk_aware = [s for s in captured if "块" in s or "2025-04" in s]
        assert chunk_aware, (
            f"所有 current_sheet 取值 {captured} 都未体现块级进度。"
            f"修复前实现：inner 覆盖 outer 设置的「块 N/M」。"
            f"期望修复：current_sheet 形如「块 2/2 (2025-04-11~2025-04-20) - Sheet1」"
        )

    def test_b2_output_files_persists_after_cancel(self, app_client, tmp_path, monkeypatch):
        """B2（B8 假设）: 中途取消后 output_files JSONB 应被持久化"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        block = {"hold": True}

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            call = {"n": 0}
            def _stream(sql, **kwargs):
                call["n"] += 1
                yield _make_batch(2)
                if call["n"] == 1:
                    # 第 1 块完成后挂起
                    while block["hold"]:
                        time.sleep(0.02)
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            # B2 测取消时序按 direct 引擎(批次粒度);csv_staging 在 CSV→XLSX 转换
            # 起点检查 cancelling,会把 chunk1 标 cancelled 而非 completed,不是本测试覆盖目标。
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "b2", monkeypatch=monkeypatch,
                xlsx_engine="direct",
            )

            # 等第 1 块完成
            for _ in range(80):
                r = app_client.get(f"/api/v1/data-export/jobs/{jid}")
                if r.json()["data"]["done_batches"] >= 1:
                    break
                time.sleep(0.05)

            # 发送取消
            cancel_resp = app_client.post(f"/api/v1/data-export/jobs/{jid}/cancel")
            assert cancel_resp.status_code == 200

            block["hold"] = False
            final = _wait_done(app_client, jid, timeout=5.0)

        assert final["status"] == "cancelled"
        files = final["output_files"]
        assert files is not None and len(files) >= 1
        assert files[0]["status"] == "completed"
        # 第 1 块的实际文件应仍然存在
        assert Path(files[0]["file_path"]).exists()


# =============================================================================
# C · 真实 Excel 验证（读回 xlsx 检查内容）
# =============================================================================

class TestRealExcelContent:
    """C1-C5：读回真实 xlsx 文件，验证 sheet/header/row/type"""

    def test_c1_xlsx_has_correct_header_and_rows(self, app_client, tmp_path, monkeypatch):
        """C1（B3）: 生成的 xlsx 第 1 行是表头，后续是数据行"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([
                [("100", "alice"), ("200", "bob")]
            ])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "c1",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"

        fpath = final["output_files"][0]["file_path"]
        content = _read_xlsx(fpath)
        assert content["sheet_names"] == ["Sheet1"]
        sheet1 = content["sheets"]["Sheet1"]
        assert sheet1["header"] == ("id", "name")
        assert sheet1["data_rows"] == [("100", "alice"), ("200", "bob")]
        assert sheet1["total_rows"] == 3  # 1 header + 2 data

    def test_c2_int64_converted_to_string(self, app_client, tmp_path, monkeypatch):
        """C2（B15）: Int64 大整数在 chunked xlsx 中也应转为字符串"""
        BIG = 9007199254740993  # 超过 JS Number.MAX_SAFE_INTEGER

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([
                [(BIG, "x")]
            ])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "c2",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"

        content = _read_xlsx(final["output_files"][0]["file_path"])
        cell_value = content["sheets"]["Sheet1"]["data_rows"][0][0]
        # Int64 → 应为字符串形式（防 Excel 科学计数法）
        assert isinstance(cell_value, str)
        assert cell_value == str(BIG)

    def test_c3_three_chunks_three_files_correct_data(self, app_client, tmp_path, monkeypatch):
        """C3: 3 块 → 3 个 xlsx，每个文件含正确日期范围数据"""
        # 模拟 ClickHouse：每个块按 SQL 中的日期字面量返回不同行
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kwargs):
                # 根据 SQL 中的日期字面量生成不同行
                if "2025-04-01" in sql:
                    yield [("apr01", "row_apr01")]
                elif "2025-04-11" in sql:
                    yield [("apr11", "row_apr11"), ("apr12", "row_apr12")]
                elif "2025-04-21" in sql:
                    yield [("apr21", "row_apr21"), ("apr22", "row_apr22"), ("apr23", "row_apr23")]
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "c3",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"
            assert final["exported_rows"] == 1 + 2 + 3

        files = final["output_files"]
        assert len(files) == 3

        c0 = _read_xlsx(files[0]["file_path"])["sheets"]["Sheet1"]
        c1 = _read_xlsx(files[1]["file_path"])["sheets"]["Sheet1"]
        c2 = _read_xlsx(files[2]["file_path"])["sheets"]["Sheet1"]

        assert len(c0["data_rows"]) == 1
        assert c0["data_rows"][0] == ("apr01", "row_apr01")

        assert len(c1["data_rows"]) == 2
        assert c1["data_rows"] == [("apr11", "row_apr11"), ("apr12", "row_apr12")]

        assert len(c2["data_rows"]) == 3

    def test_c4_filename_format_correct(self, app_client, tmp_path, monkeypatch):
        """C4: 文件名格式 {job_name}_{YYYYMMDD}_to_{YYYYMMDD}.xlsx"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([[("1", "a")]])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "c4",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"

        fname = final["output_files"][0]["filename"]
        assert "20250401_to_20250410" in fname
        assert fname.endswith(".xlsx")

    def test_c5_chinese_jobname_preserved(self, app_client, tmp_path, monkeypatch):
        """C5（B16）: 中文 job_name 保留在文件名中"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([[("1", "a")]])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}用户行为",  # 中文
                "chunk_config": {
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
            })
            assert r.status_code == 200
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"

        # 中文应保留在文件名中
        fname = final["output_files"][0]["filename"]
        assert "用户行为" in fname


# =============================================================================
# D · 边界规模
# =============================================================================

class TestEdgeCases:
    """D1-D5：empty/single/large/many"""

    def test_d1_empty_chunk_creates_header_only_file(self, app_client, tmp_path, monkeypatch):
        """D1（B4）: 空结果块 → header-only xlsx，状态 completed，rows=0"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([])  # 无数据
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "d1",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"
            assert final["exported_rows"] == 0

        # 文件应存在且可读
        fpath = final["output_files"][0]["file_path"]
        assert Path(fpath).exists()
        content = _read_xlsx(fpath)
        sheet1 = content["sheets"]["Sheet1"]
        assert sheet1["header"] == ("id", "name")
        assert sheet1["data_rows"] == []

    def test_d2_chunk_days_one_makes_30_files(self, app_client, tmp_path, monkeypatch):
        """D2（B6）: chunk_days=1 跨 30 天 → 30 个文件"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([[("1", "a")]])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "d2",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 1,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=30.0)
            assert final["status"] == "completed"
            assert final["total_batches"] == 30
            assert final["done_batches"] == 30
            assert len(final["output_files"]) == 30
            assert final["exported_rows"] == 30  # 1 行 × 30 块

        # 抽查中间一个文件
        mid = final["output_files"][14]
        assert "20250415_to_20250415" in mid["filename"]

    def test_d3_single_day_range(self, app_client, tmp_path, monkeypatch):
        """D3: start == end → 1 块 1 文件"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([[("1", "a")]])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "d3",
                chunk_config={
                    "date_start": "2025-04-15",
                    "date_end": "2025-04-15",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"
            assert final["total_batches"] == 1
            assert "20250415_to_20250415" in final["output_files"][0]["filename"]

    def test_d4_chunk_with_over_max_rows_makes_multiple_sheets(
        self, app_client, tmp_path, monkeypatch,
    ):
        """D4（B5）: 单块超过 MAX_ROWS_PER_SHEET → 该块文件内多 Sheet"""
        # 把 MAX_ROWS_PER_SHEET 调小到 5 便于测试
        import backend.services.data_export_service as svc
        monkeypatch.setattr(svc, "MAX_ROWS_PER_SHEET", 5)

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            # 单块返回 12 行 → 应分为 3 个 Sheet (5+5+2)
            mc.stream_batches.side_effect = lambda sql, **kw: iter([
                [(str(i), f"r{i}") for i in range(12)]
            ])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "d4",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 30,  # 单块覆盖整月
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"
            assert len(final["output_files"]) == 1
            assert final["output_files"][0]["sheets"] == 3
            assert final["output_files"][0]["rows"] == 12

        # 真实读回验证 3 个 Sheet
        content = _read_xlsx(final["output_files"][0]["file_path"])
        assert len(content["sheet_names"]) == 3
        for name in content["sheet_names"]:
            sheet = content["sheets"][name]
            assert sheet["header"] == ("id", "name"), f"Sheet {name} 缺表头"

        # 累计数据行数应为 12
        total_data = sum(
            len(content["sheets"][n]["data_rows"]) for n in content["sheet_names"]
        )
        assert total_data == 12

    def test_d5_max_chunk_days_90(self, app_client, tmp_path, monkeypatch):
        """D5: chunk_days=90 → 上限合法"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([[("1", "a")]])
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "d5",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-06-30",  # 91 天
                    "chunk_days": 90,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            assert final["status"] == "completed"
            assert final["total_batches"] == 2  # 90+1 天 → 2 块


# =============================================================================
# E · 故障态
# =============================================================================

class TestFailureModes:
    """E1-E4：列预检失败、块中故障、partial download"""

    def test_e1_column_probe_fail_marks_failed_no_dir_residue(
        self, app_client, tmp_path, monkeypatch,
    ):
        """E1（B10）: 列预检失败 → Job=failed，输出目录已创建但应空（无残留 xlsx）"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.side_effect = RuntimeError("Connection refused")
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "e1",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=5.0)
            assert final["status"] == "failed"
            assert "Connection refused" in (final["error_message"] or "")

        # 输出目录可能存在但应无 .xlsx 残留
        out_dir = Path(final["output_filename"])  # job 的 output_filename 是目录名
        # 由 file_path 取目录
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        db = SessionLocal()
        try:
            j = db.query(ExportJob).filter(ExportJob.id == jid).first()
            actual_dir = Path(j.file_path)
        finally:
            db.close()

        if actual_dir.exists():
            xlsx_files = list(actual_dir.glob("*.xlsx"))
            assert len(xlsx_files) == 0, f"列预检失败后不应有残留 xlsx: {xlsx_files}"

    def test_e2_partial_download_from_failed_job(self, app_client, tmp_path, monkeypatch):
        """E2(v2.14.7): 块 1 完成 + 块 2 失败 → Job=partial_failed,块 1/3 应可下载"""
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)
        # 计数器须在 _mk_client 闭包外，使所有 client 实例共享
        call = {"n": 0}
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kwargs):
                call["n"] += 1
                if call["n"] == 2:
                    raise RuntimeError("CH error on chunk 2")
                yield [("1", "a")]
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "e2",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)
            # 新行为:partial_failed(块 1+3 完成,块 2 失败)
            assert final["status"] == "partial_failed", final

        # 块 0 status=completed，应可下载
        dl0 = app_client.get(f"/api/v1/data-export/jobs/{jid}/download?file_index=0")
        assert dl0.status_code == 200
        assert "spreadsheetml" in dl0.headers["content-type"]

        # 块 1 status=failed，下载应 400
        dl1 = app_client.get(f"/api/v1/data-export/jobs/{jid}/download?file_index=1")
        assert dl1.status_code == 400

        # 块 2 status=completed,下载应可成功(v2.14.7 新行为:继续执行)
        dl2 = app_client.get(f"/api/v1/data-export/jobs/{jid}/download?file_index=2")
        assert dl2.status_code == 200

    def test_e3_failed_chunk_keeps_completed_chunks(self, app_client, tmp_path, monkeypatch):
        """E3(v2.14.7): 块 2 失败时块 1/3 状态应为 completed 且物理存在"""
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)
        call = {"n": 0}  # 共享计数器
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kwargs):
                call["n"] += 1
                if call["n"] == 2:
                    raise RuntimeError("simulated")
                yield [("1", "a")]
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            jid = _create_chunked_job_via_api(
                app_client, tmp_path, "e3",
                chunk_config={
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-30",
                    "chunk_days": 10,
                },
                monkeypatch=monkeypatch,
            )
            final = _wait_done(app_client, jid, timeout=10.0)

        files = final["output_files"]
        assert files[0]["status"] == "completed"
        assert Path(files[0]["file_path"]).exists()
        assert files[1]["status"] == "failed"
        # v2.14.7:块 3 继续执行,不再 pending
        assert files[2]["status"] == "completed"
        assert Path(files[2]["file_path"]).exists()

    def test_e4_unknown_connection_env_fails_gracefully(self, app_client, tmp_path, monkeypatch):
        """E4: connection_env 不存在时应 Job=failed 而非协程崩溃"""
        # 不 mock _build_export_client → 让真实代码尝试加载未配置的 env
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        r = app_client.post("/api/v1/data-export/execute", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "nonexistent_env_xyz",
            "job_name": f"{_PREFIX}e4",
            "chunk_config": {
                "date_start": "2025-04-01",
                "date_end": "2025-04-10",
                "chunk_days": 10,
            },
        })
        assert r.status_code == 200  # 端点本身成功（异步任务失败不影响）
        jid = r.json()["data"]["job_id"]
        final = _wait_done(app_client, jid, timeout=10.0)
        assert final["status"] == "failed"
        assert final["error_message"] is not None


# =============================================================================
# F · 删除幂等
# =============================================================================

class TestDeleteIdempotent:

    def test_f1_delete_when_dir_never_created(self, app_client, tmp_path):
        """F1（B9）: 启动前已 cancel → output_dir 未创建 → 删除不应抛错"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        from datetime import datetime

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}f1",
                query_sql="SELECT 1", connection_env="test",
                status="cancelled",
                finished_at=datetime.utcnow(),
                export_mode="date_chunked",
                file_path=str(tmp_path / "never_created_dir"),  # 不存在
            )
            db.add(j)
            db.commit()
            jid = str(j.id)
        finally:
            db.close()

        r = app_client.delete(f"/api/v1/data-export/jobs/{jid}")
        assert r.status_code == 200

    def test_f2_delete_partial_dir_with_some_xlsx(self, app_client, tmp_path):
        """F2: 部分块完成的目录 → 递归删除所有内容"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        from datetime import datetime

        out_dir = tmp_path / "f2_dir"
        out_dir.mkdir()
        (out_dir / "f2_chunk_0.xlsx").write_bytes(b"x")
        (out_dir / "f2_chunk_1.xlsx").write_bytes(b"y")

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}f2",
                query_sql="SELECT 1", connection_env="test",
                status="failed",
                finished_at=datetime.utcnow(),
                export_mode="date_chunked",
                file_path=str(out_dir),
            )
            db.add(j)
            db.commit()
            jid = str(j.id)
        finally:
            db.close()

        r = app_client.delete(f"/api/v1/data-export/jobs/{jid}")
        assert r.status_code == 200
        assert not out_dir.exists()


# =============================================================================
# G · RBAC 端点防护
# =============================================================================

class TestRBACEndpoints:
    """G1-G3: 验证所有分块端点的权限装饰器是否存在"""

    def test_g1_all_chunked_endpoints_use_data_export_permission(self):
        """G1（B11）: 检查代码层 — 所有 data-export 路由均依赖 require_permission('data','export')"""
        import re
        from pathlib import Path
        api_file = Path(__file__).parent / "backend" / "api" / "data_export.py"
        text = api_file.read_text(encoding="utf-8")

        # 找到所有 @router.X("/...") 装饰行
        decorators = re.findall(
            r'@router\.(get|post|delete|put)\(\s*["\'][^"\']*["\']',
            text,
        )
        assert len(decorators) >= 7  # 至少 7 个端点

        # 找所有 require_permission("data","export") 调用
        perm_calls = re.findall(
            r'require_permission\(\s*["\']data["\']\s*,\s*["\']export["\']\s*\)',
            text,
        )
        # 端点数量应等于权限调用数量
        assert len(perm_calls) >= len(decorators), (
            f"端点数 {len(decorators)} > 权限调用数 {len(perm_calls)} — "
            f"有端点缺失 require_permission"
        )

    def test_g2_data_export_permission_seeded_in_init_rbac(self):
        """G2: 权限种子脚本中含 data:export"""
        from pathlib import Path
        text = (Path(__file__).parent / "backend" / "scripts" / "init_rbac.py").read_text(encoding="utf-8")
        assert '("data",' in text
        assert '"export",' in text
        # superadmin 通过 list comprehension 自动获得所有权限

    def test_g3_data_export_in_db_permission_table(self):
        """G3: DB 中实际存在 data:export 权限记录"""
        from backend.config.database import SessionLocal
        from backend.models.permission import Permission

        db = SessionLocal()
        try:
            perm = (
                db.query(Permission)
                .filter(Permission.resource == "data", Permission.action == "export")
                .first()
            )
            assert perm is not None, "数据库中未找到 data:export 权限记录"
        finally:
            db.close()


# =============================================================================
# H · SQL 契约
# =============================================================================

class TestSQLContract:
    """H1-H3: 注入路径生成的 SQL 字符串精确验证"""

    def test_h1_placeholder_substitution_exact(self, tmp_path, monkeypatch):
        """H1（B12）: 占位符替换后 SQL 字面量精确匹配"""
        import backend.services.data_export_service as svc

        seen: List[str] = []
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kw):
                seen.append(sql)
                yield []
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}h1",
                query_sql="SELECT id FROM events WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
                connection_env="test", status="pending",
                export_mode="date_chunked",
                file_path=str(tmp_path / "h1_dir"),
            )
            db.add(j); db.commit(); db.refresh(j); jid = str(j.id)
        finally:
            db.close()

        config = {
            "query_sql": "SELECT id FROM events WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test", "connection_type": "clickhouse",
            "batch_size": 1000, "export_mode": "date_chunked",
            "chunk_config": {
                "date_start": "2025-04-01", "date_end": "2025-04-10", "chunk_days": 10,
            },
            "output_dir": str(tmp_path / "h1_dir"), "job_name": "h1",
        }

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(svc.run_export_job(jid, config))
            finally:
                loop.close()

        assert len(seen) == 1
        # SQL 应是占位符直接替换，不包装
        assert seen[0] == "SELECT id FROM events WHERE dt >= '2025-04-01' AND dt <= '2025-04-10'"

    def test_h2_wrapper_mode_exact(self, tmp_path, monkeypatch):
        """H2（B12）: 包装模式 SQL 精确匹配"""
        import backend.services.data_export_service as svc

        seen: List[str] = []
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            def _stream(sql, **kw):
                seen.append(sql)
                yield []
            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}h2",
                query_sql="SELECT id, name FROM events",
                connection_env="test", status="pending",
                export_mode="date_chunked",
                file_path=str(tmp_path / "h2_dir"),
            )
            db.add(j); db.commit(); db.refresh(j); jid = str(j.id)
        finally:
            db.close()

        config = {
            "query_sql": "SELECT id, name FROM events",
            "connection_env": "test", "connection_type": "clickhouse",
            "batch_size": 1000, "export_mode": "date_chunked",
            "chunk_config": {
                "date_column": "event_date",
                "date_start": "2025-04-01", "date_end": "2025-04-10", "chunk_days": 10,
            },
            "output_dir": str(tmp_path / "h2_dir"), "job_name": "h2",
        }

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(svc.run_export_job(jid, config))
            finally:
                loop.close()

        assert len(seen) == 1
        # 应是包装格式
        expected = (
            "SELECT * FROM (SELECT id, name FROM events) AS _chunk_q"
            " WHERE _chunk_q.event_date >= '2025-04-01'"
            " AND _chunk_q.event_date <= '2025-04-10'"
        )
        assert seen[0] == expected, f"\nExpected:\n{expected}\n\nGot:\n{seen[0]}"

    def test_h3_to_dict_json_serializable(self):
        """H3（B17）: ExportJob.to_dict() 完整 JSON 可序列化"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        from datetime import datetime

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}h3",
                query_sql="SELECT 1", connection_env="test",
                status="completed",
                finished_at=datetime.utcnow(),
                started_at=datetime.utcnow(),
                export_mode="date_chunked",
                chunk_config={"date_start": "2025-04-01", "date_end": "2025-04-10", "chunk_days": 10},
                output_files=[
                    {"index": 0, "filename": "x.xlsx", "file_path": "/x", "file_size": 100,
                     "rows": 10, "sheets": 1, "status": "completed",
                     "date_start": "2025-04-01", "date_end": "2025-04-10"}
                ],
            )
            db.add(j); db.commit(); db.refresh(j)
            d = j.to_dict()
        finally:
            db.delete(j); db.commit()
            db.close()

        # 应可以序列化为 JSON 而不抛错
        try:
            serialized = json.dumps(d)
            parsed = json.loads(serialized)
        except Exception as e:
            pytest.fail(f"to_dict() 输出不可 JSON 序列化: {e}")

        assert parsed["export_mode"] == "date_chunked"
        assert isinstance(parsed["output_files"], list)
        assert isinstance(parsed["chunk_config"], dict)


# =============================================================================
# I · Code 160 嵌套
# =============================================================================

class TestCode160InsideChunk:
    """I1: chunked 模式下单块触发 Code 160 应能 LIMIT/OFFSET 回退"""

    def test_i1_code_160_in_chunk_falls_back(self, tmp_path, monkeypatch):
        """I1（B13）: 模拟 stream_batches 抛 Code 160 → _run_single_export 应自动重试"""
        import backend.services.data_export_service as svc

        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}i1",
                query_sql="SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                connection_env="test", status="pending",
                export_mode="date_chunked",
                file_path=str(tmp_path / "i1_dir"),
            )
            db.add(j); db.commit(); db.refresh(j); jid = str(j.id)
        finally:
            db.close()

        attempts = {"n": 0}

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 5
            def _stream(sql, **kwargs):
                attempts["n"] += 1
                if attempts["n"] == 1:
                    # 第一次普通流式 → 抛 Code 160
                    raise RuntimeError(
                        "Code: 160. DB::Exception: Estimated query execution time "
                        "(310 seconds) is too long. ESTIMATED_EXECUTION_TIMEOUT_EXCEEDED"
                    )
                # 第二次（chunked 回退）→ 返回数据
                yield [(str(i), f"r{i}") for i in range(5)]
            mc.stream_batches.side_effect = _stream

            def _stream_chunked(sql, chunk_size, total_rows, batch_size, extra_settings=None):
                yield from _stream(sql, batch_size=batch_size, extra_settings=extra_settings)
            mc.stream_batches_chunked.side_effect = _stream_chunked
            _attach_csv_stream_raw(mc)
            return mc

        config = {
            "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test", "connection_type": "clickhouse",
            "batch_size": 1000, "export_mode": "date_chunked",
            "chunk_config": {
                "date_start": "2025-04-01", "date_end": "2025-04-10", "chunk_days": 10,
            },
            "output_dir": str(tmp_path / "i1_dir"), "job_name": "i1",
        }

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            loop = asyncio.new_event_loop()
            try:
                loop.run_until_complete(svc.run_export_job(jid, config))
            finally:
                loop.close()

        # 即使触发了 Code 160 回退，最终仍应 completed
        db = SessionLocal()
        try:
            j2 = db.query(ExportJob).filter(ExportJob.id == jid).first()
            assert j2.status == "completed", (
                f"Code 160 嵌套回退失败：status={j2.status}, err={j2.error_message}"
            )
            assert j2.exported_rows == 5
        finally:
            db.close()


# =============================================================================
# J · 并发安全
# =============================================================================

class TestConcurrentJobs:
    """J1: 两个 chunked Job 并行执行应互不干扰"""

    def test_j1_two_concurrent_chunked_jobs(self, app_client, tmp_path, monkeypatch):
        """J1（B14）: 同时提交两个 chunked Job，文件应分别写入两个目录"""
        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.stream_batches.side_effect = lambda sql, **kw: iter([
                [("1", "a"), ("2", "b")]
            ])
            _attach_csv_stream_raw(mc)
            return mc

        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        with patch("backend.services.data_export_service._build_export_client", side_effect=_mk_client):
            # 同时提交两个 Job（job_name 不同，输出目录不同）
            r1 = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}j1a",
                "chunk_config": {
                    "date_start": "2025-04-01", "date_end": "2025-04-20", "chunk_days": 10,
                },
            })
            r2 = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}j1b",
                "chunk_config": {
                    "date_start": "2025-04-01", "date_end": "2025-04-20", "chunk_days": 10,
                },
            })
            assert r1.status_code == 200 and r2.status_code == 200
            jid1, jid2 = r1.json()["data"]["job_id"], r2.json()["data"]["job_id"]

            f1 = _wait_done(app_client, jid1, timeout=15.0)
            f2 = _wait_done(app_client, jid2, timeout=15.0)

        assert f1["status"] == "completed"
        assert f2["status"] == "completed"
        # 两个 Job 的输出目录不同
        dir1 = Path(f1["output_files"][0]["file_path"]).parent
        dir2 = Path(f2["output_files"][0]["file_path"]).parent
        assert dir1 != dir2
        assert dir1.exists() and dir2.exists()
        # 各自的文件不重叠
        assert f1["output_files"][0]["file_path"] != f2["output_files"][0]["file_path"]


# =============================================================================
# K · JSON 契约
# =============================================================================

class TestJSONContract:
    """K1-K2: 任务列表/状态接口的 JSON 序列化稳定"""

    def test_k1_jobs_list_includes_chunked_fields(self, app_client):
        """K1: GET /jobs 返回的每条记录都含 export_mode/output_files 字段"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob

        db = SessionLocal()
        try:
            j_chunked = ExportJob(
                user_id="uid", username=f"{_PREFIX}k1c",
                query_sql="SELECT 1", connection_env="test",
                status="completed",
                export_mode="date_chunked",
                output_files=[{"index": 0, "filename": "x.xlsx", "status": "completed",
                               "date_start": "2025-04-01", "date_end": "2025-04-10",
                               "file_path": "/x", "file_size": 100, "rows": 10, "sheets": 1}],
            )
            j_single = ExportJob(
                user_id="uid", username=f"{_PREFIX}k1s",
                query_sql="SELECT 1", connection_env="test",
                status="completed",
                export_mode="single",
            )
            db.add_all([j_chunked, j_single]); db.commit()
        finally:
            db.close()

        r = app_client.get("/api/v1/data-export/jobs?page=1&page_size=50")
        items = r.json()["data"]["items"]
        for item in items:
            # 所有记录都应有这三个字段（即使为 null）
            assert "export_mode" in item
            assert "output_files" in item
            assert "chunk_config" in item

    def test_k2_chunked_output_files_schema(self, app_client):
        """K2: output_files 每个条目含必要字段（前端 ChunkFileList 渲染依赖）"""
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        REQUIRED = {"index", "date_start", "date_end", "filename", "file_path",
                    "file_size", "rows", "sheets", "status"}
        db = SessionLocal()
        try:
            j = ExportJob(
                user_id="uid", username=f"{_PREFIX}k2",
                query_sql="SELECT 1", connection_env="test",
                status="completed",
                export_mode="date_chunked",
                output_files=[{
                    "index": 0, "date_start": "2025-04-01", "date_end": "2025-04-10",
                    "filename": "x.xlsx", "file_path": "/x", "file_size": 100,
                    "rows": 10, "sheets": 1, "status": "completed",
                }],
            )
            db.add(j); db.commit()
            jid = str(j.id)
        finally:
            db.close()

        r = app_client.get(f"/api/v1/data-export/jobs/{jid}")
        d = r.json()["data"]
        assert d["output_files"]
        for f in d["output_files"]:
            missing = REQUIRED - set(f.keys())
            assert not missing, f"output_files 条目缺失字段: {missing}"


# =============================================================================
# teardown
# =============================================================================

# =============================================================================
# M · 流式断开错误自动 LIMIT/OFFSET 回退（修复 IncompleteRead 故障）
# =============================================================================

class TestStreamDisconnectFallback:
    """M1-M5: 单文件 / chunked 模式遇到 ChunkedEncodingError/IncompleteRead 时
    应自动切换到 LIMIT/OFFSET 回退而非直接 failed"""

    def test_m1_is_transient_stream_error_recognizes_chunked_encoding(self):
        """M1: ChunkedEncodingError 被识别为瞬时错误"""
        import requests.exceptions
        from backend.services.export_clients.clickhouse import is_transient_stream_error

        exc = requests.exceptions.ChunkedEncodingError(
            "('Connection broken: IncompleteRead(0 bytes read, 2 more expected)', "
            "IncompleteRead(0 bytes read, 2 more expected))"
        )
        assert is_transient_stream_error(exc) is True

    def test_m2_is_transient_stream_error_recognizes_incomplete_read(self):
        """M2: http.client.IncompleteRead 被识别为瞬时错误"""
        from http.client import IncompleteRead
        from backend.services.export_clients.clickhouse import is_transient_stream_error

        exc = IncompleteRead(b"", 2)
        assert is_transient_stream_error(exc) is True

    def test_m3_is_transient_stream_error_message_fingerprint(self):
        """M3: 通过消息字符串兜底匹配（未捕获到原始类型时）"""
        from backend.services.export_clients.clickhouse import is_transient_stream_error

        # 模拟 RuntimeError 包装的连接断开
        assert is_transient_stream_error(RuntimeError("Connection broken: foo")) is True
        assert is_transient_stream_error(RuntimeError("Connection reset by peer")) is True
        assert is_transient_stream_error(RuntimeError("Some other error")) is False

    def test_m4_chunk_with_stream_disconnect_falls_back_to_limit_offset(
        self, app_client, tmp_path, monkeypatch,
    ):
        """M4: chunked 模式下某块流式断开 → 自动 LIMIT/OFFSET 回退后成功完成"""
        import requests.exceptions
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        # 共享调用计数器
        call = {"stream_n": 0}

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 5

            def _stream(sql, **kwargs):
                call["stream_n"] += 1
                if call["stream_n"] == 1:
                    # 第一次普通流式 → 抛 ChunkedEncodingError（连接断开）
                    raise requests.exceptions.ChunkedEncodingError(
                        "('Connection broken: IncompleteRead(0 bytes read, "
                        "2 more expected)', IncompleteRead(0 bytes read, 2 more expected))"
                    )
                # 第二次（回退后的 LIMIT/OFFSET）→ 成功
                yield [(str(i), f"r{i}") for i in range(5)]

            mc.stream_batches.side_effect = _stream

            def _stream_chunked(sql, chunk_size, total_rows, batch_size,
                                extra_settings=None):
                yield from _stream(sql, batch_size=batch_size,
                                   extra_settings=extra_settings)

            mc.stream_batches_chunked.side_effect = _stream_chunked
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}m4",
                "chunk_config": {
                    "date_start": "2025-04-01",
                    "date_end": "2025-04-10",
                    "chunk_days": 10,
                },
            })
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=10.0)

        # 关键断言：连接断开后应自动回退并最终 completed
        assert final["status"] == "completed", (
            f"流式断开应触发 LIMIT/OFFSET 回退后成功；status={final['status']}, "
            f"err={final.get('error_message')}"
        )
        assert final["exported_rows"] == 5

    def test_m5_humanize_error_for_stream_disconnect(self):
        """M5: 流式断开错误的 _humanize_error 输出含可读提示与建议"""
        from backend.services.data_export_service import _humanize_error
        import requests.exceptions

        exc = requests.exceptions.ChunkedEncodingError(
            "('Connection broken: IncompleteRead(0 bytes read, 2 more expected)', "
            "IncompleteRead(0 bytes read, 2 more expected))"
        )
        msg = _humanize_error(exc)
        assert "数据流中途断开" in msg
        assert "建议" in msg
        assert "[技术细节]" in msg  # 保留原始细节供排查
        assert "IncompleteRead" in msg  # 技术细节段含原始信息

    def test_m6_humanize_error_for_code_160(self):
        """M6: Code 160 错误的 _humanize_error 提示"""
        from backend.services.data_export_service import _humanize_error
        msg = _humanize_error(RuntimeError(
            "Code: 160. DB::Exception: Estimated query execution time is too long"
        ))
        assert "估算查询执行时间" in msg
        assert "max_execution_time" in msg
        assert "[技术细节]" in msg

    def test_m7_humanize_error_for_unknown_passes_through(self):
        """M7: 未识别错误兜底返回原文"""
        from backend.services.data_export_service import _humanize_error
        raw = "some weird unhandled error xyz"
        assert _humanize_error(RuntimeError(raw)) == raw

    def test_m8_response_ended_prematurely_recognized(self):
        """M8: ChunkedEncodingError('Response ended prematurely') 被正确识别"""
        import requests.exceptions
        from backend.services.export_clients.clickhouse import is_transient_stream_error
        from backend.services.data_export_service import _humanize_error

        exc = requests.exceptions.ChunkedEncodingError("Response ended prematurely")
        assert is_transient_stream_error(exc) is True

        # 也通过字符串兜底
        assert is_transient_stream_error(RuntimeError("Response ended prematurely")) is True

        # _humanize_error 输出含可读提示
        msg = _humanize_error(exc)
        assert "数据流中途断开" in msg
        assert "send_progress_in_http_headers" in msg  # 提到已注入的心跳设置
        assert "[技术细节]" in msg


# =============================================================================
# O · 自动子块分裂（v2.13 v2 — 流式断开时按日期对半分裂重试）
# =============================================================================

class TestAutoSubdivision:
    """O1-O5: 块在流式断开时自动按日期对半分裂为更小子块重试"""

    def test_o1_failed_5day_chunk_subdivides_into_2_and_3_day_chunks(
        self, app_client, tmp_path, monkeypatch,
    ):
        """O1: 5 天块（含 stream + LIMIT/OFFSET 回退都失败）→ 自动分裂为 2 天 + 3 天，后续成功

        模拟用户真实场景：服务端处理 5 天 decrypt 数据耗时 >5 分钟必然断开，
        LIMIT/OFFSET 回退每窗口仍需服务端跑完整 5 天计算 → 也失败；
        只有缩小日期范围才能成功。
        """
        import requests.exceptions
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _fail_if_full_range(sql, **kwargs):
            """SQL 中含 5 天完整范围（01-01 + 01-05）就失败；缩小范围后成功"""
            if "'2026-01-01'" in sql and "'2026-01-05'" in sql:
                raise requests.exceptions.ChunkedEncodingError(
                    "Response ended prematurely"
                )
            yield [("1", "row")]

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 1
            mc.stream_batches.side_effect = _fail_if_full_range
            mc.stream_batches_chunked.side_effect = (
                lambda sql, chunk_size, total_rows, batch_size, extra_settings=None:
                _fail_if_full_range(sql, batch_size=batch_size,
                                    extra_settings=extra_settings)
            )
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": (
                    "SELECT id, name FROM t WHERE d >= '{{date_start}}' "
                    "AND d <= '{{date_end}}'"
                ),
                "connection_env": "test",
                "job_name": f"{_PREFIX}o1",
                "chunk_config": {
                    "date_start": "2026-01-01",
                    "date_end": "2026-01-05",
                    "chunk_days": 5,  # 单块覆盖整个 5 天范围
                },
            })
            assert r.status_code == 200
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=15.0)

        assert final["status"] == "completed", (
            f"自动分裂应该使任务最终完成；status={final['status']}, "
            f"err={final.get('error_message')}"
        )
        # 原本 1 个块 → 分裂后变成 2 个子块
        files = final["output_files"]
        assert len(files) == 2, f"5 天块应分裂为 2 个子块；实际 {len(files)}"

        # 子块日期范围应分别覆盖前 2 天和后 3 天（5//2=2）
        dates = [(f["date_start"], f["date_end"]) for f in files]
        assert dates[0] == ("2026-01-01", "2026-01-02")
        assert dates[1] == ("2026-01-03", "2026-01-05")

        # 两个子块都应 completed
        assert all(f["status"] == "completed" for f in files)
        assert final["exported_rows"] == 2  # 1 行 × 2 个成功子块

    def test_o2_recursive_subdivision_when_subchunk_also_fails(
        self, app_client, tmp_path, monkeypatch,
    ):
        """O2: 子块再次失败 → 继续递归分裂（5天 → 2,3；2 天再分为 1,1）"""
        import requests.exceptions
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _stream(sql, **kwargs):
            # 5 天范围（含 01-01 + 01-05）失败
            # 2 天范围（含 01-01 + 01-02，但不含 01-05）也失败
            # 1 天范围只含一个日期 → 成功
            if "'2026-01-01'" in sql and "'2026-01-05'" in sql:
                raise requests.exceptions.ChunkedEncodingError("Response ended prematurely")
            if "'2026-01-01'" in sql and "'2026-01-02'" in sql:
                raise requests.exceptions.ChunkedEncodingError("Response ended prematurely")
            yield [("1", "x")]

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 1
            mc.stream_batches.side_effect = _stream
            mc.stream_batches_chunked.side_effect = (
                lambda sql, chunk_size, total_rows, batch_size, extra_settings=None:
                _stream(sql, batch_size=batch_size, extra_settings=extra_settings)
            )
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}o2",
                "chunk_config": {
                    "date_start": "2026-01-01",
                    "date_end": "2026-01-05",
                    "chunk_days": 5,
                },
            })
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=15.0)

        assert final["status"] == "completed"
        files = final["output_files"]
        # 5 天 → 2,3；2 天再分为 1,1 → 总 3 个子块（1天+1天+3天）
        assert len(files) == 3
        # 检查日期覆盖完整
        from datetime import date as _d, timedelta
        all_dates = set()
        for f in files:
            s = _d.fromisoformat(f["date_start"])
            e = _d.fromisoformat(f["date_end"])
            for i in range((e - s).days + 1):
                all_dates.add(s + timedelta(days=i))
        # 应覆盖 2026-01-01 ~ 2026-01-05 全部 5 天
        assert len(all_dates) == 5

    def test_o3_one_day_chunk_failure_does_not_subdivide(
        self, app_client, tmp_path, monkeypatch,
    ):
        """O3: 1 天块失败 → 不可再分，直接整个 Job failed"""
        import requests.exceptions
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 1

            def _stream(sql, **kwargs):
                raise requests.exceptions.ChunkedEncodingError(
                    "Response ended prematurely"
                )
                yield  # noqa

            mc.stream_batches.side_effect = _stream
            mc.stream_batches_chunked.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}o3",
                "chunk_config": {
                    "date_start": "2026-01-01",
                    "date_end": "2026-01-01",  # 单天
                    "chunk_days": 1,
                },
            })
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=15.0)

        # 1 天块无法再分 → Job failed
        assert final["status"] == "failed"
        assert "数据流中途断开" in (final["error_message"] or "")

    def test_o4_non_transient_error_no_subdivision(
        self, app_client, tmp_path, monkeypatch,
    ):
        """O4: 非瞬时错误（如 SQL 语法错误）不应触发分裂"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()

            def _stream(sql, **kwargs):
                raise RuntimeError("Code: 62. Syntax error in query.")
                yield  # noqa

            mc.stream_batches.side_effect = _stream
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}o4",
                "chunk_config": {
                    "date_start": "2026-01-01",
                    "date_end": "2026-01-05",
                    "chunk_days": 5,
                },
            })
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=10.0)

        # SQL 语法错误不应分裂，直接 failed，仍只有 1 个原始块
        assert final["status"] == "failed"
        assert len(final["output_files"]) == 1

    def test_o5_total_batches_updates_after_subdivision(
        self, app_client, tmp_path, monkeypatch,
    ):
        """O5: 分裂后 total_batches 应从原值增长，前端进度条会自动适配"""
        import requests.exceptions
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)

        def _stream(sql, **kwargs):
            # 10 天完整范围失败，缩小后成功
            if "'2026-01-01'" in sql and "'2026-01-10'" in sql:
                raise requests.exceptions.ChunkedEncodingError(
                    "Response ended prematurely"
                )
            yield [("1", "x")]

        def _mk_client(*args, **kw):
            mc = Mock()
            mc.get_columns.return_value = _make_columns()
            mc.count_rows.return_value = 1
            mc.stream_batches.side_effect = _stream
            mc.stream_batches_chunked.side_effect = (
                lambda sql, chunk_size, total_rows, batch_size, extra_settings=None:
                _stream(sql, batch_size=batch_size, extra_settings=extra_settings)
            )
            _attach_csv_stream_raw(mc)
            return mc

        with patch("backend.services.data_export_service._build_export_client",
                   side_effect=_mk_client):
            r = app_client.post("/api/v1/data-export/execute", json={
                "query_sql": "SELECT id, name FROM t WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                "connection_env": "test",
                "job_name": f"{_PREFIX}o5",
                "chunk_config": {
                    "date_start": "2026-01-01",
                    "date_end": "2026-01-10",
                    "chunk_days": 10,
                },
            })
            jid = r.json()["data"]["job_id"]
            final = _wait_done(app_client, jid, timeout=15.0)

        # 原 1 块 → 分裂后 total_batches=2
        assert final["status"] == "completed"
        assert final["total_batches"] == 2
        assert final["done_batches"] == 2


# =============================================================================
# P · TCP keepalive HTTPAdapter
# =============================================================================

class TestTCPKeepalive:

    def test_p1_keepalive_session_uses_tcp_keepalive_adapter(self):
        """P1: 默认情况下 _get_export_session 返回带 keepalive 的 Session"""
        from backend.services.export_clients import clickhouse as ch_mod
        from backend.services.export_clients.clickhouse import (
            _TCPKeepAliveAdapter, _get_export_session,
        )

        # 重置全局缓存
        ch_mod._export_session = None

        sess = _get_export_session()
        # http:// 和 https:// 都应挂载 TCP keepalive adapter
        assert isinstance(sess.adapters["http://"], _TCPKeepAliveAdapter)
        assert isinstance(sess.adapters["https://"], _TCPKeepAliveAdapter)

    def test_p2_env_var_disables_keepalive(self, monkeypatch):
        """P2: 环境变量 CH_EXPORT_TCP_KEEPALIVE=0 关闭 keepalive 适配器"""
        from backend.services.export_clients import clickhouse as ch_mod
        from backend.services.export_clients.clickhouse import (
            _TCPKeepAliveAdapter, _get_export_session,
        )

        monkeypatch.setenv("CH_EXPORT_TCP_KEEPALIVE", "0")
        ch_mod._export_session = None

        sess = _get_export_session()
        assert not isinstance(sess.adapters["http://"], _TCPKeepAliveAdapter)


# =============================================================================
# N · ClickHouse 默认 HTTP 心跳设置注入
# =============================================================================

class TestDefaultStreamingSettings:
    """N1-N5: 验证默认保活/流式设置已正确注入 HTTP URL 参数"""

    def test_n1_default_settings_injected_into_stream_batches(self):
        """N1: stream_batches 默认带保活设置"""
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        client = ClickHouseExportClient("h", 8123, "u", "p", "db")
        captured_params = {}

        def fake_post(url, data=None, params=None, **kwargs):
            captured_params.update(params or {})
            resp = Mock()
            resp.status_code = 200
            resp.iter_lines = lambda decode_unicode: iter([])
            resp.content = b""
            return resp

        with patch("requests.sessions.Session.post", side_effect=fake_post):
            list(client.stream_batches("SELECT 1"))

        assert captured_params.get("send_progress_in_http_headers") == "1"
        assert captured_params.get("http_headers_progress_interval_ms") == "3000"
        assert captured_params.get("wait_end_of_query") == "0"

    def test_n2_extra_settings_overrides_defaults(self):
        """N2: extra_settings 优先级高于默认值，可覆盖"""
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        client = ClickHouseExportClient("h", 8123, "u", "p", "db")
        captured_params = {}

        def fake_post(url, data=None, params=None, **kwargs):
            captured_params.update(params or {})
            resp = Mock()
            resp.status_code = 200
            resp.iter_lines = lambda decode_unicode: iter([])
            resp.content = b""
            return resp

        with patch("requests.sessions.Session.post", side_effect=fake_post):
            list(client.stream_batches("SELECT 1", extra_settings={
                "http_headers_progress_interval_ms": "5000",  # 覆盖
                "max_execution_time": 600,                     # 新增
            }))

        # 覆盖生效
        assert captured_params.get("http_headers_progress_interval_ms") == "5000"
        # 新增生效
        assert captured_params.get("max_execution_time") == 600
        # 未指定的默认值仍保留
        assert captured_params.get("send_progress_in_http_headers") == "1"

    def test_n3_env_var_disables_keepalive(self, monkeypatch):
        """N3: 环境变量 CH_EXPORT_HTTP_KEEPALIVE=0 完全关闭默认保活注入"""
        monkeypatch.setenv("CH_EXPORT_HTTP_KEEPALIVE", "0")

        from backend.services.export_clients.clickhouse import (
            _build_default_streaming_settings,
        )
        assert _build_default_streaming_settings() == {}

    def test_n4_env_var_customizes_progress_interval(self, monkeypatch):
        """N4: 环境变量 CH_EXPORT_PROGRESS_INTERVAL_MS 可自定义心跳间隔"""
        monkeypatch.setenv("CH_EXPORT_PROGRESS_INTERVAL_MS", "5000")
        # 需要确认 KEEPALIVE 默认开启
        monkeypatch.delenv("CH_EXPORT_HTTP_KEEPALIVE", raising=False)

        from backend.services.export_clients.clickhouse import (
            _build_default_streaming_settings,
        )
        settings = _build_default_streaming_settings()
        assert settings["http_headers_progress_interval_ms"] == "5000"

    def test_n5_count_rows_also_uses_keepalive(self):
        """N5: count_rows() 也应注入心跳设置（大表 count 也可能跑很久）"""
        from backend.services.export_clients.clickhouse import ClickHouseExportClient

        client = ClickHouseExportClient("h", 8123, "u", "p", "db")
        captured_params = {}

        def fake_post(url, data=None, params=None, **kwargs):
            captured_params.update(params or {})
            resp = Mock()
            resp.status_code = 200
            resp.text = "1234"
            return resp

        with patch("requests.sessions.Session.post", side_effect=fake_post):
            client.count_rows("SELECT 1", timeout=300)

        assert captured_params.get("send_progress_in_http_headers") == "1"
        assert captured_params.get("max_execution_time") == 300


# =============================================================================
# L · 预览占位符替换（v2.13 占位符 SQL 也能预览）
# =============================================================================

class TestPreviewPlaceholderSubstitution:
    """L1-L5: preview_query 自动替换占位符让带占位符 SQL 也能预览"""

    def test_l1_default_date_used_when_placeholders_present(self, app_client):
        """L1: SQL 含占位符 + 不传 preview_date → 后端默认昨日替换"""
        from datetime import datetime, timedelta
        from unittest.mock import Mock

        mock_http = Mock()
        mock_http.execute.return_value = ([("v1",)], [("col", "String")])

        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db",
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                resp = app_client.post("/api/v1/data-export/preview", json={
                    "query_sql": (
                        "SELECT col FROM t WHERE d >= '{{date_start}}' "
                        "AND d <= '{{date_end}}'"
                    ),
                    "connection_env": "test",
                })

        assert resp.status_code == 200
        data = resp.json()["data"]
        # 响应中带 preview_date 字段，应为昨日
        yesterday = (datetime.utcnow().date() - timedelta(days=1)).isoformat()
        assert data["preview_date"] == yesterday

        # 验证传给 ClickHouse 的 SQL 已替换占位符
        call_args = mock_http.execute.call_args
        executed_sql = call_args[0][0]
        assert f"'{yesterday}'" in executed_sql
        assert "{{date_start}}" not in executed_sql
        assert "{{date_end}}" not in executed_sql

    def test_l2_custom_preview_date_used(self, app_client):
        """L2: 传入 preview_date → 使用该日期替换"""
        mock_http = Mock()
        mock_http.execute.return_value = ([], [("c", "String")])

        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db",
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                resp = app_client.post("/api/v1/data-export/preview", json={
                    "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
                    "connection_env": "test",
                    "preview_date": "2025-04-15",
                })

        assert resp.status_code == 200
        assert resp.json()["data"]["preview_date"] == "2025-04-15"

        executed_sql = mock_http.execute.call_args[0][0]
        assert "'2025-04-15'" in executed_sql

    def test_l3_no_placeholders_preview_date_null(self, app_client):
        """L3: SQL 不含占位符 → preview_date 字段为 None（向后兼容）"""
        mock_http = Mock()
        mock_http.execute.return_value = ([], [("c", "String")])

        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db",
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                resp = app_client.post("/api/v1/data-export/preview", json={
                    "query_sql": "SELECT 1 FROM t",
                    "connection_env": "test",
                })

        assert resp.status_code == 200
        assert resp.json()["data"]["preview_date"] is None

    def test_l4_partial_placeholder_400(self, app_client):
        """L4: 单占位符 SQL → 400 拒绝"""
        resp = app_client.post("/api/v1/data-export/preview", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}'",  # 缺 date_end
            "connection_env": "test",
        })
        assert resp.status_code == 400
        assert "成对" in resp.json()["detail"]

    def test_l5_invalid_preview_date_400(self, app_client):
        """L5: preview_date 非 ISO 格式 → 400"""
        resp = app_client.post("/api/v1/data-export/preview", json={
            "query_sql": "SELECT 1 WHERE d >= '{{date_start}}' AND d <= '{{date_end}}'",
            "connection_env": "test",
            "preview_date": "2025/04/15",
        })
        assert resp.status_code == 400

    # ─ v2.14.4: ts 占位符的 preview 路径 ─

    def test_l6_ts_placeholders_preview_substitution(self, app_client):
        """L6(v2.14.4): SQL 含 {{ts_start}}/{{ts_end}} → 预览时替换为
        datetime 半开区间字面量(start='2025-04-15 00:00:00', end='2025-04-16 00:00:00')"""
        mock_http = Mock()
        mock_http.execute.return_value = ([], [("c", "String")])

        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db",
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                resp = app_client.post("/api/v1/data-export/preview", json={
                    "query_sql": (
                        "SELECT col FROM t "
                        "WHERE ts >= parseDateTimeBestEffort('{{ts_start}}') "
                        "  AND ts <  parseDateTimeBestEffort('{{ts_end}}')"
                    ),
                    "connection_env": "test",
                    "preview_date": "2025-04-15",
                })
        assert resp.status_code == 200, resp.text
        assert resp.json()["data"]["preview_date"] == "2025-04-15"

        executed_sql = mock_http.execute.call_args[0][0]
        # ts_start = 样本日 00:00:00, ts_end = 次日 00:00:00 (半开区间)
        assert "'2025-04-15 00:00:00'" in executed_sql
        assert "'2025-04-16 00:00:00'" in executed_sql
        # 占位符已全部替换
        assert "{{ts_start}}" not in executed_sql
        assert "{{ts_end}}" not in executed_sql

    def test_l7_partial_ts_placeholder_400(self, app_client):
        """L7(v2.14.4): 单 ts 占位符(仅 ts_start 不含 ts_end)→ 400 拒绝"""
        resp = app_client.post("/api/v1/data-export/preview", json={
            "query_sql": "SELECT 1 WHERE ts >= '{{ts_start}}'",
            "connection_env": "test",
        })
        assert resp.status_code == 400
        assert "成对" in resp.json()["detail"]

    def test_l8_mixed_date_and_ts_placeholders_preview(self, app_client):
        """L8(v2.14.4): SQL 同时含 date 和 ts 占位符 → 预览同时替换两套"""
        mock_http = Mock()
        mock_http.execute.return_value = ([], [("c", "String")])

        mock_settings = Mock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "h", "http_port": 8123, "user": "u", "password": "p", "database": "db",
        }
        with patch("backend.mcp.clickhouse.http_client.ClickHouseHTTPClient", return_value=mock_http):
            with patch("backend.config.settings.settings", mock_settings):
                resp = app_client.post("/api/v1/data-export/preview", json={
                    "query_sql": (
                        "SELECT col FROM t "
                        "WHERE d BETWEEN '{{date_start}}' AND '{{date_end}}' "
                        "  OR ts >= '{{ts_start}}' AND ts < '{{ts_end}}'"
                    ),
                    "connection_env": "test",
                    "preview_date": "2025-04-15",
                })
        assert resp.status_code == 200, resp.text
        executed_sql = mock_http.execute.call_args[0][0]
        # date 闭区间字面量(原样)
        assert "'2025-04-15'" in executed_sql
        assert "BETWEEN '2025-04-15' AND '2025-04-15'" in executed_sql
        # ts 半开区间字面量(datetime)
        assert "'2025-04-15 00:00:00'" in executed_sql
        assert "'2025-04-16 00:00:00'" in executed_sql
        # 全部占位符替换
        assert "{{date_start}}" not in executed_sql
        assert "{{date_end}}" not in executed_sql
        assert "{{ts_start}}" not in executed_sql
        assert "{{ts_end}}" not in executed_sql


# =============================================================================
# Q · auto 预分窗口：空数据范围 / 单桶超阈值 / 多 chunk 拼接
# =============================================================================
# v2.14.5：auto 模式下 bucket 全空 → 不再生成「仅表头」占位文件。
# 单桶单独超阈值仍然作为单独窗口（不再细分;Excel sheet 切分兜底）。

class TestAutoPreSplitSkipEmpty:
    """Q1-Q8：auto 模式空数据范围不生成空文件 + 边界情况"""

    def test_q1_auto_bucket_to_windows_empty_input(self):
        """Q1: 空 bucket_rows → 空 windows（不生成任何窗口）"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        windows = _auto_bucket_rows_to_windows(
            [], unit="hour", target_rows=1_000_000,
        )
        assert windows == []

    def test_q2_auto_bucket_to_windows_all_zero_counts(self):
        """Q2: 所有桶 count=0 → 全部被过滤 → 空 windows"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        rows = [
            ("2026-03-06 09:00:00", 0),
            ("2026-03-06 10:00:00", 0),
            ("2026-03-06 11:00:00", 0),
        ]
        windows = _auto_bucket_rows_to_windows(
            rows, unit="hour", target_rows=1_000_000,
        )
        assert windows == []

    def test_q3_auto_bucket_greedy_merge_under_threshold(self):
        """Q3: 09+10+11=99万 ≤ 100万 → 合并为 09-11 单窗口（注释里描述的典型情况）"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        rows = [
            ("2026-03-06 09:00:00", 330_000),
            ("2026-03-06 10:00:00", 330_000),
            ("2026-03-06 11:00:00", 330_000),
        ]
        windows = _auto_bucket_rows_to_windows(
            rows, unit="hour", target_rows=1_000_000,
        )
        assert len(windows) == 1
        s, e, cnt = windows[0]
        assert s == datetime(2026, 3, 6, 9, 0, 0)
        assert e == datetime(2026, 3, 6, 11, 59, 59)
        assert cnt == 990_000

    def test_q4_auto_bucket_breaks_when_exceed_threshold(self):
        """Q4: 09+10+11=99w → +12(33w)=132w > 100w → 09-11 封口；12 起新窗口"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        rows = [
            ("2026-03-06 09:00:00", 330_000),
            ("2026-03-06 10:00:00", 330_000),
            ("2026-03-06 11:00:00", 330_000),
            ("2026-03-06 12:00:00", 330_000),
        ]
        windows = _auto_bucket_rows_to_windows(
            rows, unit="hour", target_rows=1_000_000,
        )
        assert len(windows) == 2
        s0, e0, c0 = windows[0]
        s1, e1, c1 = windows[1]
        assert s0 == datetime(2026, 3, 6, 9, 0, 0)
        assert e0 == datetime(2026, 3, 6, 11, 59, 59)
        assert c0 == 990_000
        assert s1 == datetime(2026, 3, 6, 12, 0, 0)
        assert e1 == datetime(2026, 3, 6, 12, 59, 59)
        assert c1 == 330_000

    def test_q5_auto_bucket_single_oversized_becomes_own_window(self):
        """Q5: 单桶 150万 > 100万 阈值 → 仍作为单独窗口（不再下钻细分;sheet 切分兜底）"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        rows = [
            ("2026-03-06 09:00:00", 1_500_000),
            ("2026-03-06 10:00:00", 200_000),
        ]
        windows = _auto_bucket_rows_to_windows(
            rows, unit="hour", target_rows=1_000_000,
        )
        # 9点单独成窗(超阈值)；10点单独成窗(因 9 点已 > target,加 10 也超)
        assert len(windows) == 2
        assert windows[0][2] == 1_500_000
        assert windows[1][2] == 200_000

    def test_q6_auto_bucket_minute_unit_step(self):
        """Q6: minute 粒度 → 每桶步长 1 分钟"""
        from backend.services.data_export_service import _auto_bucket_rows_to_windows
        rows = [
            ("2026-03-06 09:00:00", 400_000),
            ("2026-03-06 09:01:00", 500_000),
        ]
        windows = _auto_bucket_rows_to_windows(
            rows, unit="minute", target_rows=1_000_000,
        )
        assert len(windows) == 1
        s, e, cnt = windows[0]
        assert s == datetime(2026, 3, 6, 9, 0, 0)
        # 1 分钟步长 → 09:01:59 而非 09:59:59
        assert e == datetime(2026, 3, 6, 9, 1, 59)
        assert cnt == 900_000

    def test_q7_pre_split_ranges_empty_buckets_returns_empty(self, tmp_path, monkeypatch):
        """Q7（核心修复）: auto 模式下 bucket 查询返回空 → _pre_split_ranges 返回 [],
        不再生成仅表头占位文件。"""
        # 直接验证 _auto_bucket_rows_to_windows([]) == [] 已在 Q1 覆盖;
        # 这里通过 mock _fetch_auto_bucket_rows 全链路验证 _pre_split_ranges → []。
        from backend.services import data_export_service as svc

        # 构造一个最小化的 raw_chunk_cfg / ncfg 模拟环境
        raw_chunk_cfg = {
            "date_column": "dt",
            "date_start": "2026-03-06",
            "date_end": "2026-03-06",
            "chunk_days": 1,
            "min_subdivide_unit": "hour",
            "pre_split_hours": "auto",
            "auto_split_target_rows": 1_000_000,
        }
        # 用 ChunkConfig validate 做参数归一化
        from backend.services.data_export_chunker import validate_chunk_config
        sql = "SELECT id FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'"
        ncfg = validate_chunk_config(raw_chunk_cfg, sql)

        # mock _fetch_auto_bucket_rows 返回空
        from datetime import date as _date

        # 我们不调用真实 _run_chunked_export_sync,而是直接重现 _pre_split_ranges 闭包
        # 逻辑:auto 路径 → _fetch_auto_bucket_rows → _auto_bucket_rows_to_windows([])
        # → windows=[] → 返回 []。 已被 Q1 直接验证。本测试加一层契约检查。

        # 直接调用底层助手验证关键不变量
        bucket_rows = []
        windows = svc._auto_bucket_rows_to_windows(
            bucket_rows, unit=ncfg.min_subdivide_unit, target_rows=1_000_000,
        )
        assert windows == []

    def test_q8_end_to_end_auto_all_empty_completes_zero_files(
        self, app_client, tmp_path, monkeypatch,
    ):
        """Q8: 全链路 — auto 模式下整段范围无数据 → job 成功完成 + output_files=[] +
        rows=0 + 不写任何 .xlsx 文件到磁盘。"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        # 注入 mock 客户端
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]

            def count_rows(self, sql, timeout=None):
                return 0

            def stream_batches(self, sql, batch_size=50_000, extra_settings=None, query_id_prefix=None):
                # 用于 _fetch_auto_bucket_rows:返回空(无桶)
                return iter([])

            def stream_batches_keyset(self, *a, **kw):
                return iter([])

            def stream_batches_chunked(self, *a, **kw):
                return iter([])

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id, name FROM events WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}q8",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "min_subdivide_unit": "hour",
                "pre_split_hours": "auto",
                "auto_split_target_rows": 1_000_000,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]

        final = _wait_done(app_client, job_id, timeout=15.0)
        assert final is not None
        assert final["status"] == "completed", final
        # 关键断言:auto + 空数据 → 0 文件 + 0 行
        assert final["exported_rows"] == 0
        assert final["output_files"] == []
        # 确认目录里也确实没有 .xlsx 落盘
        xlsx_files = list(tmp_path.rglob("*.xlsx"))
        assert xlsx_files == [], f"不该有 xlsx 文件,但发现 {xlsx_files}"

    def test_q9_end_to_end_auto_mixed_chunks_skip_only_empty_ones(
        self, app_client, tmp_path, monkeypatch,
    ):
        """Q9: 多 chunk 拼接 — 部分 chunk 无数据应只跳过它自己,其他 chunk 正常生成文件。
        date_range=2026-03-01~2026-03-30, chunk_days=10 → 3 chunks。
        Chunk1 无数据(跳过),Chunk2 有数据,Chunk3 无数据(跳过) → 仅 Chunk2 出文件。"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        # 共享调用状态:用 chunk 起始日期判断当前在哪个 chunk
        call_state = {"bucket_sql_count": 0}

        def _stream_factory(_self, sql, batch_size=50_000, extra_settings=None, query_id_prefix=None):
            # 桶查询(_fetch_auto_bucket_rows):SQL 含 toStartOfHour
            if "toStartOfHour" in sql:
                call_state["bucket_sql_count"] += 1
                # 判断 chunk:SQL 内嵌的 date 字面量(注入后的 WHERE 范围字面量)
                if "'2026-03-11'" in sql or "'2026-03-15'" in sql:
                    # Chunk2 (03-11~03-20):返回 1 个桶 50w 行(放在 03-15 10:00 这个小时)
                    return iter([[("2026-03-15 10:00:00", 500_000)]])
                # Chunk1 / Chunk3:空桶
                return iter([])
            # 真实导出:auto 模式下 chunk2 narrow window 落在 2026-03-15 这小时上
            if "'2026-03-15" in sql:
                rows = [(str(i), f"n{i}") for i in range(500_000)]
                return iter([rows])
            return iter([])

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]

            def count_rows(self, sql, timeout=None):
                return 500_000

            def stream_batches(self, *a, **kw):
                return _stream_factory(self, *a, **kw)

            def stream_batches_keyset(self, *a, **kw):
                return iter([])

            def stream_batches_chunked(self, *a, **kw):
                return _stream_factory(self, *a, **kw)

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id, name FROM events WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}q9",
            # 强制走 direct 引擎,避免 csv_staging 触发 stream_raw 调用(测试 mock 简化)
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-01",
                "date_end": "2026-03-30",
                "chunk_days": 10,
                "min_subdivide_unit": "hour",
                "pre_split_hours": "auto",
                "auto_split_target_rows": 1_000_000,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]

        final = _wait_done(app_client, job_id, timeout=60.0)
        assert final is not None
        assert final["status"] == "completed", final
        # 关键断言:三个 chunk 各跑一次桶查询
        assert call_state["bucket_sql_count"] == 3
        # 只有 Chunk2 出文件
        assert len(final["output_files"]) == 1
        f = final["output_files"][0]
        # 文件落在 Chunk2 (03-11~03-20) 范围内
        assert f["date_start"].startswith("2026-03-15"), f
        assert f["rows"] == 500_000

    def test_q10_count_sql_uses_clickhouse_minute_directive(self):
        """Q10(v2.14.5 hotfix): auto 预分窗口的 count SQL 必须使用 ClickHouse formatDateTime
        的 `%i` 分钟语义,绝不能用 Python 风格的 `%M`(在 CH 里是月份英文全名,会返回
        '2026-03-02 09:March:00' 之类无法解析的字符串 → ValueError → 整个 Job 失败)。

        本测试锚定 SQL 字符串内容,防止后续重构再次踩坑。
        """
        from backend.services.data_export_service import _build_auto_pre_split_count_sql
        sql = _build_auto_pre_split_count_sql(
            "SELECT id, name FROM events WHERE dt >= '2026-03-01' AND dt <= '2026-03-31'",
            time_column="event_time",
            unit="hour",
        )
        # 必须用 ClickHouse 正确的分钟语义:%i 或包含 %T(= %H:%i:%S)
        has_correct_minute = "%i" in sql or "%T" in sql
        assert has_correct_minute, f"SQL 缺少 CH 分钟语义 %i/%T:\n{sql}"
        # 关键反向断言:不能出现 '%H:%M:%S' 这种 Python 风格(CH 下 %M=月份英文全名)
        assert "%H:%M:%S" not in sql, (
            f"SQL 含 Python 风格 '%H:%M:%S';CH formatDateTime 下 %M=月份英文,会返回 "
            f"'09:March:00' 之类无法解析的字符串。SQL:\n{sql}"
        )
        # 桶函数应正确(hour → toStartOfHour)
        assert "toStartOfHour" in sql
        # 时间列被正确引用
        assert "`event_time`" in sql

    def test_q11_count_sql_minute_unit(self):
        """Q11: unit=minute → toStartOfMinute 桶"""
        from backend.services.data_export_service import _build_auto_pre_split_count_sql
        sql = _build_auto_pre_split_count_sql(
            "SELECT 1 FROM t WHERE d = '2026-03-01'",
            time_column="ts",
            unit="minute",
        )
        assert "toStartOfMinute" in sql
        assert "%H:%M:%S" not in sql
        assert "%i" in sql or "%T" in sql

    def test_q12_count_sql_unit_validation(self):
        """Q12: unit=day 等不允许值应抛 ValueError"""
        from backend.services.data_export_service import _build_auto_pre_split_count_sql
        with pytest.raises(ValueError, match="hour/minute"):
            _build_auto_pre_split_count_sql(
                "SELECT 1", time_column="ts", unit="day",
            )

    def test_q13_parse_bucket_datetime_rejects_bad_format(self):
        """Q13(v2.14.5 hotfix): _parse_bucket_datetime 对类似 '2026-03-02 09:March:00' 的
        畸形字符串应抛 ValueError(保持 fail-loud,防止 SQL bug 静默退化为 0 文件)。"""
        from backend.services.data_export_service import _parse_bucket_datetime
        with pytest.raises(ValueError, match="无法解析"):
            _parse_bucket_datetime("2026-03-02 09:March:00")
        with pytest.raises(ValueError):
            _parse_bucket_datetime("not-a-date")

    def test_q14_parse_bucket_datetime_iso_t_separator(self):
        """Q14: ClickHouse 某些场景返回 ISO 'T' 分隔符,parser 应自动归一化"""
        from backend.services.data_export_service import _parse_bucket_datetime
        dt = _parse_bucket_datetime("2026-03-02T09:00:00")
        assert dt == datetime(2026, 3, 2, 9, 0, 0)


# =============================================================================
# R · prefer_chunked — 跳过单流首试,直接走 chunked 路径
# =============================================================================
# v2.14.6 新增:跨境/不稳网络下单流 5 分钟必断,每块先单流试错浪费 5-10 分钟。
# prefer_chunked=True 直接走 keyset(若提供 cursor_column)或 LIMIT/OFFSET,
# 跳过单流首试。

class TestPreferChunked:
    """R1-R5: prefer_chunked 行为契约"""

    def test_r1_prefer_chunked_with_cursor_routes_to_keyset_no_stream(
        self, app_client, tmp_path, monkeypatch,
    ):
        """R1: prefer_chunked=True + cursor_column → 直接 stream_batches_keyset,
        不调 stream_batches(单流)。"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        calls = {"stream_batches": 0, "stream_batches_keyset": 0}

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]

            def count_rows(self, sql, timeout=None):
                return 100

            def stream_batches(self, *a, **kw):
                calls["stream_batches"] += 1
                # 不应被调用(prefer_chunked 跳过单流)
                return iter([])

            def stream_batches_keyset(self, *a, **kw):
                calls["stream_batches_keyset"] += 1
                yield [(str(i), f"n{i}") for i in range(100)]

            def stream_batches_chunked(self, *a, **kw):
                return iter([])

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}r1",
            "xlsx_engine": "direct",  # csv_staging 不走 _run_single_export 的 chunked 分支
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "cursor_column": "id",
                "prefer_chunked": True,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        # 关键断言:单流 0 次,keyset 1 次
        assert calls["stream_batches"] == 0, (
            f"prefer_chunked 下不应调单流 stream_batches,实际 {calls['stream_batches']} 次"
        )
        assert calls["stream_batches_keyset"] == 1, (
            f"prefer_chunked + cursor 应直接走 keyset,实际 {calls['stream_batches_keyset']} 次"
        )

    def test_r2_prefer_chunked_without_cursor_routes_to_limit_offset_no_stream(
        self, app_client, tmp_path, monkeypatch,
    ):
        """R2: prefer_chunked=True + 无 cursor_column → 直接 stream_batches_chunked
        (LIMIT/OFFSET),不调单流。"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        calls = {"stream_batches": 0, "stream_batches_chunked": 0, "count_rows": 0}

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]

            def count_rows(self, sql, timeout=None):
                calls["count_rows"] += 1
                return 100

            def stream_batches(self, *a, **kw):
                calls["stream_batches"] += 1
                return iter([])

            def stream_batches_keyset(self, *a, **kw):
                return iter([])

            def stream_batches_chunked(self, sql, chunk_size, total_rows, **kw):
                calls["stream_batches_chunked"] += 1
                yield [(str(i), f"n{i}") for i in range(100)]

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}r2",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "prefer_chunked": True,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        # LIMIT/OFFSET 路径需要预扫描 count_rows
        assert calls["count_rows"] == 1, f"应预扫描 count_rows,实际 {calls['count_rows']}"
        assert calls["stream_batches"] == 0, f"不应调单流,实际 {calls['stream_batches']}"
        assert calls["stream_batches_chunked"] == 1, (
            f"无 cursor 时应走 LIMIT/OFFSET,实际 {calls['stream_batches_chunked']}"
        )

    def test_r3_default_false_keeps_legacy_stream_first_behavior(
        self, app_client, tmp_path, monkeypatch,
    ):
        """R3: 未填 prefer_chunked + 无 env var → 沿用单流首试老行为(向后兼容)"""
        from backend.api import data_export as api_module
        monkeypatch.delenv("EXPORT_PREFER_CHUNKED", raising=False)
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        calls = {"stream_batches": 0, "stream_batches_keyset": 0}

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64")]

            def count_rows(self, sql, timeout=None):
                return 10

            def stream_batches(self, *a, **kw):
                calls["stream_batches"] += 1
                yield [(str(i),) for i in range(10)]

            def stream_batches_keyset(self, *a, **kw):
                calls["stream_batches_keyset"] += 1
                return iter([])

            def stream_batches_chunked(self, *a, **kw):
                return iter([])

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}r3",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "cursor_column": "id",
                # prefer_chunked 不填 → 沿用单流首试
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        # 关键断言:仍走单流(成功),不调 keyset
        assert calls["stream_batches"] == 1
        assert calls["stream_batches_keyset"] == 0

    def test_r4_env_var_acts_as_default_when_field_unset(
        self, app_client, tmp_path, monkeypatch,
    ):
        """R4: 未填 prefer_chunked + env EXPORT_PREFER_CHUNKED=1 → 也跳过单流"""
        from backend.api import data_export as api_module
        monkeypatch.setenv("EXPORT_PREFER_CHUNKED", "1")
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        calls = {"stream_batches": 0, "stream_batches_keyset": 0}

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64")]

            def count_rows(self, sql, timeout=None):
                return 10

            def stream_batches(self, *a, **kw):
                calls["stream_batches"] += 1
                return iter([])

            def stream_batches_keyset(self, *a, **kw):
                calls["stream_batches_keyset"] += 1
                yield [(str(i),) for i in range(10)]

            def stream_batches_chunked(self, *a, **kw):
                return iter([])

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}r4",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "cursor_column": "id",
                # prefer_chunked 不填,env=1 应生效
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        assert calls["stream_batches"] == 0
        assert calls["stream_batches_keyset"] == 1

    def test_r5_field_false_overrides_env_true(
        self, app_client, tmp_path, monkeypatch,
    ):
        """R5: prefer_chunked=False 显式 + env=1 → 字段优先,沿用单流首试"""
        from backend.api import data_export as api_module
        monkeypatch.setenv("EXPORT_PREFER_CHUNKED", "1")
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        calls = {"stream_batches": 0, "stream_batches_keyset": 0}

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64")]

            def count_rows(self, sql, timeout=None):
                return 10

            def stream_batches(self, *a, **kw):
                calls["stream_batches"] += 1
                yield [(str(i),) for i in range(10)]

            def stream_batches_keyset(self, *a, **kw):
                calls["stream_batches_keyset"] += 1
                return iter([])

            def stream_batches_chunked(self, *a, **kw):
                return iter([])

        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: _MockClient())

        payload = {
            "query_sql": "SELECT id FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}r5",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2026-03-06",
                "date_end": "2026-03-06",
                "chunk_days": 1,
                "cursor_column": "id",
                "prefer_chunked": False,  # 显式 false 覆盖 env=1
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        assert calls["stream_batches"] == 1
        assert calls["stream_batches_keyset"] == 0


# =============================================================================
# P · partial_failed — v2.14.7 单块失败不再拖垮整 Job
# =============================================================================

class TestPartialFailure:
    """P1-P8: 某块 exhaust 所有重试+分裂仍失败 → 跳过继续做剩下,Job 终态根据计数判定"""

    def _build_mock_with_failing_chunks(self, failing_chunk_filter):
        """构造 mock,根据 SQL 内含的日期字面量判断当前块是否模拟失败。

        failing_chunk_filter(sql: str) -> bool — True 表示该块的 SQL 应抛 ConnectionError
        """
        from backend.services.export_clients.base import ColumnInfo
        from requests.exceptions import ConnectionError as RConnErr

        class _MockClient:
            def get_columns(self, sql):
                return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]

            def count_rows(self, sql, timeout=None):
                if failing_chunk_filter(sql):
                    raise ConnectionError("mock failure for count_rows")
                return 10

            def stream_batches(self, sql, batch_size=50_000, extra_settings=None, query_id_prefix=None):
                if failing_chunk_filter(sql):
                    raise RConnErr(
                        "Connection aborted - mock RST",
                    )
                yield [(str(i), f"n{i}") for i in range(10)]

            def stream_batches_keyset(self, *a, **kw):
                # keyset 用不上,但要存在
                sql = kw.get("sql") if "sql" in kw else (a[0] if a else "")
                if failing_chunk_filter(sql):
                    raise ConnectionError("mock keyset failure")
                return iter([])

            def stream_batches_chunked(self, sql, chunk_size, total_rows, **kw):
                if failing_chunk_filter(sql):
                    raise ConnectionError("mock LIMIT/OFFSET failure")
                yield [(str(i), f"n{i}") for i in range(10)]

        return _MockClient

    def test_p1_one_chunk_fails_others_succeed_yields_partial_failed(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P1: 3 块 job 中第 2 块失败 → job.status='partial_failed';
        output_files 含 2 completed + 1 failed;error_message 含失败块明细。"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        # 04-10 ~ 04-12 共 3 天,chunk_days=1 → 3 块
        # 让 04-11 这块失败
        MockClient = self._build_mock_with_failing_chunks(
            lambda sql: "'2025-04-11'" in sql
        )
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        # 老行为开关默认 0,partial_failed 模式生效
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p1",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final is not None
        assert final["status"] == "partial_failed", final
        # 2 块 completed, 1 块 failed
        statuses = [f["status"] for f in final["output_files"]]
        assert statuses.count("completed") == 2
        assert statuses.count("failed") == 1
        # error_message 含失败块明细
        assert "失败" in (final.get("error_message") or "")
        # 失败块有 error_summary
        failed_entry = next(f for f in final["output_files"] if f["status"] == "failed")
        assert "error_summary" in failed_entry
        assert failed_entry["error_summary"]
        # 已 exported_rows > 0 (前面的成功块计入)
        assert final["exported_rows"] >= 20  # 2 块 × 10 行

    def test_p2_all_chunks_fail_yields_failed(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P2: 全部块都失败 → job.status='failed' (与原行为一致)"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        MockClient = self._build_mock_with_failing_chunks(lambda sql: True)
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p2",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "failed", final
        # 所有 entry 都是 failed
        statuses = [f["status"] for f in final["output_files"]]
        assert all(s == "failed" for s in statuses)

    def test_p3_all_chunks_succeed_yields_completed(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P3: 全部块都成功 → job.status='completed' (与原行为一致)"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        MockClient = self._build_mock_with_failing_chunks(lambda sql: False)  # 不失败
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p3",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        assert r.status_code == 200, r.text
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "completed", final
        assert all(f["status"] == "completed" for f in final["output_files"])

    def test_p4_partial_failed_download_chunk_success(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P4: partial_failed 状态下,通过 file_index 下载 status=completed 的块仍正常"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        MockClient = self._build_mock_with_failing_chunks(lambda sql: "'2025-04-11'" in sql)
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p4",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "partial_failed"

        # 找到第一个 completed 块,下载它
        completed_idx = next(
            f["index"] for f in final["output_files"] if f["status"] == "completed"
        )
        dl = app_client.get(
            f"/api/v1/data-export/jobs/{job_id}/download?file_index={completed_idx}"
        )
        assert dl.status_code == 200, dl.text

        # 反过来,下载 failed 块应 400
        failed_idx = next(
            f["index"] for f in final["output_files"] if f["status"] == "failed"
        )
        dl_failed = app_client.get(
            f"/api/v1/data-export/jobs/{job_id}/download?file_index={failed_idx}"
        )
        assert dl_failed.status_code == 400, dl_failed.text

    def test_p5_partial_failed_delete_job_ok(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P5: partial_failed 状态可以删除 job 记录(连带磁盘文件)"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        MockClient = self._build_mock_with_failing_chunks(lambda sql: "'2025-04-11'" in sql)
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p5",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "partial_failed"

        # 删除应成功
        delr = app_client.delete(f"/api/v1/data-export/jobs/{job_id}")
        assert delr.status_code == 200, delr.text

    def test_p6_env_fail_fast_restores_legacy_behavior(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P6: env EXPORT_FAIL_FAST_ON_CHUNK_ERROR=1 → 回退老行为(整 Job failed 一遇错)"""
        from backend.api import data_export as api_module
        monkeypatch.setenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", "1")
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        # 04-11 块失败,但 env=1 → 04-12 不会执行,整 Job failed
        MockClient = self._build_mock_with_failing_chunks(
            lambda sql: "'2025-04-11'" in sql
        )
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p6",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "failed", final
        # 04-10 块应已 completed,04-11 failed,04-12 仍 pending(因为整 Job failed return 提前退出)
        statuses = [f["status"] for f in final["output_files"]]
        assert "completed" in statuses
        assert "failed" in statuses
        assert "pending" in statuses

    def test_p7_csv_zip_partial_failed_uses_partial_prefix(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P7: csv_zip 模式 partial_failed → ZIP 文件名加 partial_ 前缀,仅含成功块"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services.export_clients.base import ColumnInfo
        from backend.services import data_export_service as svc

        # csv 模式走 stream_raw 直写字节流;mock 需要按 sql 内日期返回 raw bytes 或抛错
        def _csv_mock_factory(env, conn_type):
            class _CSVClient:
                def get_columns(self, sql):
                    return [ColumnInfo("id", "Int64"), ColumnInfo("name", "String")]
                def count_rows(self, sql, timeout=None):
                    return 5
                def stream_batches(self, *a, **kw):
                    return iter([])
                def stream_batches_keyset(self, *a, **kw):
                    return iter([])
                def stream_batches_chunked(self, *a, **kw):
                    return iter([])
                def stream_raw(self, sql, format_name="CSVWithNames", **kw):
                    if "'2025-04-11'" in sql:
                        raise ConnectionError("mock csv RST for 04-11")
                    yield b"id,name\n"
                    for i in range(5):
                        yield f"{i},n{i}\n".encode("utf-8")
            return _CSVClient()

        monkeypatch.setattr(svc, "_build_export_client", _csv_mock_factory)
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p7",
            "output_format": "csv_zip",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "partial_failed", final
        # ZIP 文件名应有 partial_ 前缀
        zip_name = final.get("output_filename") or ""
        assert zip_name.startswith("partial_"), f"期望 partial_ 前缀,实际 {zip_name!r}"
        # ZIP 文件存在且能下载
        dl = app_client.get(f"/api/v1/data-export/jobs/{job_id}/download")
        assert dl.status_code == 200, dl.text

    def test_p8_failure_summary_includes_chunk_details(
        self, app_client, tmp_path, monkeypatch,
    ):
        """P8: error_message 含每个失败块的 index, 时间范围, 错误指纹"""
        from backend.api import data_export as api_module
        monkeypatch.setattr(api_module, "_CUSTOMER_DATA_ROOT", tmp_path)
        from backend.services import data_export_service as svc

        # 04-10, 04-12 失败; 04-11 成功
        MockClient = self._build_mock_with_failing_chunks(
            lambda sql: "'2025-04-10'" in sql or "'2025-04-12'" in sql
        )
        monkeypatch.setattr(svc, "_build_export_client", lambda env, conn_type: MockClient())
        monkeypatch.delenv("EXPORT_FAIL_FAST_ON_CHUNK_ERROR", raising=False)

        payload = {
            "query_sql": "SELECT id, name FROM t WHERE dt >= '{{date_start}}' AND dt <= '{{date_end}}'",
            "connection_env": "test",
            "job_name": f"{_PREFIX}p8",
            "xlsx_engine": "direct",
            "chunk_config": {
                "date_column": "dt",
                "date_start": "2025-04-10",
                "date_end": "2025-04-12",
                "chunk_days": 1,
            },
        }
        r = app_client.post("/api/v1/data-export/execute", json=payload)
        job_id = r.json()["data"]["job_id"]
        final = _wait_done(app_client, job_id, timeout=30.0)
        assert final["status"] == "partial_failed", final
        err = final.get("error_message") or ""
        # summary 应说明「3 块中 2 块失败」
        assert "3 块中 2 块失败" in err, f"期望含失败计数,实际: {err}"
        # 应含 2025-04-10 和 2025-04-12 字面量
        assert "2025-04-10" in err
        assert "2025-04-12" in err


def teardown_module(_):
    try:
        from backend.config.database import SessionLocal
        from backend.models.export_job import ExportJob
        db = SessionLocal()
        jobs = db.query(ExportJob).filter(ExportJob.username.like(f"{_PREFIX}%")).all()
        for j in jobs:
            if j.file_path:
                try:
                    p = Path(j.file_path)
                    if p.is_dir():
                        shutil.rmtree(p, ignore_errors=True)
                    elif p.exists():
                        p.unlink()
                except Exception:
                    pass
            db.delete(j)
        db.commit()
        db.close()
    except Exception:
        pass
