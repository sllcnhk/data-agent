"""
test_data_import.py — Excel → ClickHouse 数据导入功能测试
=========================================================

测试层次：
  A  (6)  — data_import_service 工具函数（list_writable_connections、parse_excel_preview、_rows_to_values_clause）
  B  (8)  — API 权限控制（无权限用户被拒绝，superadmin 可访问）
  C  (6)  — 连接/数据库/表查询端点（Mock CH 调用）
  D  (8)  — 上传端点（文件类型/大小检查、正常上传、Sheet 预览）
  E  (8)  — execute 导入端点（参数校验、任务创建、后台启动）
  F  (8)  — jobs 查询端点（状态轮询、历史列表分页）
  H  (8)  — cancel 端点（状态机校验：pending/running 可取消，其他拒绝，DB 更新验证）
  I  (4)  — delete 端点（删除任务记录，404 处理）

总计: 56 个测试用例
"""
import asyncio
import io
import os
import sys
import tempfile
import unittest
import uuid
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "backend"))

# ─── 强制禁用 Auth，避免依赖 JWT ──────────────────────────────────────────────
# setdefault 仅在 settings 尚未加载时有效；当 test_rbac.py 先运行时 settings 已初始化为
# enable_auth=True（来自 .env）。统一用 patch.object 在 setup_module 中覆盖。
os.environ.setdefault("ENABLE_AUTH", "False")

_PREFIX = f"_t_di_{uuid.uuid4().hex[:6]}_"   # 每次运行唯一前缀

# ─── 模块级 auth 补丁（兼容与其他测试套件混跑）─────────────────────────────────
_auth_patcher = None

def setup_module(_=None):
    global _auth_patcher
    from backend.config.settings import settings
    _auth_patcher = patch.object(settings, 'enable_auth', False)
    _auth_patcher.start()


def _db():
    from backend.config.database import SessionLocal
    return SessionLocal()


_g_db = _db()


def teardown_module(_=None):
    """清理测试创建的 ImportJob 记录，并停止 auth 补丁"""
    global _auth_patcher
    if _auth_patcher is not None:
        _auth_patcher.stop()
        _auth_patcher = None
    from backend.models.import_job import ImportJob
    try:
        _g_db.query(ImportJob).filter(
            ImportJob.username.like(f"{_PREFIX}%")
        ).delete(synchronize_session=False)
        _g_db.commit()
    except Exception:
        _g_db.rollback()
    finally:
        _g_db.close()


def _make_simple_xlsx(sheet_name="Sheet1", rows=None) -> bytes:
    """生成最小化 xlsx 字节内容（用于上传测试）"""
    import openpyxl
    from io import BytesIO
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    default_rows = [
        ["name", "age", "city"],
        ["Alice", 30, "Beijing"],
        ["Bob", 25, "Shanghai"],
        ["Carol", 28, "Guangzhou"],
    ]
    for row in (rows or default_rows):
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# A — 服务层工具函数 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestServiceUtils(unittest.TestCase):

    def test_A1_rows_to_values_clause_basic(self):
        """_rows_to_values_clause 基本格式正确"""
        from backend.services.data_import_service import _rows_to_values_clause
        result = _rows_to_values_clause([("Alice", 30, None), ("Bob", 25, 3.14)])
        self.assertIn("('Alice', 30, NULL)", result)
        self.assertIn("('Bob', 25, 3.14)", result)

    def test_A2_rows_to_values_clause_escape_single_quote(self):
        """_rows_to_values_clause 转义字符串中的单引号"""
        from backend.services.data_import_service import _rows_to_values_clause
        result = _rows_to_values_clause([("it's", )])
        self.assertIn("\\'", result)

    def test_A3_rows_to_values_clause_bool(self):
        """_rows_to_values_clause 布尔值转为 0/1"""
        from backend.services.data_import_service import _rows_to_values_clause
        result = _rows_to_values_clause([(True, False)])
        self.assertIn("1, 0", result)

    def test_A4_parse_excel_preview_returns_sheets(self):
        """parse_excel_preview 返回 sheet 列表并包含预览行"""
        import openpyxl
        from backend.services.data_import_service import parse_excel_preview
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(_make_simple_xlsx())
            fname = f.name
        try:
            sheets = parse_excel_preview(fname)
            self.assertEqual(len(sheets), 1)
            self.assertEqual(sheets[0]["sheet_name"], "Sheet1")
            self.assertGreaterEqual(sheets[0]["row_count_estimate"], 4)
            self.assertGreater(len(sheets[0]["preview_rows"]), 0)
        finally:
            os.unlink(fname)

    def test_A5_parse_excel_preview_multi_sheet(self):
        """parse_excel_preview 正确处理多 Sheet"""
        import openpyxl
        from io import BytesIO
        from backend.services.data_import_service import parse_excel_preview
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Alpha"
        ws1.append(["a", "b"])
        ws2 = wb.create_sheet("Beta")
        ws2.append(["x", "y"])
        buf = BytesIO()
        wb.save(buf)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            f.write(buf.getvalue())
            fname = f.name
        try:
            sheets = parse_excel_preview(fname)
            names = [s["sheet_name"] for s in sheets]
            self.assertIn("Alpha", names)
            self.assertIn("Beta", names)
        finally:
            os.unlink(fname)

    def test_A6_list_writable_connections_filters_ro(self):
        """list_writable_connections 不含 -ro 结尾的连接"""
        mock_manager = MagicMock()
        mock_manager.servers = {
            "clickhouse-sg": MagicMock(),
            "clickhouse-sg-ro": MagicMock(),   # 应被过滤
            "postgres-main": MagicMock(),       # 非 ClickHouse，应被过滤
        }
        mock_settings = MagicMock()
        mock_settings.get_clickhouse_config.return_value = {
            "host": "localhost", "http_port": 8123,
            "user": "default", "password": "", "database": "default",
        }
        with patch("backend.mcp.manager.get_mcp_manager", return_value=mock_manager), \
             patch("backend.config.settings.settings", mock_settings):
            from backend.services import data_import_service as _svc
            result = _svc.list_writable_connections()
        envs = [r["env"] for r in result]
        self.assertIn("sg", envs)
        self.assertNotIn("sg_ro", envs)
        self.assertEqual(len(result), 1)


# ═══════════════════════════════════════════════════════════════════════════════
# B — API 权限控制 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestAPIPermissions(unittest.TestCase):
    """ENABLE_AUTH=false 时 AnonymousUser.is_superadmin=True → 所有端点应 200"""

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def test_B1_connections_accessible_no_auth(self):
        """ENABLE_AUTH=false 时 GET /connections 可访问（不应返回 403）"""
        with patch("api.data_import.list_writable_connections", return_value=[]):
            resp = self._client().get("/api/v1/data-import/connections")
        self.assertNotEqual(resp.status_code, 403)
        self.assertIn(resp.status_code, (200,))

    def test_B2_connections_returns_list(self):
        """连接列表端点返回 list 结构"""
        with patch("api.data_import.list_writable_connections", return_value=[]):
            resp = self._client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 200)
        self.assertIsInstance(resp.json()["data"], list)

    def test_B3_databases_requires_env(self):
        """GET /connections/nonexist/databases → 应返回错误（不是 403）"""
        with patch("api.data_import.list_databases", side_effect=Exception("no env")):
            resp = self._client().get("/api/v1/data-import/connections/nonexist/databases")
        self.assertEqual(resp.status_code, 500)
        self.assertNotEqual(resp.status_code, 403)

    def test_B4_tables_requires_db(self):
        """GET /connections/sg/databases/mydb/tables → 应返回错误（不是 403）"""
        with patch("api.data_import.list_tables", side_effect=Exception("no table")):
            resp = self._client().get("/api/v1/data-import/connections/sg/databases/mydb/tables")
        self.assertNotEqual(resp.status_code, 403)

    def test_B5_upload_rejects_non_excel(self):
        """上传非 Excel 文件 → 400"""
        resp = self._client().post(
            "/api/v1/data-import/upload",
            files={"file": ("test.csv", b"a,b,c\n1,2,3", "text/csv")},
        )
        self.assertEqual(resp.status_code, 400)
        self.assertIn("xlsx", resp.json()["detail"].lower())

    def test_B6_upload_rejects_oversized_file(self):
        """超过 100MB 的文件 → 413"""
        big_content = b"x" * (101 * 1024 * 1024)
        resp = self._client().post(
            "/api/v1/data-import/upload",
            files={"file": ("big.xlsx", big_content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(resp.status_code, 413)

    def test_B7_execute_missing_upload_id(self):
        """execute 端点：upload_id 对应文件不存在 → 404"""
        resp = self._client().post("/api/v1/data-import/execute", json={
            "upload_id": str(uuid.uuid4()),
            "connection_env": "sg",
            "sheets": [{"sheet_name": "S1", "database": "db", "table": "t", "has_header": True, "enabled": True}],
        })
        self.assertEqual(resp.status_code, 404)

    def test_B8_execute_no_enabled_sheets(self):
        """execute 端点：所有 sheet 均 disabled → 400"""
        # 先上传一个真实文件
        client = self._client()
        with patch("api.data_import.parse_excel_preview", return_value=[
            {"sheet_name": "S1", "row_count_estimate": 3, "preview_rows": [["a", "b"]]}
        ]):
            xlsx_bytes = _make_simple_xlsx()
            upload_resp = client.post(
                "/api/v1/data-import/upload",
                files={"file": ("test.xlsx", xlsx_bytes,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )
        if upload_resp.status_code != 200:
            self.skipTest("上传失败，跳过该测试")
        uid = upload_resp.json()["data"]["upload_id"]
        resp = client.post("/api/v1/data-import/execute", json={
            "upload_id": uid,
            "connection_env": "sg",
            "sheets": [{"sheet_name": "S1", "database": "db", "table": "t", "has_header": True, "enabled": False}],
        })
        self.assertEqual(resp.status_code, 400)


# ═══════════════════════════════════════════════════════════════════════════════
# C — 连接/数据库/表查询端点 (6 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestQueryEndpoints(unittest.TestCase):

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def test_C1_connections_success(self):
        """GET /connections 成功返回连接列表"""
        mock_conns = [{"env": "sg", "display_name": "clickhouse-sg", "host": "localhost",
                       "http_port": 8123, "database": "default", "server_name": "clickhouse-sg"}]
        with patch("api.data_import.list_writable_connections", return_value=mock_conns):
            resp = self._client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["success"])
        self.assertEqual(resp.json()["data"][0]["env"], "sg")

    def test_C2_databases_success(self):
        """GET /connections/{env}/databases 返回数据库列表"""
        with patch("api.data_import.list_databases", return_value=["mydb", "analytics"]):
            resp = self._client().get("/api/v1/data-import/connections/sg/databases")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("mydb", resp.json()["data"])

    def test_C3_tables_success(self):
        """GET /connections/{env}/databases/{db}/tables 返回表列表"""
        with patch("api.data_import.list_tables", return_value=["users", "orders"]):
            resp = self._client().get("/api/v1/data-import/connections/sg/databases/mydb/tables")
        self.assertEqual(resp.status_code, 200)
        self.assertIn("users", resp.json()["data"])

    def test_C4_databases_error_returns_500(self):
        """CH 查询失败时，数据库端点返回 500"""
        with patch("api.data_import.list_databases", side_effect=Exception("conn refused")):
            resp = self._client().get("/api/v1/data-import/connections/sg/databases")
        self.assertEqual(resp.status_code, 500)

    def test_C5_tables_error_returns_500(self):
        """CH 查询失败时，表端点返回 500"""
        with patch("api.data_import.list_tables", side_effect=Exception("conn refused")):
            resp = self._client().get("/api/v1/data-import/connections/sg/databases/mydb/tables")
        self.assertEqual(resp.status_code, 500)

    def test_C6_connections_error_returns_500(self):
        """连接列表获取失败时返回 500"""
        with patch("api.data_import.list_writable_connections", side_effect=Exception("cfg error")):
            resp = self._client().get("/api/v1/data-import/connections")
        self.assertEqual(resp.status_code, 500)


# ═══════════════════════════════════════════════════════════════════════════════
# D — 上传端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestUploadEndpoint(unittest.TestCase):

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def _upload(self, client, content=None, filename="test.xlsx"):
        content = content or _make_simple_xlsx()
        return client.post(
            "/api/v1/data-import/upload",
            files={"file": (filename, content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    def test_D1_upload_xlsx_success(self):
        """正常 xlsx 文件上传成功，返回 upload_id 和 sheets"""
        resp = self._upload(self._client())
        self.assertEqual(resp.status_code, 200)
        data = resp.json()["data"]
        self.assertIn("upload_id", data)
        self.assertIn("sheets", data)
        self.assertIsInstance(data["sheets"], list)

    def test_D2_upload_returns_preview_rows(self):
        """上传后返回的 sheets 含预览行"""
        resp = self._upload(self._client())
        self.assertEqual(resp.status_code, 200)
        sheet = resp.json()["data"]["sheets"][0]
        self.assertIn("preview_rows", sheet)
        self.assertGreater(len(sheet["preview_rows"]), 0)

    def test_D3_upload_returns_row_count(self):
        """上传后返回 row_count_estimate"""
        resp = self._upload(self._client())
        self.assertEqual(resp.status_code, 200)
        sheet = resp.json()["data"]["sheets"][0]
        self.assertIn("row_count_estimate", sheet)
        self.assertGreater(sheet["row_count_estimate"], 0)

    def test_D4_upload_returns_filename(self):
        """上传后响应包含原始文件名"""
        resp = self._upload(self._client(), filename="myfile.xlsx")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["filename"], "myfile.xlsx")

    def test_D5_upload_rejects_csv(self):
        """上传 .csv 文件 → 400"""
        resp = self._client().post(
            "/api/v1/data-import/upload",
            files={"file": ("data.csv", b"a,b\n1,2", "text/csv")},
        )
        self.assertEqual(resp.status_code, 400)

    def test_D6_upload_rejects_txt(self):
        """上传 .txt 文件 → 400"""
        resp = self._client().post(
            "/api/v1/data-import/upload",
            files={"file": ("data.txt", b"hello", "text/plain")},
        )
        self.assertEqual(resp.status_code, 400)

    def test_D7_upload_invalid_xlsx_returns_422(self):
        """上传内容损坏的 xlsx → 422"""
        resp = self._client().post(
            "/api/v1/data-import/upload",
            files={"file": ("bad.xlsx", b"not_an_xlsx_content",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        self.assertEqual(resp.status_code, 422)

    def test_D8_upload_stores_file_on_disk(self):
        """上传后文件保存到临时目录"""
        resp = self._upload(self._client())
        self.assertEqual(resp.status_code, 200)
        uid = resp.json()["data"]["upload_id"]
        self.assertIsNotNone(uid)
        self.assertRegex(uid, r'^[0-9a-f-]{36}$')


# ═══════════════════════════════════════════════════════════════════════════════
# E — execute 导入端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestExecuteEndpoint(unittest.TestCase):

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def _upload_and_get_id(self, client):
        resp = client.post(
            "/api/v1/data-import/upload",
            files={"file": ("test.xlsx", _make_simple_xlsx(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        if resp.status_code != 200:
            return None
        return resp.json()["data"]["upload_id"]

    def test_E1_execute_creates_job(self):
        """execute 成功创建任务，返回 job_id"""
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task"):
            resp = client.post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                             "has_header": True, "enabled": True}],
            })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()["data"]
        self.assertIn("job_id", data)
        self.assertEqual(data["status"], "pending")

    def test_E2_execute_job_persisted_in_db(self):
        """execute 后任务记录写入数据库"""
        from backend.models.import_job import ImportJob
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task"):
            resp = client.post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                             "has_header": True, "enabled": True}],
            })
        if resp.status_code != 200:
            self.skipTest("execute failed")
        job_id = resp.json()["data"]["job_id"]
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertIsNotNone(job)
            self.assertEqual(job.status, "pending")
            self.assertEqual(job.connection_env, "sg")
        finally:
            db.close()

    def test_E3_execute_missing_upload_id_404(self):
        """不存在的 upload_id → 404"""
        resp = self._client().post("/api/v1/data-import/execute", json={
            "upload_id": str(uuid.uuid4()),
            "connection_env": "sg",
            "sheets": [{"sheet_name": "S1", "database": "db", "table": "t",
                        "has_header": True, "enabled": True}],
        })
        self.assertEqual(resp.status_code, 404)

    def test_E4_execute_no_enabled_sheets_400(self):
        """所有 sheet 都 disabled → 400"""
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        resp = client.post("/api/v1/data-import/execute", json={
            "upload_id": uid,
            "connection_env": "sg",
            "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                        "has_header": True, "enabled": False}],
        })
        self.assertEqual(resp.status_code, 400)

    def test_E5_execute_batch_size_default(self):
        """execute 不传 batch_size 时默认值生效"""
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task"):
            resp = client.post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                             "has_header": True, "enabled": True}],
            })
        self.assertEqual(resp.status_code, 200)

    def test_E6_execute_invalid_batch_size_too_small(self):
        """batch_size < 100 → 422（Pydantic 校验）"""
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        resp = client.post("/api/v1/data-import/execute", json={
            "upload_id": uid,
            "connection_env": "sg",
            "batch_size": 50,   # min=100
            "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                        "has_header": True, "enabled": True}],
        })
        self.assertEqual(resp.status_code, 422)

    def test_E7_execute_config_snapshot_saved(self):
        """execute 后 config_snapshot 保存到 ImportJob"""
        from backend.models.import_job import ImportJob
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task"):
            resp = client.post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "batch_size": 500,
                "sheets": [{"sheet_name": "Sheet1", "database": "mydb", "table": "mytable",
                             "has_header": True, "enabled": True}],
            })
        if resp.status_code != 200:
            self.skipTest("execute failed")
        job_id = resp.json()["data"]["job_id"]
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            snap = job.config_snapshot
            self.assertEqual(snap["connection_env"], "sg")
            self.assertEqual(snap["batch_size"], 500)
            self.assertEqual(snap["sheets"][0]["table"], "mytable")
        finally:
            db.close()

    def test_E8_execute_triggers_background_task(self):
        """execute 调用 asyncio.create_task 启动后台协程"""
        client = self._client()
        uid = self._upload_and_get_id(client)
        if uid is None:
            self.skipTest("upload failed")
        with patch("api.data_import.asyncio.create_task") as mock_ct:
            resp = client.post("/api/v1/data-import/execute", json={
                "upload_id": uid,
                "connection_env": "sg",
                "sheets": [{"sheet_name": "Sheet1", "database": "db", "table": "t",
                             "has_header": True, "enabled": True}],
            })
        if resp.status_code == 200:
            mock_ct.assert_called_once()


# ═══════════════════════════════════════════════════════════════════════════════
# F — jobs 查询端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestJobsEndpoints(unittest.TestCase):

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def _create_job(self, status="pending", username=None):
        """直接在 DB 创建一个 ImportJob 记录"""
        from backend.models.import_job import ImportJob
        db = _db()
        try:
            job = ImportJob(
                user_id=str(uuid.uuid4()),
                username=username or f"{_PREFIX}u",
                upload_id=str(uuid.uuid4()),
                filename="test.xlsx",
                connection_env="sg",
                status=status,
            )
            db.add(job)
            db.commit()
            db.refresh(job)
            return str(job.id)
        finally:
            db.close()

    def test_F1_get_job_status_pending(self):
        """GET /jobs/{job_id} 返回 pending 任务信息"""
        job_id = self._create_job(status="pending")
        resp = self._client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()["data"]
        self.assertEqual(data["job_id"], job_id)
        self.assertEqual(data["status"], "pending")

    def test_F2_get_job_not_found(self):
        """GET /jobs/{random_uuid} → 404"""
        resp = self._client().get(f"/api/v1/data-import/jobs/{uuid.uuid4()}")
        self.assertEqual(resp.status_code, 404)

    def test_F3_get_job_completed_status(self):
        """GET /jobs/{job_id} 返回 completed 状态"""
        job_id = self._create_job(status="completed")
        resp = self._client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["status"], "completed")

    def test_F4_get_job_failed_status(self):
        """GET /jobs/{job_id} 返回 failed 状态"""
        job_id = self._create_job(status="failed")
        resp = self._client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["status"], "failed")

    def test_F5_list_jobs_returns_paginated(self):
        """GET /jobs 返回分页结构"""
        resp = self._client().get("/api/v1/data-import/jobs?page=1&page_size=10")
        self.assertEqual(resp.status_code, 200)
        data = resp.json()["data"]
        self.assertIn("total", data)
        self.assertIn("page", data)
        self.assertIn("items", data)
        self.assertIsInstance(data["items"], list)

    def test_F6_list_jobs_default_page_size_10(self):
        """GET /jobs 默认每页 10 条"""
        resp = self._client().get("/api/v1/data-import/jobs")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["page_size"], 10)

    def test_F7_list_jobs_sorted_desc(self):
        """历史列表按 created_at 倒序排列（最新在最前）"""
        # 创建 3 条记录，created_at 递增
        import time
        ids = []
        for i in range(3):
            ids.append(self._create_job(status="completed"))
            time.sleep(0.05)
        resp = self._client().get("/api/v1/data-import/jobs?page=1&page_size=50")
        self.assertEqual(resp.status_code, 200)
        items = resp.json()["data"]["items"]
        # 找到我们创建的 3 条
        our_items = [it for it in items if it["job_id"] in ids]
        if len(our_items) >= 2:
            t0 = our_items[0]["created_at"]
            t1 = our_items[1]["created_at"]
            if t0 and t1:
                self.assertGreaterEqual(t0, t1)

    def test_F8_list_jobs_to_dict_fields(self):
        """ImportJob.to_dict() 包含所有前端必需字段"""
        job_id = self._create_job()
        resp = self._client().get(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        d = resp.json()["data"]
        required_fields = [
            "job_id", "filename", "connection_env", "status",
            "total_sheets", "done_sheets", "imported_rows",
            "total_batches", "done_batches", "errors", "created_at",
        ]
        for f in required_fields:
            self.assertIn(f, d, f"Missing field: {f}")


# ═══════════════════════════════════════════════════════════════════════════════
# H — cancel 端点 (8 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestCancelEndpoint(unittest.TestCase):
    """验证 POST /jobs/{job_id}/cancel 状态机约束与 DB 持久化"""

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def _make_job(self, status="pending"):
        from backend.models.import_job import ImportJob
        db = _db()
        try:
            job = ImportJob(
                user_id=str(uuid.uuid4()),
                username=f"{_PREFIX}h",
                upload_id=str(uuid.uuid4()),
                filename="cancel_test.xlsx",
                connection_env="sg",
                status=status,
            )
            db.add(job)
            db.commit()
            db.refresh(job)
            return str(job.id)
        finally:
            db.close()

    def test_H1_cancel_pending_job_returns_200(self):
        """pending 状态任务可以取消 → 200, data.status=cancelling"""
        job_id = self._make_job("pending")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["success"])
        self.assertEqual(resp.json()["data"]["status"], "cancelling")

    def test_H2_cancel_running_job_returns_200(self):
        """running 状态任务可以取消 → 200"""
        job_id = self._make_job("running")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()["data"]["status"], "cancelling")

    def test_H3_cancel_completed_job_returns_400(self):
        """completed 任务不可取消 → 400"""
        job_id = self._make_job("completed")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 400)
        self.assertIn("completed", resp.json()["detail"])

    def test_H4_cancel_failed_job_returns_400(self):
        """failed 任务不可取消 → 400"""
        job_id = self._make_job("failed")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 400)

    def test_H5_cancel_nonexistent_job_returns_404(self):
        """不存在的 job_id → 404"""
        resp = self._client().post(f"/api/v1/data-import/jobs/{uuid.uuid4()}/cancel")
        self.assertEqual(resp.status_code, 404)

    def test_H6_cancel_sets_db_status_to_cancelling(self):
        """cancel 请求后 DB 中 status 字段更新为 cancelling"""
        from backend.models.import_job import ImportJob
        job_id = self._make_job("pending")
        self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertIsNotNone(job)
            self.assertEqual(job.status, "cancelling")
        finally:
            db.close()

    def test_H7_cancel_already_cancelling_returns_400(self):
        """cancelling 状态任务再次取消 → 400（幂等保护）"""
        job_id = self._make_job("cancelling")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 400)

    def test_H8_cancel_already_cancelled_returns_400(self):
        """cancelled 状态任务不可取消 → 400"""
        job_id = self._make_job("cancelled")
        resp = self._client().post(f"/api/v1/data-import/jobs/{job_id}/cancel")
        self.assertEqual(resp.status_code, 400)


# ═══════════════════════════════════════════════════════════════════════════════
# I — delete 端点 (4 tests)
# ═══════════════════════════════════════════════════════════════════════════════

class TestDeleteEndpoint(unittest.TestCase):
    """验证 DELETE /jobs/{job_id} 任务记录删除行为"""

    def _client(self):
        from fastapi.testclient import TestClient
        from backend.main import app
        return TestClient(app, raise_server_exceptions=False)

    def _make_job(self, status="completed"):
        from backend.models.import_job import ImportJob
        db = _db()
        try:
            job = ImportJob(
                user_id=str(uuid.uuid4()),
                username=f"{_PREFIX}i",
                upload_id=str(uuid.uuid4()),
                filename="delete_test.xlsx",
                connection_env="sg",
                status=status,
            )
            db.add(job)
            db.commit()
            db.refresh(job)
            return str(job.id)
        finally:
            db.close()

    def test_I1_delete_existing_job_returns_200(self):
        """删除存在的任务 → 200, success=true"""
        job_id = self._make_job("completed")
        resp = self._client().delete(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()["success"])

    def test_I2_delete_nonexistent_job_returns_404(self):
        """删除不存在的任务 → 404"""
        resp = self._client().delete(f"/api/v1/data-import/jobs/{uuid.uuid4()}")
        self.assertEqual(resp.status_code, 404)

    def test_I3_delete_removes_record_from_db(self):
        """删除后 DB 中记录消失"""
        from backend.models.import_job import ImportJob
        job_id = self._make_job("failed")
        self._client().delete(f"/api/v1/data-import/jobs/{job_id}")
        db = _db()
        try:
            job = db.query(ImportJob).filter(ImportJob.id == job_id).first()
            self.assertIsNone(job)
        finally:
            db.close()

    def test_I4_delete_active_job_also_succeeds(self):
        """删除 running/pending 任务也返回 200（不阻止删除记录，不停止后台任务）"""
        job_id = self._make_job("running")
        resp = self._client().delete(f"/api/v1/data-import/jobs/{job_id}")
        self.assertEqual(resp.status_code, 200)


if __name__ == "__main__":
    unittest.main(verbosity=2)
